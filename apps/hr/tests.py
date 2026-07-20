from io import BytesIO

import openpyxl
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from datetime import date

from .models import Candidate


User = get_user_model()


class HRCandidateImportTests(TestCase):
    def setUp(self):
        self.hr_user = User.objects.create_user(username='hr_test', password='password', role='hr')
        self.client = Client()
        self.client.login(username='hr_test', password='password')

    def make_workbook_file(self, row_values):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([])
        sheet.append([])
        sheet.append([
            'Sr.No',
            'Interview Date',
            'Candidate Name',
            'Position',
            'Contact No',
            'Year of Experience',
            'Mail Id',
            'Current ctc',
            'Expect ctc',
            'Notice Period',
            'Location',
            'Interview Status',
            'Remark',
            'Source',
        ])
        sheet.append(row_values)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            'candidates.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_candidate_excel_import_creates_sheet_fields(self):
        upload = self.make_workbook_file([
            1,
            '15-07-2026',
            'Riya Sharma',
            'Backend Developer',
            '9876501234',
            '3 Years',
            'riya@example.com',
            '5 LPA',
            '7 LPA',
            '30 Days',
            'Mumbai',
            'Interview Scheduled',
            'Strong Django profile',
            'LinkedIn',
        ])

        response = self.client.post(reverse('hr:candidate_import'), {'candidates_file': upload})

        self.assertRedirects(response, reverse('hr:candidate_list'))
        candidate = Candidate.objects.get(mobile='9876501234')
        self.assertEqual(candidate.full_name, 'Riya Sharma')
        self.assertEqual(candidate.applying_position, 'Backend Developer')
        self.assertEqual(candidate.experience, '3 Years')
        self.assertEqual(candidate.email, 'riya@example.com')
        self.assertEqual(candidate.current_salary, '5 LPA')
        self.assertEqual(candidate.expected_salary, '7 LPA')
        self.assertEqual(candidate.notice_period, '30 Days')
        self.assertEqual(candidate.location, 'Mumbai')
        self.assertEqual(candidate.interview_date, date(2026, 7, 15))
        self.assertEqual(candidate.status, 'interview_scheduled')
        self.assertEqual(candidate.remarks, 'Strong Django profile')
        self.assertEqual(candidate.source, 'LinkedIn')
        self.assertEqual(candidate.assigned_hr, self.hr_user)
        self.assertTrue(candidate.activities.filter(title='Candidate Imported').exists())

    def test_candidate_import_updates_existing_contact(self):
        Candidate.objects.create(
            full_name='Old Name',
            mobile='9876501234',
            applying_position='Old Role',
            assigned_hr=self.hr_user,
            created_by=self.hr_user,
        )
        upload = self.make_workbook_file([
            1,
            '15-07-2026',
            'Updated Name',
            'Frontend Developer',
            '9876501234',
            '2 Years',
            'updated@example.com',
            '4 LPA',
            '6 LPA',
            'Immediate',
            'Pune',
            'Selected',
            'Ready to join',
            'Naukri',
        ])

        response = self.client.post(reverse('hr:candidate_import'), {'candidates_file': upload})

        self.assertRedirects(response, reverse('hr:candidate_list'))
        self.assertEqual(Candidate.objects.filter(mobile='9876501234').count(), 1)
        candidate = Candidate.objects.get(mobile='9876501234')
        self.assertEqual(candidate.full_name, 'Updated Name')
        self.assertEqual(candidate.applying_position, 'Frontend Developer')
        self.assertEqual(candidate.status, 'selected')
        self.assertEqual(candidate.notice_period, 'Immediate')
        self.assertTrue(candidate.activities.filter(title='Candidate Import Updated').exists())

    def test_candidate_import_uses_selected_added_date(self):
        upload = self.make_workbook_file([
            1,
            '15-07-2026',
            'Imported Date Candidate',
            'Counsellor',
            '9000005555',
            '1 Year',
            'added-date@example.com',
            '',
            '',
            '',
            'Mumbai',
            'Applied',
            '',
            'Walk In',
        ])

        response = self.client.post(
            reverse('hr:candidate_import'),
            {'date_added': '2026-05-04', 'candidates_file': upload},
        )

        self.assertRedirects(response, reverse('hr:candidate_list'))
        candidate = Candidate.objects.get(mobile='9000005555')
        self.assertEqual(candidate.date_added, date(2026, 5, 4))

    def test_candidate_import_defaults_added_date_to_today(self):
        upload = self.make_workbook_file([
            1,
            '15-07-2026',
            'Default Date Candidate',
            'Recruiter',
            '9000006666',
            '2 Years',
            'default-date@example.com',
            '',
            '',
            '',
            'Pune',
            'Applied',
            '',
            'Referral',
        ])

        response = self.client.post(reverse('hr:candidate_import'), {'candidates_file': upload})

        self.assertRedirects(response, reverse('hr:candidate_list'))
        candidate = Candidate.objects.get(mobile='9000006666')
        self.assertEqual(candidate.date_added, timezone.localdate())

    def test_candidate_import_handles_exported_date_and_header_variants(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([])
        sheet.append([
            'Interview Date',
            'Candidate Name',
            'Position',
            'Contact No',
            'Year of Expereince',
            'Mail Id',
            'Source',
        ])
        sheet.append([
            46218,
            'Neha Patel',
            'HR Executive',
            '9000001111',
            '4.5 Years',
            'neha@example.com',
            'Consultant Sheet',
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            'candidates.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(reverse('hr:candidate_import'), {'candidates_file': upload})

        self.assertRedirects(response, reverse('hr:candidate_list'))
        candidate = Candidate.objects.get(mobile='9000001111')
        self.assertEqual(candidate.interview_date, date(2026, 7, 15))
        self.assertEqual(candidate.experience, '4.5 Years')
        self.assertEqual(candidate.source, 'Consultant Sheet')

    def test_candidate_import_keeps_date_and_experience_with_typos_and_blank_duplicates(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([
            'Candidate Name',
            'Contact No',
            'Interview Date',
            'Year of Experenice',
            'Year of Experience',
            'Position',
        ])
        sheet.append([
            'Asha Rao',
            '9000002222',
            '15-07-2026 00:00:00',
            '6 Years',
            '',
            'Recruiter',
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            'candidates.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(reverse('hr:candidate_import'), {'candidates_file': upload})

        self.assertRedirects(response, reverse('hr:candidate_list'))
        candidate = Candidate.objects.get(mobile='9000002222')
        self.assertEqual(candidate.interview_date, date(2026, 7, 15))
        self.assertEqual(candidate.experience, '6 Years')

    def test_candidate_import_uses_standard_date_column_when_header_is_messy(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([
            'Sr.No',
            'Interview Dt.',
            'Candidate Name',
            'Position',
            'Contact No',
            'Yr Exp',
        ])
        sheet.append([
            1,
            'Interview fixed for Jul 15, 2026',
            'Kiran Sen',
            'Accountant',
            '9000003333',
            '7 Years',
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            'candidates.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(reverse('hr:candidate_import'), {'candidates_file': upload})

        self.assertRedirects(response, reverse('hr:candidate_list'))
        candidate = Candidate.objects.get(mobile='9000003333')
        self.assertEqual(candidate.interview_date, date(2026, 7, 15))
        self.assertEqual(candidate.experience, '7 Years')

    def test_candidate_import_accepts_day_month_interview_date_without_year(self):
        upload = self.make_workbook_file([
            1,
            '4th May',
            'Maya Iyer',
            'HR Recruiter',
            '9000004444',
            '5 Years',
            'maya@example.com',
            '',
            '',
            '',
            'Delhi',
            'Interview Scheduled',
            '',
            'Referral',
        ])

        response = self.client.post(reverse('hr:candidate_import'), {'candidates_file': upload})

        self.assertRedirects(response, reverse('hr:candidate_list'))
        candidate = Candidate.objects.get(mobile='9000004444')
        self.assertEqual(candidate.interview_date, date(2026, 5, 4))
        self.assertEqual(candidate.experience, '5 Years')

    def test_candidate_edit_updates_added_date(self):
        candidate = Candidate.objects.create(
            full_name='Editable Date Candidate',
            mobile='9000007777',
            email='edit-date@example.com',
            applying_position='Counsellor',
            status='new',
            source='Indeed',
            assigned_hr=self.hr_user,
            created_by=self.hr_user,
            date_added=date(2026, 7, 15),
        )

        response = self.client.post(
            reverse('hr:candidate_edit', args=[candidate.id]),
            {
                'full_name': 'Editable Date Candidate',
                'mobile': '9000007777',
                'email': 'edit-date@example.com',
                'location': 'Mumbai',
                'experience': '2 Years',
                'current_salary': '3 LPA',
                'expected_salary': '5 LPA',
                'notice_period': 'Immediate',
                'applying_position': 'Counsellor',
                'date_added': '2026-05-04',
                'interview_date': '',
                'source': 'Indeed',
                'assigned_hr': self.hr_user.id,
                'status': 'selected',
                'remarks': 'Updated added date',
            },
        )

        self.assertRedirects(response, reverse('hr:candidate_detail', args=[candidate.id]))
        candidate.refresh_from_db()
        self.assertEqual(candidate.date_added, date(2026, 5, 4))
