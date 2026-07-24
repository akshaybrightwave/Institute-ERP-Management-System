import calendar
from io import BytesIO
from functools import wraps
from urllib.parse import urlencode
from xml.sax.saxutils import escape

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.accounts.models import User

from .forms import (
    CaseWorkItemFormSet,
    CaseWorkReportForm,
    FraudTypeForm,
    InvestigatorUserCreateForm,
    InvestigatorUserEditForm,
    PoliceStationForm,
    RecoveryReportForm,
)
from .models import CaseWorkReport, FraudType, PoliceStation, RecoveryReport


def investigator_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and request.user.role in ('admin', 'investigator'):
            return view_func(request, *args, **kwargs)
        return HttpResponseForbidden("Access Denied: Investigator Panel only.")
    return wrapper


def investigator_admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and request.user.role == 'admin':
            return view_func(request, *args, **kwargs)
        return HttpResponseForbidden("Access Denied: Admins only.")
    return wrapper


def render_panel_page(request, title, description, phase_note, admin_only=False, primary_action=None):
    return render(request, 'investigator/panel_page.html', {
        'title': title,
        'description': description,
        'phase_note': phase_note,
        'admin_only': admin_only,
        'primary_action': primary_action,
    })


def user_can_edit_recovery_report(user, report):
    if user.role == 'admin':
        return True
    return report.submitted_by_id == user.id and report.can_investigator_edit


def user_can_edit_case_work_report(user, report):
    if user.role == 'admin':
        return True
    return report.submitted_by_id == user.id and report.can_investigator_edit


def get_user_display_name(user):
    if not user:
        return ''
    return user.get_full_name() or user.username


def get_report_investigator_name(report, request_user):
    if report and report.name:
        return report.name
    if report and report.submitted_by:
        return get_user_display_name(report.submitted_by)
    return get_user_display_name(request_user)


def build_universal_report_data(request):
    filters = {
        'month': request.GET.get('month', '').strip(),
        'year': request.GET.get('year', '').strip(),
        'investigator': request.GET.get('investigator', '').strip(),
        'investigator_name': request.GET.get('investigator_name', '').strip(),
        'police_station': request.GET.get('police_station', '').strip(),
        'fraud_type': request.GET.get('fraud_type', '').strip(),
        'case_status': request.GET.get('case_status', '').strip(),
        'approval_status': request.GET.get('approval_status', RecoveryReport.STATUS_APPROVED).strip(),
    }
    if 'approval_status' not in request.GET:
        filters['approval_status'] = RecoveryReport.STATUS_APPROVED

    recovery_reports = RecoveryReport.objects.select_related(
        'submitted_by',
        'police_station',
        'fraud_type',
    ).filter(is_archived=False)
    case_reports = CaseWorkReport.objects.select_related(
        'submitted_by',
        'police_station',
    ).filter(is_archived=False)

    if filters['month'].isdigit():
        recovery_reports = recovery_reports.filter(report_month=filters['month'])
        case_reports = case_reports.filter(entry_date__month=filters['month'])
    if filters['year'].isdigit():
        recovery_reports = recovery_reports.filter(report_year=filters['year'])
        case_reports = case_reports.filter(entry_date__year=filters['year'])
    if filters['investigator'].isdigit():
        recovery_reports = recovery_reports.filter(submitted_by_id=filters['investigator'])
        case_reports = case_reports.filter(submitted_by_id=filters['investigator'])
    if filters['investigator_name']:
        recovery_reports = recovery_reports.filter(
            Q(submitted_by__first_name__icontains=filters['investigator_name']) |
            Q(submitted_by__last_name__icontains=filters['investigator_name']) |
            Q(submitted_by__username__icontains=filters['investigator_name'])
        )
        case_reports = case_reports.filter(
            Q(submitted_by__first_name__icontains=filters['investigator_name']) |
            Q(submitted_by__last_name__icontains=filters['investigator_name']) |
            Q(submitted_by__username__icontains=filters['investigator_name'])
        )
    if filters['police_station'].isdigit():
        recovery_reports = recovery_reports.filter(police_station_id=filters['police_station'])
        case_reports = case_reports.filter(police_station_id=filters['police_station'])
    if filters['fraud_type'].isdigit():
        recovery_reports = recovery_reports.filter(fraud_type_id=filters['fraud_type'])
        case_reports = case_reports.none()
    if filters['case_status']:
        recovery_reports = recovery_reports.none()
        case_reports = case_reports.filter(case_status=filters['case_status'])
    if filters['approval_status']:
        recovery_reports = recovery_reports.filter(approval_status=filters['approval_status'])
        case_reports = case_reports.filter(approval_status=filters['approval_status'])

    recovery_reports = recovery_reports.order_by('police_station__name', 'name', 'report_number')
    case_reports = case_reports.order_by('-created_at')
    recovery_totals = recovery_reports.aggregate(
        mobile=Sum('mobile_recovery_count'),
        financial=Sum('financial_recovery_amount'),
    )
    totals = {
        'recovery_count': recovery_reports.count(),
        'case_count': case_reports.count(),
        'mobile': recovery_totals['mobile'] or 0,
        'financial': recovery_totals['financial'] or 0,
        'investigators': len(set(
            list(recovery_reports.exclude(submitted_by__isnull=True).values_list('submitted_by_id', flat=True)) +
            list(case_reports.exclude(submitted_by__isnull=True).values_list('submitted_by_id', flat=True))
        )),
    }
    return filters, recovery_reports, case_reports, totals


def get_recovery_report_queryset(request):
    reports = RecoveryReport.objects.select_related(
        'police_station',
        'fraud_type',
        'submitted_by',
    ).filter(is_archived=False)
    if request.user.role == 'investigator':
        reports = reports.filter(submitted_by=request.user)
    return reports


def get_case_work_report_queryset(request):
    reports = CaseWorkReport.objects.select_related(
        'police_station',
        'submitted_by',
    ).prefetch_related('work_items').filter(is_archived=False)
    if request.user.role == 'investigator':
        reports = reports.filter(submitted_by=request.user)
    return reports


def case_work_formset_item_count(formset):
    count = 0
    for form in formset.forms:
        if not form.cleaned_data or form.cleaned_data.get('DELETE'):
            continue
        title = (form.cleaned_data.get('work_title') or '').strip()
        description = (form.cleaned_data.get('work_description') or '').strip()
        if title and description:
            count += 1
    return count


def apply_recovery_report_filters(request, reports):
    filters = {
        'name': request.GET.get('name', request.GET.get('student_name', '')).strip(),
        'investigator_name': request.GET.get('investigator_name', '').strip(),
        'police_station': request.GET.get('police_station', '').strip(),
        'fraud_type': request.GET.get('fraud_type', '').strip(),
        'month': request.GET.get('month', '').strip(),
        'year': request.GET.get('year', '').strip(),
        'approval_status': request.GET.get('approval_status', '').strip(),
    }

    if filters['name']:
        reports = reports.filter(name__icontains=filters['name'])
    if request.user.role == 'admin' and filters['investigator_name']:
        reports = reports.filter(
            Q(submitted_by__first_name__icontains=filters['investigator_name']) |
            Q(submitted_by__last_name__icontains=filters['investigator_name']) |
            Q(submitted_by__username__icontains=filters['investigator_name'])
        )
    if filters['police_station'].isdigit():
        reports = reports.filter(police_station_id=filters['police_station'])
    if filters['fraud_type'].isdigit():
        reports = reports.filter(fraud_type_id=filters['fraud_type'])
    if filters['month'].isdigit():
        reports = reports.filter(report_month=filters['month'])
    if filters['year'].isdigit():
        reports = reports.filter(report_year=filters['year'])
    if filters['approval_status']:
        reports = reports.filter(approval_status=filters['approval_status'])

    return reports, filters


def apply_case_work_report_filters(request, reports):
    filters = {
        'name': request.GET.get('name', '').strip(),
        'case_no': request.GET.get('case_no', '').strip(),
        'investigator_name': request.GET.get('investigator_name', '').strip(),
        'police_station': request.GET.get('police_station', '').strip(),
        'investigating_officer': request.GET.get('investigating_officer', '').strip(),
        'case_status': request.GET.get('case_status', '').strip(),
        'approval_status': request.GET.get('approval_status', '').strip(),
        'month': request.GET.get('month', '').strip(),
        'year': request.GET.get('year', '').strip(),
    }

    if filters['name']:
        reports = reports.filter(name__icontains=filters['name'])
    if filters['case_no']:
        reports = reports.filter(case_no__icontains=filters['case_no'])
    if request.user.role == 'admin' and filters['investigator_name']:
        reports = reports.filter(
            Q(submitted_by__first_name__icontains=filters['investigator_name']) |
            Q(submitted_by__last_name__icontains=filters['investigator_name']) |
            Q(submitted_by__username__icontains=filters['investigator_name'])
        )
    if filters['police_station'].isdigit():
        reports = reports.filter(police_station_id=filters['police_station'])
    if filters['investigating_officer']:
        reports = reports.filter(investigating_officer__icontains=filters['investigating_officer'])
    if filters['case_status']:
        reports = reports.filter(case_status=filters['case_status'])
    if filters['approval_status']:
        reports = reports.filter(approval_status=filters['approval_status'])
    if filters['month'].isdigit():
        reports = reports.filter(entry_date__month=filters['month'])
    if filters['year'].isdigit():
        reports = reports.filter(entry_date__year=filters['year'])

    return reports, filters


def apply_dashboard_filters(request, recovery_reports, case_reports):
    filters = {
        'month': request.GET.get('month', '').strip(),
        'year': request.GET.get('year', '').strip(),
        'approval_status': request.GET.get('approval_status', '').strip(),
    }

    if filters['month'].isdigit():
        recovery_reports = recovery_reports.filter(report_month=filters['month'])
        case_reports = case_reports.filter(entry_date__month=filters['month'])
    if filters['year'].isdigit():
        recovery_reports = recovery_reports.filter(report_year=filters['year'])
        case_reports = case_reports.filter(entry_date__year=filters['year'])
    if filters['approval_status']:
        recovery_reports = recovery_reports.filter(approval_status=filters['approval_status'])
        case_reports = case_reports.filter(approval_status=filters['approval_status'])

    return recovery_reports, case_reports, filters


def combined_status_count(recovery_reports, case_reports, status):
    return (
        recovery_reports.filter(approval_status=status).count() +
        case_reports.filter(approval_status=status).count()
    )


def get_print_period_label(filters):
    month = filters.get('month', '')
    year = filters.get('year', '')
    if month.isdigit() and 1 <= int(month) <= 12 and year:
        return f'{calendar.month_name[int(month)].upper()} - {year}'
    if year:
        return str(year)
    return timezone.localdate().strftime('%B - %Y').upper()


def get_export_period_label(filters):
    return get_print_period_label(filters)


def format_submitted_at(value):
    if not value:
        return ''
    return timezone.localtime(value).strftime('%d %b %Y, %I:%M %p')


def format_mobile_count(value):
    count = value or 0
    if not count:
        return '-'
    label = 'Mobile' if count == 1 else 'Mobiles'
    return f'{count} {label}'


def format_financial_amount(value):
    amount = value or 0
    if not amount:
        return '-'
    return f'Rs. {amount:.2f} /-'


def pdf_text(value):
    return escape(str(value or ''))


def make_excel_response(workbook, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def style_excel_sheet(worksheet):
    thin = Side(style='thin', color='9aa7c7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in worksheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EEF2FF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for index, width in enumerate([8, 28, 34, 24, 18, 22, 18, 20], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def add_recovery_excel_sheet(workbook, title, reports, filters):
    worksheet = workbook.create_sheet(title=title[:31])
    period = get_export_period_label(filters)
    worksheet.append([f'ZONE - 1    {period}'])
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    worksheet.append([
        'Sr.N',
        'Name',
        'Name of the Police Station',
        'Types of Fraud',
        'Mobile Recovery',
        'Financial Recovery',
    ])
    total_mobile = 0
    total_financial = 0
    for index, report in enumerate(reports, start=1):
        total_mobile += report.mobile_recovery_count or 0
        total_financial += report.financial_recovery_amount or 0
        worksheet.append([
            index,
            report.name,
            report.police_station.name,
            report.fraud_type.name,
            format_mobile_count(report.mobile_recovery_count),
            format_financial_amount(report.financial_recovery_amount),
        ])

    worksheet.append([])
    worksheet.append(['Total', '', '', '', '', ''])
    worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=2)
    worksheet.append(['Total Mobile Recovery', format_mobile_count(total_mobile)])
    worksheet.append(['Total Financial Recovery', format_financial_amount(total_financial)])
    style_excel_sheet(worksheet)
    return worksheet


def add_case_work_excel_sheet(workbook, title, reports, filters):
    worksheet = workbook.create_sheet(title=title[:31])
    period = get_export_period_label(filters)
    worksheet.append(['THANE CITY CYBER CELL REPORT'])
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    worksheet.append([period])
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    worksheet.append([])
    for index, report in enumerate(reports, start=1):
        worksheet.append([f'{index}) {report.name} - {report.police_station.name}'])
        worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=4)
        worksheet.append([f'{index}) Case No.: {report.case_no}'])
        worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=4)
        worksheet.append(['Police Station:', report.police_station.name, 'Investigating Officer:', report.investigating_officer])
        worksheet.append(['Case Status:', report.case_status, 'Approval Status:', report.approval_status])
        worksheet.append(['Submitted Date:', format_submitted_at(report.submitted_at), 'Report No:', report.report_number])
        worksheet.append(['Work Description:'])
        worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=4)
        for item in report.work_items.all():
            worksheet.append([item.work_title])
            worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=4)
            worksheet.append([item.work_description])
            worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=4)
        worksheet.append([])

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color='9aa7c7'),
                right=Side(style='thin', color='9aa7c7'),
                top=Side(style='thin', color='9aa7c7'),
                bottom=Side(style='thin', color='9aa7c7'),
            )
    for row_number in (1, 2):
        for cell in worksheet[row_number]:
            cell.font = Font(bold=True, size=14 if row_number == 1 else 12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for index, width in enumerate([24, 28, 24, 30], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    return worksheet


def build_recovery_workbook(reports, filters, title='Recovery Reports'):
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_recovery_excel_sheet(workbook, title, reports, filters)
    return workbook


def build_combined_workbook(recovery_reports, case_reports, filters, title='Investigation Report'):
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_recovery_excel_sheet(workbook, 'Recovery Reports', recovery_reports, filters)
    add_case_work_excel_sheet(workbook, 'Case Work Reports', case_reports, filters)
    workbook.properties.title = title
    return workbook


def build_case_work_workbook(reports, filters, title='Case Work Reports'):
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_case_work_excel_sheet(workbook, title, reports, filters)
    return workbook


def make_pdf_response(elements, filename, pagesize=landscape(A4)):
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    document.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def get_pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='SmallCell',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    ))
    styles.add(ParagraphStyle(
        name='TitleCenter',
        parent=styles['Title'],
        alignment=1,
        fontSize=14,
        leading=18,
    ))
    styles.add(ParagraphStyle(
        name='BlockText',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
    ))
    return styles


def build_recovery_pdf_elements(reports, filters, title='ZONE - 1'):
    styles = get_pdf_styles()
    elements = [
        Paragraph(title, styles['TitleCenter']),
        Paragraph(get_export_period_label(filters), styles['TitleCenter']),
        Spacer(1, 8),
    ]
    total_mobile = 0
    total_financial = 0
    rows = [[
        'Sr.N',
        'Name',
        'Name of the Police Station',
        'Types of Fraud',
        'Mobile Recovery',
        'Financial Recovery',
    ]]
    for index, report in enumerate(reports, start=1):
        total_mobile += report.mobile_recovery_count or 0
        total_financial += report.financial_recovery_amount or 0
        rows.append([
            str(index),
            Paragraph(pdf_text(report.name), styles['SmallCell']),
            Paragraph(pdf_text(report.police_station.name), styles['SmallCell']),
            Paragraph(pdf_text(report.fraud_type.name), styles['SmallCell']),
            format_mobile_count(report.mobile_recovery_count),
            format_financial_amount(report.financial_recovery_amount),
        ])
    rows.append(['', 'Total', '', '', format_mobile_count(total_mobile), format_financial_amount(total_financial)])
    table = Table(rows, colWidths=[16 * mm, 45 * mm, 65 * mm, 42 * mm, 35 * mm, 43 * mm])
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    return elements


def build_case_work_pdf_elements(reports, filters, title='THANE CITY CYBER CELL REPORT'):
    styles = get_pdf_styles()
    elements = [
        Paragraph(title, styles['TitleCenter']),
        Paragraph(get_export_period_label(filters), styles['TitleCenter']),
        Spacer(1, 8),
    ]
    for index, report in enumerate(reports, start=1):
        rows = [
            [Paragraph(f'<b>{index}) {pdf_text(report.name)} - {pdf_text(report.police_station.name)}</b>', styles['BlockText'])],
            [Paragraph(f'<b>{index}) Case No.:</b> {pdf_text(report.case_no)}', styles['BlockText'])],
            [Paragraph(f'Police Station: {pdf_text(report.police_station.name)}', styles['BlockText'])],
            [Paragraph(f'Investigating Officer: {pdf_text(report.investigating_officer)}', styles['BlockText'])],
            [Paragraph('Work Description:', styles['BlockText'])],
        ]
        for item in report.work_items.all():
            rows.append([Paragraph(f'<b>{pdf_text(item.work_title)}</b>', styles['BlockText'])])
            rows.append([Paragraph(pdf_text(item.work_description).replace('\n', '<br/>'), styles['BlockText'])])
        table = Table(rows, colWidths=[180 * mm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 8))
    return elements


def save_recovery_report_form(request, form, report=None):
    action = request.POST.get('action', 'draft')
    recovery_report = form.save(commit=False)
    if not recovery_report.submitted_by_id:
        recovery_report.submitted_by = request.user
    recovery_report.name = get_user_display_name(recovery_report.submitted_by or request.user)

    if action == 'submit':
        recovery_report.approval_status = RecoveryReport.STATUS_SUBMITTED
        recovery_report.submitted_at = timezone.now()
        recovery_report.admin_remarks = ''
        message = 'Recovery report submitted successfully.'
    else:
        recovery_report.approval_status = RecoveryReport.STATUS_DRAFT
        if report is None:
            recovery_report.submitted_at = None
        message = 'Recovery report saved as draft.'

    recovery_report.save()
    form.save_m2m()
    messages.success(request, message)
    return recovery_report


def save_case_work_report_form(request, form, formset, report=None):
    action = request.POST.get('action', 'draft')
    case_report = form.save(commit=False)
    if not case_report.submitted_by_id:
        case_report.submitted_by = request.user
    case_report.name = get_user_display_name(case_report.submitted_by or request.user)

    if action == 'submit':
        case_report.approval_status = CaseWorkReport.STATUS_SUBMITTED
        case_report.submitted_at = timezone.now()
        case_report.admin_remarks = ''
        message = 'Case work report submitted successfully.'
    else:
        case_report.approval_status = CaseWorkReport.STATUS_DRAFT
        if report is None:
            case_report.submitted_at = None
        message = 'Case work report saved as draft.'

    case_report.save()
    formset.instance = case_report
    instances = formset.save(commit=False)
    for index, item in enumerate(instances, start=1):
        item.sort_order = index
        item.save()
    for deleted_item in formset.deleted_objects:
        deleted_item.delete()
    form.save_m2m()
    messages.success(request, message)
    return case_report


@login_required
@investigator_required
def investigator_dashboard(request):
    recovery_reports = RecoveryReport.objects.filter(is_archived=False)
    case_reports = CaseWorkReport.objects.filter(is_archived=False)
    if request.user.role == 'investigator':
        recovery_reports = recovery_reports.filter(submitted_by=request.user)
        case_reports = case_reports.filter(submitted_by=request.user)

    recovery_reports, case_reports, filters = apply_dashboard_filters(request, recovery_reports, case_reports)
    recovery_totals = recovery_reports.aggregate(
        mobile=Sum('mobile_recovery_count'),
        financial=Sum('financial_recovery_amount'),
    )
    recovery_count = recovery_reports.count()
    case_count = case_reports.count()
    total_reports = recovery_count + case_count
    submitted_count = combined_status_count(recovery_reports, case_reports, RecoveryReport.STATUS_SUBMITTED)
    approved_count = combined_status_count(recovery_reports, case_reports, RecoveryReport.STATUS_APPROVED)
    rejected_count = combined_status_count(recovery_reports, case_reports, RecoveryReport.STATUS_REJECTED)
    draft_count = combined_status_count(recovery_reports, case_reports, RecoveryReport.STATUS_DRAFT)
    recent_recovery_reports = recovery_reports.select_related('police_station').order_by('-created_at')[:5]
    recent_case_reports = case_reports.select_related('police_station').order_by('-created_at')[:5]

    return render(request, 'investigator/dashboard.html', {
        'filters': filters,
        'month_choices': range(1, 13),
        'year_choices': range(timezone.localdate().year - 5, timezone.localdate().year + 2),
        'approval_status_choices': RecoveryReport.APPROVAL_STATUS_CHOICES,
        'stats': {
            'total_investigators': User.objects.filter(role='investigator').count(),
            'investigators_involved': len(set(
                list(recovery_reports.exclude(submitted_by__isnull=True).values_list('submitted_by_id', flat=True)) +
                list(case_reports.exclude(submitted_by__isnull=True).values_list('submitted_by_id', flat=True))
            )),
            'recovery_reports': recovery_count,
            'case_work_reports': case_count,
            'total_reports': total_reports,
            'mobile_recovery': recovery_totals['mobile'] or 0,
            'financial_recovery': recovery_totals['financial'] or 0,
            'submitted_reports': submitted_count,
            'approved_reports': approved_count,
            'rejected_reports': rejected_count,
            'draft_reports': draft_count,
        },
        'recent_recovery_reports': recent_recovery_reports,
        'recent_case_reports': recent_case_reports,
    })


@login_required
@investigator_required
def recovery_reports(request):
    reports, filters = apply_recovery_report_filters(request, get_recovery_report_queryset(request))
    reports = reports.order_by('police_station__name', 'name', 'report_number')
    total_values = reports.aggregate(
        mobile=Sum('mobile_recovery_count'),
        financial=Sum('financial_recovery_amount'),
    )
    totals = {
        'mobile': total_values['mobile'] or 0,
        'financial': total_values['financial'] or 0,
    }
    paginator = Paginator(reports, 10)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    query_params = request.GET.copy()
    query_params.pop('page', None)
    return render(request, 'investigator/recovery_report_list.html', {
        'page_obj': page_obj,
        'totals': totals,
        'title': 'Recovery Reports' if request.user.role == 'admin' else 'My Recovery Reports',
        'filters': filters,
        'police_stations': PoliceStation.objects.filter(is_active=True).order_by('name'),
        'fraud_types': FraudType.objects.filter(is_active=True).order_by('name'),
        'approval_status_choices': RecoveryReport.APPROVAL_STATUS_CHOICES,
        'month_choices': range(1, 13),
        'year_choices': range(timezone.localdate().year - 5, timezone.localdate().year + 2),
        'query_string': urlencode(query_params, doseq=True),
    })


@login_required
@investigator_required
def recovery_report_add(request):
    if request.method == 'POST':
        form = RecoveryReportForm(request.POST, request.FILES)
        if form.is_valid():
            report = save_recovery_report_form(request, form)
            return redirect('investigator_recovery_detail', pk=report.pk)
    else:
        form = RecoveryReportForm(initial={
            'entry_date': timezone.localdate(),
            'report_month': timezone.localdate().month,
            'report_year': timezone.localdate().year,
        })

    return render(request, 'investigator/recovery_report_form.html', {
        'form': form,
        'report': None,
        'title': 'Add Recovery Report',
        'investigator_display_name': get_report_investigator_name(None, request.user),
    })


@login_required
@investigator_required
def recovery_report_edit(request, pk):
    report = get_object_or_404(get_recovery_report_queryset(request), pk=pk)
    if not user_can_edit_recovery_report(request.user, report):
        return HttpResponseForbidden("Approved or submitted reports cannot be edited by Investigator.")

    if request.method == 'POST':
        form = RecoveryReportForm(request.POST, request.FILES, instance=report)
        if form.is_valid():
            report = save_recovery_report_form(request, form, report)
            return redirect('investigator_recovery_detail', pk=report.pk)
    else:
        form = RecoveryReportForm(instance=report)

    return render(request, 'investigator/recovery_report_form.html', {
        'form': form,
        'report': report,
        'title': f'Edit {report.report_number}',
        'investigator_display_name': get_report_investigator_name(report, request.user),
    })


@login_required
@investigator_required
def recovery_report_detail(request, pk):
    report = get_object_or_404(get_recovery_report_queryset(request), pk=pk)
    can_edit = user_can_edit_recovery_report(request.user, report)
    return render(request, 'investigator/recovery_report_detail.html', {
        'report': report,
        'can_edit': can_edit,
        'show_actions': request.user.role == 'admin' or (
            request.user.role == 'investigator' and report.can_investigator_edit
        ) or can_edit,
    })


@login_required
@investigator_admin_required
def recovery_report_delete(request, pk):
    report = get_object_or_404(RecoveryReport, pk=pk, is_archived=False)
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")

    report_number = report.report_number
    if report.attachment:
        report.attachment.delete(save=False)
    report.delete()
    messages.success(request, f'{report_number} permanently deleted successfully.')
    return redirect('investigator_recovery_reports')


@login_required
@investigator_required
def recovery_report_submit(request, pk):
    report = get_object_or_404(get_recovery_report_queryset(request), pk=pk)
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")
    if request.user.role != 'admin' and report.submitted_by_id != request.user.id:
        return HttpResponseForbidden("You cannot submit another investigator's report.")
    if report.approval_status not in {RecoveryReport.STATUS_DRAFT, RecoveryReport.STATUS_REJECTED}:
        messages.error(request, 'Only Draft or Rejected reports can be submitted.')
        return redirect('investigator_recovery_reports')

    report.approval_status = RecoveryReport.STATUS_SUBMITTED
    report.submitted_at = timezone.now()
    report.admin_remarks = ''
    report.save(update_fields=['approval_status', 'submitted_at', 'admin_remarks', 'updated_at'])
    messages.success(request, f'{report.report_number} submitted successfully.')
    return redirect('investigator_recovery_reports')


@login_required
@investigator_required
def recovery_report_export(request):
    reports, filters = apply_recovery_report_filters(request, get_recovery_report_queryset(request))
    reports = reports.order_by('police_station__name', 'name', 'report_number')
    workbook = build_recovery_workbook(reports, filters)
    return make_excel_response(workbook, 'recovery_reports.xlsx')


@login_required
@investigator_required
def recovery_report_pdf(request):
    reports, filters = apply_recovery_report_filters(request, get_recovery_report_queryset(request))
    reports = reports.order_by('police_station__name', 'name', 'report_number')
    elements = build_recovery_pdf_elements(reports, filters)
    return make_pdf_response(elements, 'recovery_reports.pdf')


@login_required
@investigator_required
def recovery_report_print(request):
    reports, filters = apply_recovery_report_filters(request, get_recovery_report_queryset(request))
    reports = reports.order_by('police_station__name', 'name', 'report_number')
    totals = reports.aggregate(
        mobile=Sum('mobile_recovery_count'),
        financial=Sum('financial_recovery_amount'),
    )
    return render(request, 'investigator/print_report.html', {
        'title': 'Recovery Reports',
        'print_type': 'recovery',
        'filters': filters,
        'print_period': get_print_period_label(filters),
        'recovery_reports': reports,
        'totals': {
            'mobile': totals['mobile'] or 0,
            'financial': totals['financial'] or 0,
        },
    })


@login_required
@investigator_admin_required
def recovery_report_review(request, pk, action):
    report = get_object_or_404(RecoveryReport, pk=pk, is_archived=False)
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")
    if report.approval_status != RecoveryReport.STATUS_SUBMITTED:
        messages.error(request, 'Only submitted recovery reports can be reviewed.')
        return redirect('investigator_recovery_detail', pk=report.pk)

    remarks = (request.POST.get('admin_remarks') or '').strip()
    if action == 'approve':
        report.approval_status = RecoveryReport.STATUS_APPROVED
        report.admin_remarks = remarks
        report.save(update_fields=['approval_status', 'admin_remarks', 'updated_at'])
        messages.success(request, f'{report.report_number} approved successfully.')
    elif action == 'reject':
        if not remarks:
            messages.error(request, 'Admin remarks are required when rejecting a report.')
            return redirect('investigator_recovery_detail', pk=report.pk)
        report.approval_status = RecoveryReport.STATUS_REJECTED
        report.admin_remarks = remarks
        report.save(update_fields=['approval_status', 'admin_remarks', 'updated_at'])
        messages.success(request, f'{report.report_number} rejected successfully.')
    else:
        return HttpResponseForbidden("Invalid review action.")

    return redirect('investigator_recovery_detail', pk=report.pk)


@login_required
@investigator_required
def case_work_reports(request):
    reports, filters = apply_case_work_report_filters(request, get_case_work_report_queryset(request))
    reports = reports.order_by('-created_at')
    paginator = Paginator(reports, 10)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    query_params = request.GET.copy()
    query_params.pop('page', None)
    return render(request, 'investigator/case_work_report_list.html', {
        'page_obj': page_obj,
        'title': 'Case Work Reports' if request.user.role == 'admin' else 'My Case Work Reports',
        'filters': filters,
        'police_stations': PoliceStation.objects.filter(is_active=True).order_by('name'),
        'case_status_choices': CaseWorkReport.CASE_STATUS_CHOICES,
        'approval_status_choices': CaseWorkReport.APPROVAL_STATUS_CHOICES,
        'month_choices': range(1, 13),
        'year_choices': range(timezone.localdate().year - 5, timezone.localdate().year + 2),
        'query_string': urlencode(query_params, doseq=True),
    })


@login_required
@investigator_required
def case_work_report_export(request):
    reports, filters = apply_case_work_report_filters(request, get_case_work_report_queryset(request))
    reports = reports.order_by('-created_at')
    workbook = build_case_work_workbook(reports, filters)
    return make_excel_response(workbook, 'case_work_reports.xlsx')


@login_required
@investigator_required
def case_work_report_pdf(request):
    reports, filters = apply_case_work_report_filters(request, get_case_work_report_queryset(request))
    reports = reports.order_by('-created_at')
    elements = build_case_work_pdf_elements(reports, filters)
    return make_pdf_response(elements, 'case_work_reports.pdf', pagesize=A4)


@login_required
@investigator_required
def case_work_report_print(request):
    reports, filters = apply_case_work_report_filters(request, get_case_work_report_queryset(request))
    return render(request, 'investigator/print_report.html', {
        'title': 'Case Work Reports',
        'print_type': 'case_work',
        'filters': filters,
        'print_period': get_print_period_label(filters),
        'case_reports': reports.order_by('-created_at'),
    })


@login_required
@investigator_required
def case_work_report_add(request):
    if request.method == 'POST':
        form = CaseWorkReportForm(request.POST, request.FILES)
        formset = CaseWorkItemFormSet(request.POST, prefix='work_items')
        if form.is_valid() and formset.is_valid():
            if request.POST.get('action') == 'submit' and case_work_formset_item_count(formset) < 1:
                messages.error(request, 'At least one work title and description is required before submit.')
            else:
                report = save_case_work_report_form(request, form, formset)
                return redirect('investigator_case_work_detail', pk=report.pk)
    else:
        form = CaseWorkReportForm(initial={'entry_date': timezone.localdate()})
        formset = CaseWorkItemFormSet(prefix='work_items')

    return render(request, 'investigator/case_work_report_form.html', {
        'form': form,
        'formset': formset,
        'report': None,
        'title': 'Add Case Work Report',
        'investigator_display_name': get_report_investigator_name(None, request.user),
    })


@login_required
@investigator_required
def case_work_report_edit(request, pk):
    report = get_object_or_404(get_case_work_report_queryset(request), pk=pk)
    if not user_can_edit_case_work_report(request.user, report):
        return HttpResponseForbidden("Approved or submitted reports cannot be edited by Investigator.")

    if request.method == 'POST':
        form = CaseWorkReportForm(request.POST, request.FILES, instance=report)
        formset = CaseWorkItemFormSet(request.POST, instance=report, prefix='work_items')
        if form.is_valid() and formset.is_valid():
            if request.POST.get('action') == 'submit' and case_work_formset_item_count(formset) < 1:
                messages.error(request, 'At least one work title and description is required before submit.')
            else:
                report = save_case_work_report_form(request, form, formset, report)
                return redirect('investigator_case_work_detail', pk=report.pk)
    else:
        form = CaseWorkReportForm(instance=report)
        formset = CaseWorkItemFormSet(instance=report, prefix='work_items')

    return render(request, 'investigator/case_work_report_form.html', {
        'form': form,
        'formset': formset,
        'report': report,
        'title': f'Edit {report.report_number}',
        'investigator_display_name': get_report_investigator_name(report, request.user),
    })


@login_required
@investigator_required
def case_work_report_detail(request, pk):
    report = get_object_or_404(get_case_work_report_queryset(request), pk=pk)
    can_edit = user_can_edit_case_work_report(request.user, report)
    return render(request, 'investigator/case_work_report_detail.html', {
        'report': report,
        'can_edit': can_edit,
        'show_actions': request.user.role == 'admin' or (
            request.user.role == 'investigator' and report.can_investigator_edit
        ) or can_edit,
    })


@login_required
@investigator_admin_required
def case_work_report_delete(request, pk):
    report = get_object_or_404(CaseWorkReport, pk=pk, is_archived=False)
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")

    report_number = report.report_number
    if report.attachment:
        report.attachment.delete(save=False)
    report.delete()
    messages.success(request, f'{report_number} permanently deleted successfully.')
    return redirect('investigator_case_work_reports')


@login_required
@investigator_required
def case_work_report_submit(request, pk):
    report = get_object_or_404(get_case_work_report_queryset(request), pk=pk)
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")
    if request.user.role != 'admin' and report.submitted_by_id != request.user.id:
        return HttpResponseForbidden("You cannot submit another investigator's report.")
    if report.approval_status not in {CaseWorkReport.STATUS_DRAFT, CaseWorkReport.STATUS_REJECTED}:
        messages.error(request, 'Only Draft or Rejected reports can be submitted.')
        return redirect('investigator_case_work_detail', pk=report.pk)
    if not report.work_items.exists():
        messages.error(request, 'At least one work title and description is required before submit.')
        return redirect('investigator_case_work_edit', pk=report.pk)

    report.approval_status = CaseWorkReport.STATUS_SUBMITTED
    report.submitted_at = timezone.now()
    report.admin_remarks = ''
    report.save(update_fields=['approval_status', 'submitted_at', 'admin_remarks', 'updated_at'])
    messages.success(request, f'{report.report_number} submitted successfully.')
    return redirect('investigator_case_work_detail', pk=report.pk)


@login_required
@investigator_admin_required
def case_work_report_review(request, pk, action):
    report = get_object_or_404(CaseWorkReport, pk=pk, is_archived=False)
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")
    if report.approval_status != CaseWorkReport.STATUS_SUBMITTED:
        messages.error(request, 'Only submitted case work reports can be reviewed.')
        return redirect('investigator_case_work_detail', pk=report.pk)

    remarks = (request.POST.get('admin_remarks') or '').strip()
    if action == 'approve':
        report.approval_status = CaseWorkReport.STATUS_APPROVED
        report.admin_remarks = remarks
        report.save(update_fields=['approval_status', 'admin_remarks', 'updated_at'])
        messages.success(request, f'{report.report_number} approved successfully.')
    elif action == 'reject':
        if not remarks:
            messages.error(request, 'Admin remarks are required when rejecting a report.')
            return redirect('investigator_case_work_detail', pk=report.pk)
        report.approval_status = CaseWorkReport.STATUS_REJECTED
        report.admin_remarks = remarks
        report.save(update_fields=['approval_status', 'admin_remarks', 'updated_at'])
        messages.success(request, f'{report.report_number} rejected successfully.')
    else:
        return HttpResponseForbidden("Invalid review action.")

    return redirect('investigator_case_work_detail', pk=report.pk)


@login_required
@investigator_admin_required
def investigator_wise_report(request):
    filters = {
        'investigator': request.GET.get('investigator', '').strip(),
        'month': request.GET.get('month', '').strip(),
        'year': request.GET.get('year', '').strip(),
        'approval_status': request.GET.get('approval_status', '').strip(),
    }
    investigators = User.objects.filter(role='investigator').order_by('first_name', 'last_name', 'username')
    selected_investigator = None
    recovery_reports = RecoveryReport.objects.none()
    case_reports = CaseWorkReport.objects.none()

    if filters['investigator'].isdigit():
        selected_investigator = investigators.filter(pk=filters['investigator']).first()

    if selected_investigator:
        recovery_reports = RecoveryReport.objects.select_related(
            'police_station',
            'fraud_type',
        ).filter(is_archived=False, submitted_by=selected_investigator)
        case_reports = CaseWorkReport.objects.select_related(
            'police_station',
        ).prefetch_related('work_items').filter(is_archived=False, submitted_by=selected_investigator)

        if filters['month'].isdigit():
            recovery_reports = recovery_reports.filter(report_month=filters['month'])
            case_reports = case_reports.filter(entry_date__month=filters['month'])
        if filters['year'].isdigit():
            recovery_reports = recovery_reports.filter(report_year=filters['year'])
            case_reports = case_reports.filter(entry_date__year=filters['year'])
        if filters['approval_status']:
            recovery_reports = recovery_reports.filter(approval_status=filters['approval_status'])
            case_reports = case_reports.filter(approval_status=filters['approval_status'])

    recovery_reports = recovery_reports.order_by('-created_at')
    case_reports = case_reports.order_by('-created_at')
    recovery_totals = recovery_reports.aggregate(
        mobile=Sum('mobile_recovery_count'),
        financial=Sum('financial_recovery_amount'),
    )
    police_stations = sorted(set(
        list(recovery_reports.values_list('police_station__name', flat=True)) +
        list(case_reports.values_list('police_station__name', flat=True))
    ))
    fraud_types = sorted(set(recovery_reports.values_list('fraud_type__name', flat=True)))

    return render(request, 'investigator/investigator_wise_report.html', {
        'filters': filters,
        'investigators': investigators,
        'selected_investigator': selected_investigator,
        'recovery_reports': recovery_reports,
        'case_reports': case_reports,
        'month_choices': range(1, 13),
        'year_choices': range(timezone.localdate().year - 5, timezone.localdate().year + 2),
        'approval_status_choices': RecoveryReport.APPROVAL_STATUS_CHOICES,
        'totals': {
            'recovery_count': recovery_reports.count(),
            'case_count': case_reports.count(),
            'mobile': recovery_totals['mobile'] or 0,
            'financial': recovery_totals['financial'] or 0,
            'police_station_count': len(police_stations),
            'fraud_type_count': len(fraud_types),
        },
        'police_stations_used': police_stations,
        'fraud_types_used': fraud_types,
    })


@login_required
@investigator_admin_required
def investigator_wise_report_export(request):
    filters, selected_investigator, recovery_reports, case_reports = get_investigator_wise_export_data(request)
    if not selected_investigator:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Investigator Report'
        worksheet.append(['Select an investigator to export report data.'])
        return make_excel_response(workbook, 'investigator_wise_report.xlsx')
    workbook = build_combined_workbook(
        recovery_reports.order_by('police_station__name', 'name', 'report_number'),
        case_reports.order_by('-created_at'),
        filters,
        'Investigator-wise Report',
    )
    summary = workbook.create_sheet(title='Summary', index=0)
    recovery_totals = recovery_reports.aggregate(
        mobile=Sum('mobile_recovery_count'),
        financial=Sum('financial_recovery_amount'),
    )
    summary_rows = [
        ['Investigator-wise Report'],
        ['Investigator', get_user_display_name(selected_investigator)],
        ['Username', selected_investigator.username],
        ['Email', selected_investigator.email or ''],
        ['Period', get_export_period_label(filters)],
        ['Total Recovery Reports', recovery_reports.count()],
        ['Total Case Work Reports', case_reports.count()],
        ['Total Mobile Recovery', recovery_totals['mobile'] or 0],
        ['Total Financial Recovery', recovery_totals['financial'] or 0],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions['A'].width = 28
    summary.column_dimensions['B'].width = 38
    summary['A1'].font = Font(bold=True, size=14)
    return make_excel_response(workbook, 'investigator_wise_report.xlsx')


def get_investigator_wise_export_data(request):
    filters = {
        'investigator': request.GET.get('investigator', '').strip(),
        'month': request.GET.get('month', '').strip(),
        'year': request.GET.get('year', '').strip(),
        'approval_status': request.GET.get('approval_status', '').strip(),
    }
    selected_investigator = None
    if filters['investigator'].isdigit():
        selected_investigator = User.objects.filter(role='investigator', pk=filters['investigator']).first()

    if not selected_investigator:
        return filters, selected_investigator, RecoveryReport.objects.none(), CaseWorkReport.objects.none()

    recovery_reports = RecoveryReport.objects.select_related(
        'police_station',
        'fraud_type',
    ).filter(is_archived=False, submitted_by=selected_investigator)
    case_reports = CaseWorkReport.objects.select_related(
        'police_station',
    ).filter(is_archived=False, submitted_by=selected_investigator)

    if filters['month'].isdigit():
        recovery_reports = recovery_reports.filter(report_month=filters['month'])
        case_reports = case_reports.filter(entry_date__month=filters['month'])
    if filters['year'].isdigit():
        recovery_reports = recovery_reports.filter(report_year=filters['year'])
        case_reports = case_reports.filter(entry_date__year=filters['year'])
    if filters['approval_status']:
        recovery_reports = recovery_reports.filter(approval_status=filters['approval_status'])
        case_reports = case_reports.filter(approval_status=filters['approval_status'])

    return filters, selected_investigator, recovery_reports, case_reports


@login_required
@investigator_admin_required
def investigator_wise_report_pdf(request):
    filters, selected_investigator, recovery_reports, case_reports = get_investigator_wise_export_data(request)
    if not selected_investigator:
        return HttpResponse('Select an investigator to download report data.', status=400)
    elements = [
        Paragraph(
            f'INVESTIGATOR-WISE REPORT - {pdf_text(get_user_display_name(selected_investigator))}',
            get_pdf_styles()['TitleCenter'],
        ),
        Spacer(1, 8),
    ]
    elements.extend(build_recovery_pdf_elements(
        recovery_reports.order_by('police_station__name', 'name', 'report_number'),
        filters,
        title='RECOVERY REPORT',
    ))
    elements.append(Spacer(1, 12))
    elements.extend(build_case_work_pdf_elements(
        case_reports.order_by('-created_at'),
        filters,
        title='CASE WORK REPORT',
    ))
    return make_pdf_response(elements, 'investigator_wise_report.pdf')


@login_required
@investigator_admin_required
def investigator_wise_report_print(request):
    filters = {
        'investigator': request.GET.get('investigator', '').strip(),
        'month': request.GET.get('month', '').strip(),
        'year': request.GET.get('year', '').strip(),
        'approval_status': request.GET.get('approval_status', '').strip(),
    }
    selected_investigator = None
    recovery_reports = RecoveryReport.objects.none()
    case_reports = CaseWorkReport.objects.none()
    if filters['investigator'].isdigit():
        selected_investigator = User.objects.filter(role='investigator', pk=filters['investigator']).first()

    if selected_investigator:
        recovery_reports = RecoveryReport.objects.select_related('police_station', 'fraud_type').filter(
            is_archived=False,
            submitted_by=selected_investigator,
        )
        case_reports = CaseWorkReport.objects.select_related('police_station').filter(
            is_archived=False,
            submitted_by=selected_investigator,
        )
        if filters['month'].isdigit():
            recovery_reports = recovery_reports.filter(report_month=filters['month'])
            case_reports = case_reports.filter(entry_date__month=filters['month'])
        if filters['year'].isdigit():
            recovery_reports = recovery_reports.filter(report_year=filters['year'])
            case_reports = case_reports.filter(entry_date__year=filters['year'])
        if filters['approval_status']:
            recovery_reports = recovery_reports.filter(approval_status=filters['approval_status'])
            case_reports = case_reports.filter(approval_status=filters['approval_status'])

    recovery_totals = recovery_reports.aggregate(
        mobile=Sum('mobile_recovery_count'),
        financial=Sum('financial_recovery_amount'),
    )
    return render(request, 'investigator/print_report.html', {
        'title': 'Investigator-wise Report',
        'print_type': 'investigator_wise',
        'filters': filters,
        'print_period': get_print_period_label(filters),
        'selected_investigator': selected_investigator,
        'recovery_reports': recovery_reports.order_by('-created_at'),
        'case_reports': case_reports.order_by('-created_at'),
        'totals': {
            'mobile': recovery_totals['mobile'] or 0,
            'financial': recovery_totals['financial'] or 0,
            'recovery_count': recovery_reports.count(),
            'case_count': case_reports.count(),
        },
    })


@login_required
@investigator_admin_required
def universal_report(request):
    filters, recovery_reports, case_reports, totals = build_universal_report_data(request)
    query_params = request.GET.copy()

    return render(request, 'investigator/universal_report.html', {
        'filters': filters,
        'investigators': User.objects.filter(role='investigator').order_by('first_name', 'last_name', 'username'),
        'police_stations': PoliceStation.objects.filter(is_active=True).order_by('name'),
        'fraud_types': FraudType.objects.filter(is_active=True).order_by('name'),
        'case_status_choices': CaseWorkReport.CASE_STATUS_CHOICES,
        'approval_status_choices': RecoveryReport.APPROVAL_STATUS_CHOICES,
        'month_choices': range(1, 13),
        'year_choices': range(timezone.localdate().year - 5, timezone.localdate().year + 2),
        'recovery_reports': recovery_reports,
        'case_reports': case_reports,
        'totals': totals,
        'query_string': urlencode(query_params, doseq=True),
    })


@login_required
@investigator_admin_required
def universal_report_export(request):
    filters, recovery_reports, case_reports, totals = build_universal_report_data(request)
    workbook = build_combined_workbook(recovery_reports, case_reports, filters, 'Universal Report')
    summary = workbook.create_sheet(title='Summary', index=0)
    for row in [
        ['Universal Report'],
        ['Period', get_export_period_label(filters)],
        ['Recovery Reports', totals['recovery_count']],
        ['Case Work Reports', totals['case_count']],
        ['Investigators', totals['investigators']],
        ['Mobile Recovery', totals['mobile']],
        ['Financial Recovery', totals['financial']],
    ]:
        summary.append(row)
    summary.column_dimensions['A'].width = 28
    summary.column_dimensions['B'].width = 38
    summary['A1'].font = Font(bold=True, size=14)
    return make_excel_response(workbook, 'universal_report.xlsx')


@login_required
@investigator_admin_required
def universal_report_pdf(request):
    filters, recovery_reports, case_reports, _totals = build_universal_report_data(request)
    elements = [
        Paragraph('UNIVERSAL INVESTIGATION REPORT', get_pdf_styles()['TitleCenter']),
        Spacer(1, 8),
    ]
    elements.extend(build_recovery_pdf_elements(recovery_reports, filters, title='RECOVERY REPORT'))
    elements.append(Spacer(1, 12))
    elements.extend(build_case_work_pdf_elements(case_reports, filters, title='CASE WORK REPORT'))
    return make_pdf_response(elements, 'universal_report.pdf')


@login_required
@investigator_admin_required
def universal_report_print(request):
    filters, recovery_reports, case_reports, totals = build_universal_report_data(request)
    return render(request, 'investigator/print_report.html', {
        'title': 'Universal Report',
        'print_type': 'universal',
        'filters': filters,
        'print_period': get_print_period_label(filters),
        'recovery_reports': recovery_reports,
        'case_reports': case_reports,
        'totals': totals,
    })


@login_required
@investigator_admin_required
def approval_review(request):
    return render_panel_page(
        request,
        'Approval / Review',
        'Admin approval and rejection page placeholder.',
        'Approval workflow will become active after report forms are built.',
        admin_only=True,
    )


@login_required
@investigator_admin_required
def master_settings(request):
    return render(request, 'investigator/master_settings.html')


@login_required
@investigator_admin_required
def export_print(request):
    return render(request, 'investigator/export_print.html')


@login_required
@investigator_admin_required
def investigator_user_management(request):
    form = InvestigatorUserCreateForm()
    if request.method == 'POST':
        form = InvestigatorUserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Investigator "{user.username}" created successfully. They can login with this username and password.')
            return redirect('investigator_user_management')

    query = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    investigators = User.objects.filter(role='investigator').order_by('first_name', 'username')
    if query:
        investigators = investigators.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(username__icontains=query) |
            Q(email__icontains=query)
        )
    if status == 'active':
        investigators = investigators.filter(is_active=True)
    elif status == 'inactive':
        investigators = investigators.filter(is_active=False)

    paginator = Paginator(investigators, 10)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'investigator/user_management.html', {
        'form': form,
        'page_obj': page_obj,
        'query': query,
        'status': status,
    })


@login_required
@investigator_admin_required
def investigator_user_edit(request, pk):
    investigator = get_object_or_404(User, pk=pk, role='investigator', is_deleted=False)
    if request.method == 'POST':
        form = InvestigatorUserEditForm(request.POST, instance=investigator)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Investigator "{user.username}" updated successfully.')
            return redirect('investigator_user_management')
    else:
        form = InvestigatorUserEditForm(instance=investigator)

    return render(request, 'investigator/user_edit.html', {
        'form': form,
        'investigator': investigator,
    })


@login_required
@investigator_admin_required
def investigator_user_toggle(request, pk):
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")
    investigator = get_object_or_404(User, pk=pk, role='investigator', is_deleted=False)
    investigator.is_active = not investigator.is_active
    investigator.save(update_fields=['is_active'])
    status = 'activated' if investigator.is_active else 'deactivated'
    messages.success(request, f'Investigator "{investigator.username}" {status} successfully.')
    return redirect('investigator_user_management')


def master_list_view(request, model, form_class, template_name, redirect_name, title, description):
    form = form_class()
    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            item = form.save()
            messages.success(request, f'{title} "{item.name}" added successfully.')
            return redirect(redirect_name)

    query = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    items = model.objects.all().order_by('name')

    if query:
        items = items.filter(Q(name__icontains=query))
    if status == 'active':
        items = items.filter(is_active=True)
    elif status == 'inactive':
        items = items.filter(is_active=False)

    paginator = Paginator(items, 10)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, template_name, {
        'form': form,
        'page_obj': page_obj,
        'query': query,
        'status': status,
        'title': title,
        'description': description,
        'redirect_name': redirect_name,
    })


def master_edit_view(request, pk, model, form_class, template_name, redirect_name, title):
    item = get_object_or_404(model, pk=pk)
    if request.method == 'POST':
        form = form_class(request.POST, instance=item)
        if form.is_valid():
            item = form.save()
            messages.success(request, f'{title} "{item.name}" updated successfully.')
            return redirect(redirect_name)
    else:
        form = form_class(instance=item)

    return render(request, template_name, {
        'form': form,
        'item': item,
        'title': title,
        'redirect_name': redirect_name,
    })


def master_toggle_view(request, pk, model, redirect_name, title):
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request.")

    item = get_object_or_404(model, pk=pk)
    item.is_active = not item.is_active
    item.save(update_fields=['is_active', 'updated_at'])
    status = 'activated' if item.is_active else 'deactivated'
    messages.success(request, f'{title} "{item.name}" {status} successfully.')
    return redirect(redirect_name)


@login_required
@investigator_admin_required
def fraud_type_list(request):
    return master_list_view(
        request,
        FraudType,
        FraudTypeForm,
        'investigator/master_list.html',
        'investigator_fraud_type_list',
        'Fraud Type',
        'Manage fraud type dropdown values for recovery reports.',
    )


@login_required
@investigator_admin_required
def fraud_type_edit(request, pk):
    return master_edit_view(
        request,
        pk,
        FraudType,
        FraudTypeForm,
        'investigator/master_edit.html',
        'investigator_fraud_type_list',
        'Fraud Type',
    )


@login_required
@investigator_admin_required
def fraud_type_toggle(request, pk):
    return master_toggle_view(request, pk, FraudType, 'investigator_fraud_type_list', 'Fraud Type')


@login_required
@investigator_admin_required
def police_station_list(request):
    return master_list_view(
        request,
        PoliceStation,
        PoliceStationForm,
        'investigator/master_list.html',
        'investigator_police_station_list',
        'Police Station',
        'Manage police station dropdown values for investigation reports.',
    )


@login_required
@investigator_admin_required
def police_station_edit(request, pk):
    return master_edit_view(
        request,
        pk,
        PoliceStation,
        PoliceStationForm,
        'investigator/master_edit.html',
        'investigator_police_station_list',
        'Police Station',
    )


@login_required
@investigator_admin_required
def police_station_toggle(request, pk):
    return master_toggle_view(request, pk, PoliceStation, 'investigator_police_station_list', 'Police Station')
