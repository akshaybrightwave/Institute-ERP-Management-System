from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.contrib import messages
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.db import transaction
from django.urls import reverse
import datetime
from datetime import datetime, date, time, timedelta
import json
import csv
from django.utils import timezone
from django.utils.dateparse import parse_date

from .models import Inquiry, Lead, CallLog, FollowUp, LeadImport, LeadNote, LeadActivity, ImportErrorLog, CounselingSession, VisitSheet, AdmissionSheet
from .forms import InquiryForm, LeadConversionForm, LeadForm, CallLogForm, FollowUpForm, CounselingSessionForm, CounselorFollowUpForm, CounselorLeadStatusForm, VisitSheetForm, AdmissionSheetForm
from .decorators import superadmin_required, admin_required, telecaller_required, counselor_required, telecaller_counselor_admin_required

ADMIN_ROLES = ('admin', 'superadmin', 'SUPER_ADMIN')
TELECALLER_CALL_STATUS_EXCLUDED = {'NEW', 'ACCEPTED'}

def is_admin_user(user):
    return user.is_authenticated and user.role in ADMIN_ROLES

def get_call_status_choices_for_user(user):
    if getattr(user, 'role', None) == 'telecaller':
        return [
            choice for choice in Inquiry.CALL_STATUS_CHOICES
            if choice[0] not in TELECALLER_CALL_STATUS_EXCLUDED
        ]
    return Inquiry.CALL_STATUS_CHOICES

def can_access_inquiry(user, inquiry):
    if is_admin_user(user) or user.role == 'counselor':
        return True
    if user.role != 'telecaller':
        return False
    if inquiry.created_by_id == user.id:
        return True
    return Lead.objects.filter(inquiry=inquiry, assigned_telecaller=user).exists()

def initial_assignment_queue():
    """Admin upload/import contacts that have never entered a user work queue."""
    return admin_uploaded_assignment_records().filter(
        assigned_telecaller__isnull=True,
        assigned_counselor__isnull=True,
        first_assigned_user__isnull=True,
        first_assigned_telecaller__isnull=True,
        first_assigned_counselor__isnull=True,
    )

def admin_uploaded_assignment_records():
    """Admin-created pre-conversion contacts shown on first-assignment pages."""
    return Lead.objects.select_related('inquiry', 'assigned_telecaller', 'assigned_counselor').filter(
        converted_at__isnull=True,
        inquiry__created_by__role__in=ADMIN_ROLES,
    )

def remember_first_assignment(lead_ids, user, assigned_at, assignment_type):
    """Keep the original owner metadata stable while allowing reassignments."""
    if not lead_ids or not user:
        return

    leads = Lead.objects.filter(pk__in=lead_ids)
    if assignment_type == 'telecaller':
        leads.filter(first_assigned_telecaller__isnull=True).update(first_assigned_telecaller=user)
    elif assignment_type == 'counselor':
        leads.filter(first_assigned_counselor__isnull=True).update(first_assigned_counselor=user)

    leads.filter(first_assigned_user__isnull=True).update(
        first_assigned_user=user,
        first_assigned_date=assigned_at,
    )

def normalize_timestamp(value):
    if isinstance(value, date) and not isinstance(value, datetime):
        # Convert a date (naive) to a datetime at midnight and make it timezone‑aware
        dt = datetime.combine(value, time.min)
        return timezone.make_aware(dt) if not timezone.is_aware(dt) else dt
    return value

def parse_filter_date(value):
    """Accept native date input and the dd-mm-yyyy format shown in the UI."""
    if not value:
        return None

    parsed = parse_date(value)
    if parsed:
        return parsed

    try:
        return datetime.strptime(value, "%d-%m-%Y").date()
    except ValueError:
        return None

def filter_datetime_field_by_local_date(queryset, field_name, selected_date):
    start = timezone.make_aware(datetime.combine(selected_date, time.min), timezone.get_current_timezone())
    end = start + timedelta(days=1)
    return queryset.filter(**{
        f"{field_name}__gte": start,
        f"{field_name}__lt": end,
    })

@login_required
@telecaller_counselor_admin_required
def management_dashboard(request):
    if request.user.role in ('admin', 'superadmin', 'SUPER_ADMIN'):
        if request.user.role == 'superadmin':
            return redirect('management_super_admin_dashboard')
        return redirect('management_admin_dashboard')

    today = timezone.localdate()

    # Telecaller Dashboard: shows contacts/leads assigned to this telecaller
    # Contacts: Inquiries with a Lead record where assigned_telecaller=this user (pending conversion)
    # The inquiry may also have been created by this user
    from django.db.models import Q

    # Active contacts: assigned to telecaller (Lead exists, not yet converted)
    active_contacts_qs = Inquiry.objects.filter(
        lead__assigned_telecaller=request.user,
        lead__converted_at__isnull=True
    ).distinct()

    # All inquiries this telecaller can see (created by them OR assigned to them)
    inquiries_qs = Inquiry.objects.filter(
        Q(created_by=request.user) | Q(lead__assigned_telecaller=request.user)
    ).distinct()

    # Converted leads: assigned_telecaller=this user AND converted_at is set
    converted_leads_qs = Lead.objects.filter(
        assigned_telecaller=request.user,
        converted_at__isnull=False
    )

    # Pending inquiry followups: contacts needing follow-up (call_status=PENDING_FOLLOW_UP)
    timeout_threshold = timezone.now() - timedelta(hours=48)
    pending_inquiry_followups = active_contacts_qs.filter(
        call_status='PENDING_FOLLOW_UP',
        updated_at__gte=timeout_threshold
    ).order_by('-updated_at')[:5]

    calls_qs = CallLog.objects.filter(created_by=request.user)
    followups_qs = FollowUp.objects.filter(created_by=request.user)

    # Dashboard stats
    assigned_inquiries = {
        'today': active_contacts_qs.filter(lead__assigned_at__date=today).count(),
        'total': active_contacts_qs.count()
    }
    converted_to_lead = {
        'today': converted_leads_qs.filter(converted_at__date=today).count(),
        'total': converted_leads_qs.count()
    }
    qualified_leads = converted_to_lead
    calls_stats = {'today': calls_qs.filter(call_date__date=today).count(), 'total': calls_qs.count()}
    followups_stats = {
        'pending': followups_qs.filter(status='Pending').count(),
        'overdue': followups_qs.filter(status='Pending', followup_date__lt=today).count()
    }

    call_outcomes = {
        'busy': active_contacts_qs.filter(call_status='BUSY').count(),
        'ringing': active_contacts_qs.filter(call_status='NO_ANSWER').count(),
        'call_back': active_contacts_qs.filter(call_status='CALL_BACK').count(),
        'wrong_number': active_contacts_qs.filter(call_status='WRONG_NUMBER').count(),
        'interested': active_contacts_qs.filter(call_status='INTERESTED').count(),
        'not_interested': active_contacts_qs.filter(call_status='NOT_INTERESTED').count(),
        'pending_follow_up': active_contacts_qs.filter(call_status='PENDING_FOLLOW_UP', updated_at__gte=timeout_threshold).count(),
        'overdue_follow_up': active_contacts_qs.filter(call_status='PENDING_FOLLOW_UP', updated_at__lt=timeout_threshold).count(),
    }

    # Recent leads = recently converted leads this telecaller worked on
    recent_leads = converted_leads_qs.select_related('inquiry', 'assigned_telecaller', 'assigned_counselor').order_by('-converted_at')[:5]
    recent_activities = LeadActivity.objects.filter(
        lead__assigned_telecaller=request.user
    ).select_related('lead__inquiry', 'created_by').order_by('-created_at')[:5]
    today_followups_list = followups_qs.filter(status='Pending', followup_date=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]
    overdue_followups_list = followups_qs.filter(status='Pending', followup_date__lt=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]
    overdue_inquiry_followups = active_contacts_qs.filter(call_status='PENDING_FOLLOW_UP', updated_at__lt=timeout_threshold).order_by('updated_at')[:5]

    context = {
        'assigned_inquiries': assigned_inquiries,
        'converted_to_lead': converted_to_lead,
        'qualified_leads': qualified_leads,
        'today_calls': calls_stats,
        'pending_followups': followups_stats['pending'],
        'overdue_followups': followups_stats['overdue'],
        'call_outcomes': call_outcomes,
        'recent_leads': recent_leads,
        'recent_activities': recent_activities,
        'today_followups_list': today_followups_list,
        'overdue_followups_list': overdue_followups_list,
        'pending_inquiry_followups': pending_inquiry_followups,
        'overdue_inquiry_followups': overdue_inquiry_followups,
        'active_contacts': active_contacts_qs.select_related('lead').order_by('-lead__assigned_at')[:10],
    }
    return render(request, 'management/telecaller_dashboard.html', context)



@login_required
@admin_required
def management_admin_dashboard(request):
    today = timezone.localdate()

    # Admin strictly manages leads and assignments. No calling features.
    # Total contacts (inquiries) in the system
    total_uploaded = {'today': Inquiry.objects.filter(created_at__date=today).count(), 'total': Inquiry.objects.count()}

    # Unassigned contacts: no Lead record at all (never imported as a lead)
    unassigned_inquiries = {
        'today': Inquiry.objects.filter(lead__isnull=True, created_at__date=today).count(),
        'total': Inquiry.objects.filter(lead__isnull=True).count()
    }

    # Telecaller-assigned contacts: has a Lead with assigned_telecaller, not yet converted
    assigned_to_telecaller = {
        'today': Lead.objects.filter(assigned_telecaller__isnull=False, converted_at__isnull=True, assigned_at__date=today).count(),
        'total': Lead.objects.filter(assigned_telecaller__isnull=False, converted_at__isnull=True).count()
    }

    # Counselor telecalling contacts: has a Lead with assigned_counselor, not yet converted (telecalling mode)
    counselor_telecalling_contacts = {
        'today': Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=True, assigned_at__date=today).count(),
        'total': Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=True).count()
    }

    # Admin Queue: converted leads without a counselor (waiting for counselor assignment)
    admin_queue = {
        'today': Lead.objects.filter(converted_at__isnull=False, assigned_counselor__isnull=True, converted_at__date=today).count(),
        'total': Lead.objects.filter(converted_at__isnull=False, assigned_counselor__isnull=True).count()
    }

    # Counselor-assigned leads: properly assigned to counselor for counseling
    assigned_to_counselor = {
        'today': Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=False, counselor_assigned_at__date=today).count(),
        'total': Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=False).count(),
    }

    # Total assigned count for summary card
    total_assigned = {
        'today': Lead.objects.filter(assigned_at__date=today).count(),
        'total': Lead.objects.count()
    }

    # Tables for Admin (Uploads, Assignments)
    recent_inquiries = Inquiry.objects.order_by('-created_at')[:5]
    recent_leads = Lead.objects.select_related('inquiry', 'assigned_telecaller', 'assigned_counselor').order_by('-assigned_at')[:5]
    pending_admissions = Lead.objects.filter(counselor_status='ADMISSION').select_related('inquiry', 'assigned_counselor')

    context = {
        'total_uploaded': total_uploaded,
        'total_assigned': total_assigned,
        'unassigned_inquiries': unassigned_inquiries,
        'assigned_to_telecaller': assigned_to_telecaller,
        'counselor_telecalling_contacts': counselor_telecalling_contacts,
        'admin_queue': admin_queue,
        'assigned_to_counselor': assigned_to_counselor,
        'recent_inquiries': recent_inquiries,
        'recent_leads': recent_leads,
        'pending_admissions': pending_admissions,
    }
    return render(request, 'management/admin_dashboard.html', context)


@login_required
@admin_required
def management_super_admin_dashboard(request):
    if request.user.role == 'admin':
        return redirect('management_admin_dashboard')

    today = timezone.localdate()

    # ── Date filter support ──
    date_filter = request.GET.get('date_filter', 'all')
    start_date = None
    end_date = today

    if date_filter == 'today':
        start_date = today
    elif date_filter == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = today - timedelta(days=1)
    elif date_filter == 'this_week':
        start_date = today - timedelta(days=today.weekday())
    elif date_filter == 'this_month':
        start_date = today.replace(day=1)
    elif date_filter == 'this_quarter':
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=quarter_start_month, day=1)
    elif date_filter == 'this_year':
        start_date = today.replace(month=1, day=1)
    elif date_filter == 'custom':
        try:
            s = request.GET.get('start_date', '')
            e = request.GET.get('end_date', '')
            if s:
                start_date = datetime.strptime(s, "%Y-%m-%d").date()
            if e:
                end_date = datetime.strptime(e, "%Y-%m-%d").date()
        except ValueError:
            pass

    # Base querysets (filtered)
    leads_qs = Lead.objects.all()
    calls_qs = CallLog.objects.all()
    followups_qs = FollowUp.objects.all()

    if start_date:
        leads_qs = leads_qs.filter(created_at__date__gte=start_date)
        calls_qs = calls_qs.filter(call_date__date__gte=start_date)
        followups_qs = followups_qs.filter(followup_date__gte=start_date)
    if end_date and date_filter not in ('all', ''):
        leads_qs = leads_qs.filter(created_at__date__lte=end_date)
        calls_qs = calls_qs.filter(call_date__date__lte=end_date)
        followups_qs = followups_qs.filter(followup_date__lte=end_date)

    dashboard_cache_key = (
        f"management:super_admin_dashboard:v2:"
        f"{date_filter}:{start_date or ''}:{end_date or ''}:"
        f"{request.GET.get('view', '')}"
    )
    cached_dashboard_context = cache.get(dashboard_cache_key)
    if cached_dashboard_context:
        cached_dashboard_context = cached_dashboard_context.copy()
        recent_leads = leads_qs.select_related('inquiry', 'assigned_telecaller').order_by('-created_at')[:5]
        recent_assignments = Lead.objects.filter(
            assigned_counselor__isnull=False,
            converted_at__isnull=False,
        ).select_related('inquiry', 'assigned_telecaller', 'assigned_counselor').order_by('-assigned_at', '-converted_at')[:5]
        recent_activities = LeadActivity.objects.select_related('lead__inquiry', 'created_by').order_by('-created_at')[:5]
        today_followups_list = FollowUp.objects.filter(status='Pending', followup_date=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]
        overdue_followups_list = FollowUp.objects.filter(status='Pending', followup_date__lt=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]

        cached_dashboard_context.update({
            'recent_leads': recent_leads,
            'recent_assignments': recent_assignments,
            'recent_activities': recent_activities,
            'today_followups_list': today_followups_list,
            'overdue_followups_list': overdue_followups_list,
            'export_history': request.session.get('export_history', []),
            'date_filter': date_filter,
            'saved_view': request.GET.get('view', ''),
        })
        return render(request, 'management/super_admin_dashboard.html', cached_dashboard_context)

    # Top dashboard cards: keep these aligned with the Admin Dashboard cards.
    total_uploaded = {
        'today': Inquiry.objects.filter(created_at__date=today).count(),
        'total': Inquiry.objects.count(),
    }
    total_assigned = {
        'today': Lead.objects.filter(assigned_at__date=today).count(),
        'total': Lead.objects.count(),
    }
    unassigned_inquiries = {
        'today': Inquiry.objects.filter(lead__isnull=True, created_at__date=today).count(),
        'total': Inquiry.objects.filter(lead__isnull=True).count(),
    }
    assigned_to_telecaller = {
        'today': Lead.objects.filter(assigned_telecaller__isnull=False, assigned_at__date=today).count(),
        'total': Lead.objects.filter(assigned_telecaller__isnull=False).count(),
    }
    assigned_to_counselor = {
        'today': Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=False, assigned_at__date=today).count(),
        'total': Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=False).count(),
    }
    total_admissions_count = {
        'today': AdmissionSheet.objects.filter(admission_date=today).count(),
        'total': AdmissionSheet.objects.count(),
    }
    lost_leads_qs = Lead.objects.filter(
        Q(status='Rejected') | Q(counselor_status__in=['LOST', 'NOT_INTERESTED'])
    ).distinct()
    lost_leads_count = {
        'today': lost_leads_qs.filter(
            Q(updated_at__date=today) | Q(counselor_status_updated_at__date=today)
        ).count(),
        'total': lost_leads_qs.count(),
    }

    total_leads = {
        'today': Lead.objects.filter(created_at__date=today).count(),
        'total': leads_qs.count(),
    }
    assigned_leads = {
        'today': Lead.objects.filter(assigned_telecaller__isnull=False, assigned_at__date=today).count(),
        'total': leads_qs.filter(assigned_telecaller__isnull=False).count(),
    }
    contacted_leads = {
        'today': Lead.objects.filter(status='Contacted', updated_at__date=today).count(),
        'total': leads_qs.filter(status='Contacted').count(),
    }
    interested_leads = {
        'today': Lead.objects.filter(status='Interested', updated_at__date=today).count(),
        'total': leads_qs.filter(status='Interested').count(),
    }
    qualified_leads = {
        'today': Lead.objects.filter(status='Qualified', updated_at__date=today).count(),
        'total': leads_qs.filter(status='Qualified').count(),
    }
    rejected_leads = {
        'today': Lead.objects.filter(status='Rejected', updated_at__date=today).count(),
        'total': leads_qs.filter(status='Rejected').count(),
    }
    calls_stats = {
        'today': CallLog.objects.filter(call_date__date=today).count(),
        'total': calls_qs.count(),
    }
    followups_stats = {
        'pending': followups_qs.filter(status='Pending').count(),
        'overdue': followups_qs.filter(status='Pending', followup_date__lt=today).count(),
    }

    # Call Outcomes (unfiltered — always show totals)
    call_outcomes = {
        'accepted': Inquiry.objects.filter(call_status='ACCEPTED').count(),
        'busy': Inquiry.objects.filter(call_status='BUSY').count(),
        'call_back': Inquiry.objects.filter(call_status='CALL_BACK').count(),
        'interested': Inquiry.objects.filter(call_status='INTERESTED').count(),
        'not_interested': Inquiry.objects.filter(call_status='NOT_INTERESTED').count(),
    }

    # Admission Metrics
    admission_metrics = {
        'total': AdmissionSheet.objects.count(),
        'confirmed': AdmissionSheet.objects.filter(admission_status='CONFIRMED').count(),
        'pending': AdmissionSheet.objects.filter(admission_status='PENDING').count(),
        'cancelled': AdmissionSheet.objects.filter(admission_status='CANCELLED').count(),
    }

    # Counselor Metrics
    counseling_done = Lead.objects.filter(counselor_status='COUNSELING_DONE').count()
    today_visits = VisitSheet.objects.filter(visit_date=today).count()

    # Telecalling specific metrics
    timeout_threshold = timezone.now() - timedelta(hours=48)
    User = get_user_model()
    normal_telecallers = User.objects.filter(role='telecaller')
    telecalling_converted_leads = Lead.objects.filter(
        converted_at__isnull=False,
        first_assigned_user__in=normal_telecallers,
    )
    telecalling_leads_generated = telecalling_converted_leads.count()
    telecalling_assigned_by_admin = Lead.objects.filter(assigned_telecaller__in=normal_telecallers).count()
    telecalling_called = CallLog.objects.filter(created_by__in=normal_telecallers).count()
    telecalling_pending_followups = FollowUp.objects.filter(
        created_by__in=normal_telecallers, status='Pending', followup_date__gte=today
    ).count()
    telecalling_overdue_followups = FollowUp.objects.filter(
        created_by__in=normal_telecallers, status='Pending', followup_date__lt=today
    ).count()

    counselor_telecallers = User.objects.filter(role='counselor')
    ct_converted_leads = Lead.objects.filter(
        converted_at__isnull=False,
        first_assigned_user__in=counselor_telecallers,
    )
    ct_leads_generated = ct_converted_leads.count()
    ct_assigned_by_admin = ct_converted_leads.count()
    ct_called = CallLog.objects.filter(created_by__in=counselor_telecallers).count()
    ct_pending_followups = ct_converted_leads.filter(
        counselor_status='FOLLOW_UP_REQUIRED',
        counselor_status_updated_at__gte=timeout_threshold
    ).count()
    ct_overdue_followups = ct_converted_leads.filter(
        counselor_status='FOLLOW_UP_REQUIRED',
        counselor_status_updated_at__lt=timeout_threshold
    ).count()

    counselling_leads = Lead.objects.filter(
        assigned_counselor__isnull=False,
        converted_at__isnull=False,
    )
    counselling_assigned = counselling_leads.count()
    counselling_interested = counselling_leads.filter(counselor_status='INTERESTED').count()
    counselling_admissions_done = counselling_leads.filter(counselor_status='ADMISSION').count()
    counselling_pending_followups = counselling_leads.filter(
        counselor_status='FOLLOW_UP_REQUIRED',
        counselor_status_updated_at__gte=timeout_threshold,
    ).count()
    counselling_overdue_followups = counselling_leads.filter(
        counselor_status='FOLLOW_UP_REQUIRED',
        counselor_status_updated_at__lt=timeout_threshold,
    ).count()

    active_leads_today = (
        telecalling_converted_leads.filter(converted_at__date=today).count() +
        ct_converted_leads.filter(converted_at__date=today).count()
    )
    active_leads_count = {'today': active_leads_today, 'total': telecalling_leads_generated + ct_leads_generated}

    # Tables
    recent_leads = leads_qs.select_related('inquiry', 'assigned_telecaller').order_by('-created_at')[:5]
    recent_assignments = counselling_leads.select_related(
        'inquiry',
        'assigned_telecaller',
        'assigned_counselor',
    ).order_by('-assigned_at', '-converted_at')[:5]
    recent_activities = LeadActivity.objects.select_related('lead__inquiry', 'created_by').order_by('-created_at')[:5]
    today_followups_list = FollowUp.objects.filter(status='Pending', followup_date=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]
    overdue_followups_list = FollowUp.objects.filter(status='Pending', followup_date__lt=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]

    # Export history from session
    export_history = request.session.get('export_history', [])

    context = {
        'total_uploaded': total_uploaded,
        'total_assigned': total_assigned,
        'unassigned_inquiries': unassigned_inquiries,
        'assigned_to_telecaller': assigned_to_telecaller,
        'assigned_to_counselor': assigned_to_counselor,
        'total_admissions_count': total_admissions_count,
        'lost_leads_count': lost_leads_count,
        'total_leads': total_leads,
        'assigned_leads': assigned_leads,
        'contacted_leads': contacted_leads,
        'interested_leads': interested_leads,
        'qualified_leads': qualified_leads,
        'rejected_leads': rejected_leads,
        'today_calls': calls_stats,
        'pending_followups': followups_stats['pending'],
        'overdue_followups': followups_stats['overdue'],
        'call_outcomes': call_outcomes,
        'admission_metrics': admission_metrics,
        'recent_leads': recent_leads,
        'recent_activities': recent_activities,
        'today_followups_list': today_followups_list,
        'overdue_followups_list': overdue_followups_list,
        'export_history': export_history,
        'date_filter': date_filter,
        'saved_view': request.GET.get('view', ''),
        'counseling_done': counseling_done,
        'today_visits': today_visits,
        'telecalling_assigned_by_admin': telecalling_assigned_by_admin,
        'telecalling_called': telecalling_called,
        'telecalling_leads_generated': telecalling_leads_generated,
        'telecalling_pending_followups': telecalling_pending_followups,
        'telecalling_overdue_followups': telecalling_overdue_followups,

        'ct_assigned_by_admin': ct_assigned_by_admin,
        'ct_called': ct_called,
        'ct_leads_generated': ct_leads_generated,
        'ct_pending_followups': ct_pending_followups,
        'ct_overdue_followups': ct_overdue_followups,

        'counselling_assigned': counselling_assigned,
        'counselling_interested': counselling_interested,
        'counselling_admissions_done': counselling_admissions_done,
        'counselling_pending_followups': counselling_pending_followups,
        'counselling_overdue_followups': counselling_overdue_followups,

        'active_leads_count': active_leads_count,
        'recent_assignments': recent_assignments,
    }
    dashboard_cache_keys = [
        'total_uploaded', 'total_assigned', 'unassigned_inquiries',
        'assigned_to_telecaller', 'assigned_to_counselor',
        'total_admissions_count', 'lost_leads_count', 'total_leads',
        'assigned_leads', 'contacted_leads', 'interested_leads',
        'qualified_leads', 'rejected_leads', 'today_calls',
        'pending_followups', 'overdue_followups', 'call_outcomes',
        'admission_metrics', 'counseling_done', 'today_visits',
        'telecalling_assigned_by_admin', 'telecalling_called',
        'telecalling_leads_generated', 'telecalling_pending_followups',
        'telecalling_overdue_followups', 'ct_assigned_by_admin',
        'ct_called', 'ct_leads_generated', 'ct_pending_followups',
        'ct_overdue_followups', 'counselling_assigned',
        'counselling_interested', 'counselling_admissions_done',
        'counselling_pending_followups', 'counselling_overdue_followups',
        'active_leads_count',
    ]
    cache.set(
        dashboard_cache_key,
        {key: context[key] for key in dashboard_cache_keys},
        20,
    )
    return render(request, 'management/super_admin_dashboard.html', context)


# ==================================================
# INQUIRY CRUD
# ==================================================

@login_required
@telecaller_counselor_admin_required
def inquiry_list(request):
    scope = request.GET.get('scope', '').strip()

    if request.user.role == 'telecaller':
        if scope == 'active_contacts':
            inquiries = Inquiry.objects.filter(
                lead__assigned_telecaller=request.user,
                lead__converted_at__isnull=True,
            ).distinct()
        elif scope == 'converted_leads':
            inquiries = Inquiry.objects.filter(
                lead__assigned_telecaller=request.user,
                lead__converted_at__isnull=False,
            ).distinct()
        else:
            inquiries = Inquiry.objects.filter(
                Q(created_by=request.user) | Q(lead__assigned_telecaller=request.user)
            ).distinct()
    elif request.user.role == 'counselor':
        if scope == 'counselor_telecalling':
            inquiries = Inquiry.objects.filter(
                lead__assigned_counselor=request.user,
                lead__converted_at__isnull=True,
            ).distinct()
        else:
            inquiries = Inquiry.objects.filter(
                Q(created_by=request.user) | Q(lead__assigned_telecaller=request.user) | Q(lead__assigned_counselor=request.user)
            ).distinct()
    else:
        inquiries = Inquiry.objects.all()

    # Search
    q = request.GET.get('q', '').strip()
    if q:
        inquiries = inquiries.filter(
            Q(full_name__icontains=q) | Q(mobile_number__icontains=q) | Q(city__icontains=q)
        )

    # Filters
    status = request.GET.get('status', '').strip()
    if status:
        inquiries = inquiries.filter(status=status)

    source = request.GET.get('source', '').strip()
    if source:
        inquiries = inquiries.filter(source=source)

    call_status = request.GET.get('call_status', '').strip()
    if call_status:
        inquiries = inquiries.filter(call_status=call_status)

    overdue = request.GET.get('overdue', '').strip()
    if call_status == 'PENDING_FOLLOW_UP' and overdue in ('true', 'false'):
        timeout_threshold = timezone.now() - timedelta(hours=48)
        if overdue == 'true':
            inquiries = inquiries.filter(updated_at__lt=timeout_threshold)
        else:
            inquiries = inquiries.filter(updated_at__gte=timeout_threshold)

    date_filter = request.GET.get('date', '').strip()
    selected_date = parse_filter_date(date_filter)
    if selected_date:
        inquiries = filter_datetime_field_by_local_date(inquiries, 'created_at', selected_date)
        date_filter = selected_date.isoformat()

    inquiries = inquiries.order_by('-created_at')
    paginator = Paginator(inquiries, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    call_status_choices = get_call_status_choices_for_user(request.user)

    return render(request, 'management/inquiry_list.html', {
        'page_obj': page_obj,
        'inquiry_form': InquiryForm(),
        'q': q,
        'status': status,
        'source': source,
        'call_status': call_status,
        'scope': scope,
        'overdue': overdue,
        'date_filter': date_filter,
        'status_choices': Inquiry.STATUS_CHOICES,
        'source_choices': Inquiry.SOURCE_CHOICES,
        'call_status_choices': call_status_choices,
        'call_status_choice_values': [choice[0] for choice in call_status_choices],
        'can_delete_inquiries': is_admin_user(request.user),
    })


@login_required
@telecaller_counselor_admin_required
def inquiry_add(request):
    if request.method == 'POST':
        form = InquiryForm(request.POST)
        if form.is_valid():
            inquiry = form.save(commit=False)
            inquiry.created_by = request.user
            with transaction.atomic():
                inquiry.save()
                if is_admin_user(request.user):
                    Lead.objects.create(inquiry=inquiry, assigned_by=request.user)
            messages.success(request, f"Inquiry for {inquiry.full_name} created successfully.")
            return redirect('inquiry_list')
    else:
        form = InquiryForm()
    return render(request, 'management/inquiry_form.html', {'form': form, 'title': 'Add Inquiry'})


@login_required
@telecaller_counselor_admin_required
def inquiry_detail(request, pk):
    inquiry = get_object_or_404(Inquiry, pk=pk)
    if not can_access_inquiry(request.user, inquiry):
        return HttpResponseForbidden("Access Denied: You do not own this record.")
    return render(request, 'management/inquiry_detail.html', {'inquiry': inquiry})


@login_required
@telecaller_counselor_admin_required
def inquiry_edit(request, pk):
    inquiry = get_object_or_404(Inquiry, pk=pk)
    # Security Data Isolation
    if not can_access_inquiry(request.user, inquiry):
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    if request.method == 'POST':
        form = InquiryForm(request.POST, instance=inquiry)
        if form.is_valid():
            form.save()
            messages.success(request, f"Inquiry for {inquiry.full_name} updated successfully.")
            return redirect('inquiry_detail', pk=inquiry.pk)
    else:
        form = InquiryForm(instance=inquiry)
    return render(request, 'management/inquiry_form.html', {'form': form, 'inquiry': inquiry, 'title': 'Edit Inquiry'})


@login_required
@admin_required
def inquiry_delete(request, pk):
    inquiry = get_object_or_404(Inquiry, pk=pk)

    if request.method == 'POST':
        name = inquiry.full_name
        inquiry.delete()
        messages.success(request, f"Inquiry for {name} deleted successfully.")
        return redirect('inquiry_list')
    return render(request, 'management/inquiry_confirm_delete.html', {'inquiry': inquiry})


@login_required
@admin_required
def inquiry_bulk_delete(request):
    if request.method != 'POST':
        return HttpResponseForbidden("Bulk delete requires a POST request.")

    inquiry_ids = request.POST.getlist('inquiries')
    if not inquiry_ids:
        messages.warning(request, "Please select at least one inquiry to delete.")
        return redirect('inquiry_list')

    inquiries = Inquiry.objects.filter(pk__in=inquiry_ids)
    selected_count = inquiries.count()

    if selected_count == 0:
        messages.warning(request, "No valid inquiries were selected.")
        return redirect('inquiry_list')

    with transaction.atomic():
        inquiries.delete()

    label = "inquiry" if selected_count == 1 else "inquiries"
    messages.success(request, f"{selected_count} {label} deleted successfully.")
    return redirect('inquiry_list')


@login_required
@telecaller_counselor_admin_required
def inquiry_convert(request, pk):
    inquiry = get_object_or_404(Inquiry, pk=pk)

    # Security: telecaller can only convert inquiries they created or were assigned.
    if not can_access_inquiry(request.user, inquiry):
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    existing_lead = getattr(inquiry, 'lead', None)

    # TELECALLER CONVERSION FLOW:
    # When a telecaller is assigned a lead by admin (assigned_telecaller set, converted_at=None),
    # they should be able to "convert" it — which simply marks converted_at and puts it in the
    # Admin Queue. The admin then assigns a counselor. No counselor assignment at this step.
    telecaller_assigned_lead_convert = (
        existing_lead
        and request.user.role == 'telecaller'
        and existing_lead.assigned_telecaller_id == request.user.id
        and existing_lead.converted_at is None
    )

    # COUNSELOR TELECALLING CONVERSION FLOW:
    # When a counselor is assigned a contact for telecalling (assigned_counselor set, converted_at=None),
    # they convert it → goes to Admin Queue → Admin assigns it to their Counselor Dashboard.
    counselor_telecalling_convert = (
        existing_lead
        and request.user.role == 'counselor'
        and existing_lead.assigned_counselor_id == request.user.id
        and existing_lead.converted_at is None
    )

    if telecaller_assigned_lead_convert:
        # Telecaller marks inquiry as qualified and sends to admin queue.
        # No counselor assignment here — admin decides later.
        if request.method == 'POST':
            converted_at = timezone.now()
            existing_lead.converted_at = converted_at
            existing_lead.telecaller_assigned_at = existing_lead.telecaller_assigned_at or existing_lead.assigned_at or converted_at
            existing_lead.first_assigned_telecaller = existing_lead.first_assigned_telecaller or request.user
            existing_lead.first_assigned_user = existing_lead.first_assigned_user or request.user
            existing_lead.first_assigned_date = existing_lead.first_assigned_date or existing_lead.telecaller_assigned_at or converted_at
            existing_lead.status = existing_lead.status or 'New'
            # Clear telecaller assignment so lead goes to admin queue (unassigned for counselor step)
            # Keep assigned_telecaller so history is preserved, but remove counselor link
            existing_lead.assigned_counselor = None
            existing_lead.save()

            inquiry.status = 'Qualified'
            inquiry.save()

            log_lead_activity(
                existing_lead,
                'LEAD_CREATED',
                f"Lead converted from inquiry by telecaller {request.user.username}. Sent to admin queue for counselor assignment.",
                request.user,
            )
            messages.success(
                request,
                f"Inquiry for {inquiry.full_name} successfully converted to Lead. Admin will assign a counselor.",
            )
            return redirect('management_dashboard')

        return render(request, 'management/inquiry_convert.html', {
            'inquiry': inquiry,
            'form': None,
            'telecaller_convert_mode': True,
            'existing_lead': existing_lead,
        })

    if counselor_telecalling_convert:
        # Counselor telecalling: convert contact, goes to Admin Queue, admin assigns to their counselor dashboard.
        if request.method == 'POST':
            converted_at = timezone.now()
            existing_lead.converted_at = converted_at
            existing_lead.counselor_assigned_at = existing_lead.counselor_assigned_at or existing_lead.assigned_at or converted_at
            existing_lead.first_assigned_counselor = existing_lead.first_assigned_counselor or request.user
            existing_lead.first_assigned_user = existing_lead.first_assigned_user or request.user
            existing_lead.first_assigned_date = existing_lead.first_assigned_date or existing_lead.counselor_assigned_at or converted_at
            existing_lead.status = existing_lead.status or 'New'
            # Remove the telecalling counselor assignment so it goes to admin queue
            # Admin will re-assign this lead to the counselor's counseling dashboard
            existing_lead.assigned_counselor = None
            existing_lead.save()

            inquiry.status = 'Qualified'
            inquiry.save()

            log_lead_activity(
                existing_lead,
                'LEAD_CREATED',
                f"Lead converted from counselor telecalling by {request.user.username}. Sent to admin queue for counselor assignment.",
                request.user,
            )
            messages.success(
                request,
                f"Contact {inquiry.full_name} successfully converted to Lead. Admin will assign it to your counseling queue.",
            )
            return redirect('counselor_telecalling_dashboard')

        return render(request, 'management/inquiry_convert.html', {
            'inquiry': inquiry,
            'form': None,
            'counselor_telecalling_convert_mode': True,
            'existing_lead': existing_lead,
        })

    # Guard: prevent duplicate lead creation
    if existing_lead:
        messages.warning(request, "This inquiry has already been converted to a Lead.")
        if request.user.role == 'counselor':
            return redirect('counselor_lead_detail', pk=existing_lead.pk)
        return redirect('lead_detail', pk=existing_lead.pk)

    # NEW INQUIRY CONVERSION (no existing lead)
    # Admins can create a lead and optionally assign counselor.
    # Telecallers create a lead with themselves as telecaller, no counselor (goes to admin queue).
    # Counselors on their Inquiry Directory can create a lead assigned to themselves as counselor.
    form = LeadConversionForm(request.POST or None)
    has_active_counselors = form.fields['assigned_counselor'].queryset.exists()

    if request.method == 'POST':
        telecaller = None
        assigned_counselor = None

        if request.user.role == 'telecaller':
            # Telecaller converts: lead assigned to themselves, no counselor (admin decides)
            telecaller = request.user
            is_valid = True
            assigned_counselor = None
        elif request.user.role == 'counselor':
            # Counselor converts from their inquiry directory: assigned to themselves
            assigned_counselor = request.user
            if inquiry.created_by and getattr(inquiry.created_by, 'role', '') == 'telecaller':
                telecaller = inquiry.created_by
            is_valid = True
        else:
            # Admin: can optionally assign counselor
            is_valid = form.is_valid()
            if is_valid:
                assigned_counselor = form.cleaned_data.get('assigned_counselor')
            if inquiry.created_by and getattr(inquiry.created_by, 'role', '') == 'telecaller':
                telecaller = inquiry.created_by

        if is_valid:
            converted_at = timezone.now()
            lead = Lead.objects.create(
                inquiry=inquiry,
                assigned_telecaller=telecaller,
                assigned_counselor=assigned_counselor,
                status='New',
                priority='Warm',
                assigned_by=request.user,
                assigned_at=converted_at,
                converted_at=converted_at,
                telecaller_assigned_at=converted_at if telecaller else None,
                counselor_assigned_at=converted_at if assigned_counselor else None,
                first_assigned_telecaller=telecaller,
                first_assigned_counselor=assigned_counselor,
                first_assigned_user=telecaller or assigned_counselor,
                first_assigned_date=converted_at if (telecaller or assigned_counselor) else None,
            )

            # Auto-qualify the inquiry
            inquiry.status = 'Qualified'
            inquiry.save()

            # Audit trail
            log_lead_activity(lead, 'LEAD_CREATED', f"Lead created from inquiry conversion by {request.user.username}.", request.user)
            if telecaller:
                log_lead_activity(lead, 'ASSIGNED', f"Telecaller {telecaller.username} assigned upon conversion.", request.user)
            if assigned_counselor:
                log_lead_activity(lead, 'ASSIGNED', f"Counselor {assigned_counselor.username} assigned during conversion by {request.user.username}.", request.user)

            if request.user.role == 'telecaller':
                messages.success(request, f"Inquiry for {inquiry.full_name} converted to Lead. Admin will assign a counselor.")
                return redirect('management_dashboard')
            elif request.user.role == 'counselor':
                messages.success(request, f"Inquiry for {inquiry.full_name} converted to Lead and assigned to you.")
                return redirect('counselor_lead_detail', pk=lead.pk)
            elif assigned_counselor:
                messages.success(request, f"Inquiry for {inquiry.full_name} converted to Lead and assigned to counselor {assigned_counselor.username}.")
                return redirect('lead_detail', pk=lead.pk)
            else:
                messages.success(request, f"Inquiry for {inquiry.full_name} converted to Lead. Assign a counselor from the Counselor Assignment page.")
                return redirect('lead_detail', pk=lead.pk)

    return render(request, 'management/inquiry_convert.html', {
        'inquiry': inquiry,
        'form': form,
        'has_active_counselors': has_active_counselors,
    })


@login_required
@telecaller_counselor_admin_required
def update_call_status(request, pk):
    """AJAX endpoint to update call status inline from Inquiry Directory."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    inquiry = get_object_or_404(Inquiry, pk=pk)

    # Access control: All allowed roles can update call status
    # (Removed created_by check as Inquiries are shared in the directory)

    try:
        data = json.loads(request.body)
        new_status = data.get('call_status', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'success': False, 'message': 'Invalid data.'}, status=400)

    valid_statuses = [choice[0] for choice in get_call_status_choices_for_user(request.user)]
    if new_status not in valid_statuses:
        return JsonResponse({'success': False, 'message': f'Invalid call status: {new_status}'}, status=400)

    inquiry.call_status = new_status
    inquiry.save(update_fields=['call_status', 'updated_at'])

    return JsonResponse({
        'success': True,
        'message': 'Call status updated successfully.',
        'call_status': new_status,
        'call_status_display': dict(Inquiry.CALL_STATUS_CHOICES).get(new_status, new_status),
    })


# ==================================================
# LEAD CRUD
# ==================================================

@login_required
@telecaller_required
def lead_list(request):
    if is_admin_user(request.user):
        leads = Lead.objects.filter(converted_at__isnull=False)
    elif request.user.role == 'counselor':
        leads = Lead.objects.filter(Q(assigned_telecaller=request.user) | Q(assigned_counselor=request.user), converted_at__isnull=False)
    else:
        leads = Lead.objects.filter(assigned_telecaller=request.user, converted_at__isnull=False)

    # Search
    q = request.GET.get('q', '').strip()
    if q:
        leads = leads.filter(
            Q(inquiry__full_name__icontains=q) | Q(inquiry__mobile_number__icontains=q)
        )

    # Filters
    status = request.GET.get('status', '').strip()
    if status:
        leads = leads.filter(status=status)

    priority = request.GET.get('priority', '').strip()
    if priority:
        leads = leads.filter(priority=priority)

    date_filter = request.GET.get('date', '').strip()
    selected_date = parse_filter_date(date_filter)
    if selected_date:
        leads = filter_datetime_field_by_local_date(leads, 'converted_at', selected_date)
        date_filter = selected_date.isoformat()

    status_date_filter = request.GET.get('status_date', '').strip()
    selected_status_date = parse_filter_date(status_date_filter)
    if selected_status_date:
        leads = filter_datetime_field_by_local_date(leads, 'counselor_status_updated_at', selected_status_date)

    assigned_date_filter = request.GET.get('assigned_date', '').strip()
    selected_assigned_date = parse_filter_date(assigned_date_filter)
    if selected_assigned_date:
        leads = filter_datetime_field_by_local_date(leads, 'assigned_at', selected_assigned_date)

    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/lead_list.html', {
        'page_obj': page_obj,
        'q': q,
        'status': status,
        'priority': priority,
        'date_filter': date_filter,
        'status_choices': Lead.STATUS_CHOICES,
        'priority_choices': Lead.PRIORITY_CHOICES,
    })


@login_required
@telecaller_required
def lead_detail(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    # Security Data Isolation
    if not is_admin_user(request.user) and lead.assigned_telecaller != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    call_logs = lead.call_logs.all().order_by('-call_date')
    followups = lead.followups.all().order_by('-followup_date')

    # Check admission sheet
    try:
        admission = lead.admission_sheet
    except AdmissionSheet.DoesNotExist:
        admission = None

    notes = lead.notes_timeline.all().select_related('created_by').order_by('-created_at')

    return render(request, 'management/lead_detail.html', {
        'lead': lead,
        'call_logs': call_logs,
        'followups': followups,
        'admission': admission,
        'notes': notes,
    })


@login_required
@telecaller_required
def lead_edit(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    # Security Data Isolation
    if not is_admin_user(request.user) and lead.assigned_telecaller != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    if request.method == 'POST':
        old_status = lead.status
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            updated_lead = form.save()
            if old_status != updated_lead.status:
                log_lead_activity(updated_lead, 'STATUS_CHANGED', f"Status updated from {old_status} to {updated_lead.status}.", request.user)
            messages.success(request, f"Lead for {lead.inquiry.full_name} updated successfully.")
            return redirect('lead_detail', pk=lead.pk)
    else:
        form = LeadForm(instance=lead)
    return render(request, 'management/lead_form.html', {'form': form, 'lead': lead})


# ==================================================
# CALL LOGS
# ==================================================

@login_required
@telecaller_required
def call_log_list(request):
    if is_admin_user(request.user):
        call_logs = CallLog.objects.all()
    else:
        call_logs = CallLog.objects.filter(created_by=request.user)

    lead_id = request.GET.get('lead_id', '').strip()
    if lead_id:
        call_logs = call_logs.filter(lead_id=lead_id)

    date_filter = request.GET.get('date', '').strip()
    if date_filter:
        try:
            date_val = datetime.strptime(date_filter, "%Y-%m-%d").date()
            call_logs = call_logs.filter(call_date__date=date_val)
        except ValueError:
            pass

    paginator = Paginator(call_logs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/calllog_list.html', {
        'page_obj': page_obj,
        'lead_id': lead_id,
        'date_filter': date_filter,
    })


@login_required
@telecaller_required
def call_log_add(request):
    lead_id = request.GET.get('lead_id', '')
    lead = None
    if lead_id:
        lead = get_object_or_404(Lead, pk=lead_id)
        if not is_admin_user(request.user) and lead.assigned_telecaller != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

    if request.method == 'POST':
        form = CallLogForm(request.POST)
        selected_lead_id = request.POST.get('lead')
        selected_lead = get_object_or_404(Lead, pk=selected_lead_id)

        if not is_admin_user(request.user) and selected_lead.assigned_telecaller != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

        if form.is_valid():
            log = form.save(commit=False)
            log.lead = selected_lead
            log.created_by = request.user
            log.save()
            # Log activity
            log_lead_activity(selected_lead, 'CALL_LOG_ADDED', f"Call log added: Status '{log.call_status}', Duration {log.call_duration}s.", request.user)
            messages.success(request, f"Call log saved successfully for {selected_lead.inquiry.full_name}.")
            return redirect('lead_detail', pk=selected_lead.pk)
    else:
        form = CallLogForm()

    # Get available leads for selection
    if is_admin_user(request.user):
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_telecaller=request.user)

    return render(request, 'management/calllog_form.html', {
        'form': form,
        'leads': leads,
        'selected_lead': lead,
    })


# ==================================================
# FOLLOW UPS
# ==================================================

@login_required
@telecaller_required
def followup_list(request):
    if is_admin_user(request.user):
        followups = FollowUp.objects.all()
    else:
        followups = FollowUp.objects.filter(created_by=request.user)

    status = request.GET.get('status', '').strip()
    if status:
        followups = followups.filter(status=status)

    date_filter = request.GET.get('date', '').strip()
    if date_filter:
        followups = followups.filter(followup_date=date_filter)

    paginator = Paginator(followups, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/followup_list.html', {
        'page_obj': page_obj,
        'status': status,
        'date_filter': date_filter,
    })


@login_required
@telecaller_required
def followup_add(request):
    lead_id = request.GET.get('lead_id', '')
    lead = None
    if lead_id:
        lead = get_object_or_404(Lead, pk=lead_id)
        if not is_admin_user(request.user) and lead.assigned_telecaller != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

    if request.method == 'POST':
        form = FollowUpForm(request.POST)
        selected_lead_id = request.POST.get('lead')
        selected_lead = get_object_or_404(Lead, pk=selected_lead_id)

        if not is_admin_user(request.user) and selected_lead.assigned_telecaller != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

        if form.is_valid():
            followup = form.save(commit=False)
            followup.lead = selected_lead
            followup.created_by = request.user
            followup.save()

            # Update Lead next follow up date
            selected_lead.next_followup_date = followup.followup_date
            selected_lead.save()

            # Log activity
            log_lead_activity(selected_lead, 'FOLLOWUP_CREATED', f"Follow-up scheduled for {followup.followup_date}.", request.user)

            messages.success(request, f"Follow-up scheduled successfully for {selected_lead.inquiry.full_name}.")
            return redirect('lead_detail', pk=selected_lead.pk)
    else:
        form = FollowUpForm()

    if is_admin_user(request.user):
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_telecaller=request.user)

    return render(request, 'management/followup_form.html', {
        'form': form,
        'leads': leads,
        'selected_lead': lead,
        'title': 'Schedule Follow-Up',
    })


@login_required
@telecaller_required
def followup_edit(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if not is_admin_user(request.user) and followup.created_by != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    if request.method == 'POST':
        form = FollowUpForm(request.POST, instance=followup)
        if form.is_valid():
            form.save()

            # Sync next followup date on the lead if updated
            lead = followup.lead
            if followup.next_followup_date:
                lead.next_followup_date = followup.next_followup_date
            lead.save()

            messages.success(request, "Follow-up updated successfully.")
            return redirect('lead_detail', pk=followup.lead.pk)
    else:
        form = FollowUpForm(instance=followup)
    return render(request, 'management/followup_form.html', {
        'form': form,
        'followup': followup,
        'selected_lead': followup.lead,
        'title': 'Edit Follow-Up',
    })


@login_required
@telecaller_required
def followup_complete(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if not is_admin_user(request.user) and followup.created_by != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    followup.status = 'Completed'
    followup.save()
    # Log activity
    log_lead_activity(followup.lead, 'FOLLOWUP_COMPLETED', f"Follow-up scheduled on {followup.followup_date} completed.", request.user)
    messages.success(request, "Follow-up marked as Completed.")
    return redirect('lead_detail', pk=followup.lead.pk)


def map_header(raw_header):
    if raw_header is None:
        return None
    h = str(raw_header).strip().lower().replace('_', ' ').replace('-', ' ')
    # Normalize multiple whitespace to single space
    h = ' '.join(h.split())

    # Full Name mapping
    if h in ['full_name', 'fullname', 'full name', 'name', 'student name', 'candidate name']:
        return 'full_name'

    # Mobile Number mapping
    if h in ['mobile_number', 'mobilenumber', 'mobile number', 'mobile', 'phone', 'phone_number', 'phone number', 'contact number', 'contact_number']:
        return 'mobile_number'

    # Email mapping
    if h in ['email', 'email_address', 'email address']:
        return 'email'

    # City mapping
    if h in ['city', 'location']:
        return 'city'

    # Course interest mapping
    if h in ['course_interest', 'course interest', 'course']:
        return 'course_interest'

    # Source mapping
    if h in ['source']:
        return 'source'

    # Remarks mapping
    if h in ['remarks']:
        return 'remarks'

    return None


def format_cell_value(val):
    if val is None:
        return ''
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def log_lead_activity(lead, activity_type, description, user):
    created_by_user = user if (user and user.is_authenticated) else None
    LeadActivity.objects.create(
        lead=lead,
        activity_type=activity_type,
        description=description,
        created_by=created_by_user
    )


@login_required
@telecaller_required
def inquiry_import(request):
    if request.method == 'POST':
        file_data = request.FILES.get('file')
        if not file_data:
            messages.error(request, "Please select a file to upload.")
            return render(request, 'management/import.html')

        filename = file_data.name.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
            messages.error(request, "Only CSV and Excel files are allowed.")
            return render(request, 'management/import.html')

        if filename.endswith('.csv'):
            try:
                import csv
                import io
                decoded_file = file_data.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                header_row = next(reader, None)
                if not header_row:
                    messages.error(request, "The uploaded CSV file is empty.")
                    return render(request, 'management/import.html')

                mapped_headers = [map_header(h) for h in header_row]
                if 'full_name' not in mapped_headers or 'mobile_number' not in mapped_headers:
                    messages.error(request, "Missing required columns: full_name and mobile_number.")
                    return render(request, 'management/import.html')

                name_idx = mapped_headers.index('full_name')
                mobile_idx = mapped_headers.index('mobile_number')
                email_idx = mapped_headers.index('email') if 'email' in mapped_headers else -1
                city_idx = mapped_headers.index('city') if 'city' in mapped_headers else -1
                course_idx = mapped_headers.index('course_interest') if 'course_interest' in mapped_headers else -1
                source_idx = mapped_headers.index('source') if 'source' in mapped_headers else -1
                remarks_idx = mapped_headers.index('remarks') if 'remarks' in mapped_headers else -1

                lead_import = LeadImport.objects.create(
                    uploaded_by=request.user,
                    file=file_data,
                    total_records=0,
                    successful_records=0,
                    duplicate_records=0,
                    failed_records=0
                )

                total = 0
                success = 0
                duplicates = 0
                failed = 0
                row_idx = 1

                for row in reader:
                    if not row or not any(row):
                        continue
                    while len(row) < len(mapped_headers):
                        row.append(None)

                    row_idx += 1
                    total += 1
                    full_name = format_cell_value(row[name_idx]).strip() if name_idx != -1 else ''
                    mobile_number = format_cell_value(row[mobile_idx]).strip() if mobile_idx != -1 else ''

                    if not full_name:
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Missing Full Name"
                        )
                        continue

                    if not mobile_number:
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Missing Mobile Number"
                        )
                        continue

                    if not mobile_number.isdigit() or len(mobile_number) < 10:
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Invalid Mobile Number"
                        )
                        continue

                    if Inquiry.objects.filter(mobile_number=mobile_number).exists():
                        duplicates += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Duplicate Mobile Number"
                        )
                        continue

                    email = format_cell_value(row[email_idx]).strip() if email_idx != -1 else ''
                    city = format_cell_value(row[city_idx]).strip() if city_idx != -1 else ''
                    course_interest = format_cell_value(row[course_idx]).strip() if course_idx != -1 else ''
                    source = format_cell_value(row[source_idx]).strip() if source_idx != -1 else ''
                    remarks = format_cell_value(row[remarks_idx]).strip() if remarks_idx != -1 else ''

                    if email and ('@' not in email or '.' not in email):
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Invalid Email"
                        )
                        continue

                    matched_source = 'Website'
                    if source:
                        for c in Inquiry.SOURCE_CHOICES:
                            if c[0].lower() == source.lower():
                                matched_source = c[0]
                                break
                        else:
                            matched_source = 'Other'

                    with transaction.atomic():
                        inq = Inquiry.objects.create(
                            full_name=full_name,
                            mobile_number=mobile_number,
                            email=email or None,
                            city=city,
                            course_interest=course_interest,
                            source=matched_source,
                            remarks=remarks,
                            status='New',
                            created_by=request.user
                        )
                        Lead.objects.create(
                            inquiry=inq,
                            assigned_by=request.user if request.user.role in ['admin', 'superadmin'] else None
                        )
                    success += 1

                lead_import.total_records = total
                lead_import.successful_records = success
                lead_import.duplicate_records = duplicates
                lead_import.failed_records = failed
                lead_import.save()

                messages.success(request, (
                    f"Import summary:\n"
                    f"Total Records: {total}\n\n"
                    f"Imported Successfully: {success}\n\n"
                    f"Duplicate Records: {duplicates}\n\n"
                    f"Failed Records: {failed}"
                ))
                return redirect('import_history')
            except Exception as e:
                messages.error(request, f"Failed to process CSV file: {str(e)}")
                return render(request, 'management/import.html')

        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_data, read_only=True)
                sheet = wb.active
                rows_iter = sheet.iter_rows(values_only=True)
                header_row = next(rows_iter, None)
                if not header_row:
                    messages.error(request, "The uploaded Excel file is empty.")
                    return render(request, 'management/import.html')

                mapped_headers = [map_header(h) for h in header_row]
                if 'full_name' not in mapped_headers or 'mobile_number' not in mapped_headers:
                    messages.error(request, "Missing required columns: full_name and mobile_number.")
                    return render(request, 'management/import.html')

                name_idx = mapped_headers.index('full_name')
                mobile_idx = mapped_headers.index('mobile_number')
                email_idx = mapped_headers.index('email') if 'email' in mapped_headers else -1
                city_idx = mapped_headers.index('city') if 'city' in mapped_headers else -1
                course_idx = mapped_headers.index('course_interest') if 'course_interest' in mapped_headers else -1
                source_idx = mapped_headers.index('source') if 'source' in mapped_headers else -1
                remarks_idx = mapped_headers.index('remarks') if 'remarks' in mapped_headers else -1

                lead_import = LeadImport.objects.create(
                    uploaded_by=request.user,
                    file=file_data,
                    total_records=0,
                    successful_records=0,
                    duplicate_records=0,
                    failed_records=0
                )

                total = 0
                success = 0
                duplicates = 0
                failed = 0
                row_idx = 1

                for row in rows_iter:
                    if not row or not any(row):
                        continue
                    row_idx += 1
                    total += 1

                    full_name = format_cell_value(row[name_idx]).strip() if name_idx != -1 else ''
                    mobile_number = format_cell_value(row[mobile_idx]).strip() if mobile_idx != -1 else ''

                    if not full_name:
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Missing Full Name"
                        )
                        continue

                    if not mobile_number:
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Missing Mobile Number"
                        )
                        continue

                    if not mobile_number.isdigit() or len(mobile_number) < 10:
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Invalid Mobile Number"
                        )
                        continue

                    if Inquiry.objects.filter(mobile_number=mobile_number).exists():
                        duplicates += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Duplicate Mobile Number"
                        )
                        continue

                    email = format_cell_value(row[email_idx]).strip() if email_idx != -1 else ''
                    city = format_cell_value(row[city_idx]).strip() if city_idx != -1 else ''
                    course_interest = format_cell_value(row[course_idx]).strip() if course_idx != -1 else ''
                    source = format_cell_value(row[source_idx]).strip() if source_idx != -1 else ''
                    remarks = format_cell_value(row[remarks_idx]).strip() if remarks_idx != -1 else ''

                    if email and ('@' not in email or '.' not in email):
                        failed += 1
                        ImportErrorLog.objects.create(
                            lead_import=lead_import,
                            row_number=row_idx,
                            error_message="Invalid Email"
                        )
                        continue

                    matched_source = 'Website'
                    if source:
                        for c in Inquiry.SOURCE_CHOICES:
                            if c[0].lower() == source.lower():
                                matched_source = c[0]
                                break
                        else:
                            matched_source = 'Other'

                    with transaction.atomic():
                        inq = Inquiry.objects.create(
                            full_name=full_name,
                            mobile_number=mobile_number,
                            email=email or None,
                            city=city,
                            course_interest=course_interest,
                            source=matched_source,
                            remarks=remarks,
                            status='New',
                            created_by=request.user
                        )
                        Lead.objects.create(
                            inquiry=inq,
                            assigned_by=request.user if request.user.role in ['admin', 'superadmin'] else None
                        )
                    success += 1

                lead_import.total_records = total
                lead_import.successful_records = success
                lead_import.duplicate_records = duplicates
                lead_import.failed_records = failed
                lead_import.save()

                messages.success(request, (
                    f"Import summary:\n"
                    f"Total Records: {total}\n\n"
                    f"Imported Successfully: {success}\n\n"
                    f"Duplicate Records: {duplicates}\n\n"
                    f"Failed Records: {failed}"
                ))
                return redirect('import_history')
            except Exception as e:
                messages.error(request, f"Failed to process Excel file: {str(e)}")
                return render(request, 'management/import.html')

    return render(request, 'management/import.html')


@login_required
@telecaller_required
def download_sample_csv(request):
    import csv
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_inquiries.csv"'
    writer = csv.writer(response)
    writer.writerow(['full_name', 'mobile_number', 'email', 'city', 'course_interest', 'source', 'remarks'])
    writer.writerow(['Rahul Sharma', '9876543210', 'rahul.sharma@example.com', 'Thane', 'Java', 'Website', ''])
    writer.writerow(['Priya Patel', '9123456780', 'priya.patel@example.com', 'Mumbai', 'Python', 'Walk-In', ''])
    return response


@login_required
@telecaller_required
def download_sample_excel(request):
    import openpyxl
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sample_inquiries.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Inquiries"
    ws.append(['full_name', 'mobile_number', 'email', 'city', 'course_interest', 'source', 'remarks'])
    ws.append(['Rahul Sharma', '9876543210', 'rahul.sharma@example.com', 'Thane', 'Java', 'Website', ''])
    ws.append(['Priya Patel', '9123456780', 'priya.patel@example.com', 'Mumbai', 'Python', 'Walk-In', ''])
    wb.save(response)
    return response


@login_required
@telecaller_required
def import_history(request):
    if is_admin_user(request.user):
        imports = LeadImport.objects.all()
    else:
        imports = LeadImport.objects.filter(uploaded_by=request.user)

    paginator = Paginator(imports, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/import_history.html', {
        'page_obj': page_obj,
        'can_delete_imports': is_admin_user(request.user),
    })


@login_required
@admin_required
def import_history_delete(request, pk):
    if request.method != 'POST':
        return HttpResponseForbidden("Import history delete requires a POST request.")

    lead_import = get_object_or_404(LeadImport, pk=pk)
    lead_import.delete()
    messages.success(request, "Import history record deleted successfully.")
    return redirect(request.POST.get('next') or 'import_history')


@login_required
@admin_required
def import_history_bulk_delete(request):
    if request.method != 'POST':
        return HttpResponseForbidden("Bulk import history delete requires a POST request.")

    import_ids = request.POST.getlist('imports')
    if not import_ids:
        messages.warning(request, "Please select at least one import record to delete.")
        return redirect(request.POST.get('next') or 'import_history')

    imports = LeadImport.objects.filter(pk__in=import_ids)
    selected_count = imports.count()
    if selected_count == 0:
        messages.warning(request, "No valid import records were selected.")
        return redirect(request.POST.get('next') or 'import_history')

    imports.delete()
    label = "record" if selected_count == 1 else "records"
    messages.success(request, f"{selected_count} import history {label} deleted successfully.")
    return redirect(request.POST.get('next') or 'import_history')


@login_required
@telecaller_required
def lead_notes_list(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    if not is_admin_user(request.user) and lead.assigned_telecaller != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this lead.")

    notes = lead.notes_timeline.all().order_by('created_at')
    return render(request, 'management/lead_notes.html', {
        'lead': lead,
        'notes': notes,
    })


@login_required
@telecaller_required
def lead_note_add(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    if not is_admin_user(request.user) and lead.assigned_telecaller != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this lead.")

    if request.method == 'POST':
        note_text = request.POST.get('note', '').strip()
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', '')

        if note_text:
            note_obj = LeadNote.objects.create(
                lead=lead,
                note=note_text,
                created_by=request.user
            )
            # Log activity
            log_lead_activity(lead, 'NOTE_ADDED', f"Note added: '{note_text[:50]}...'", request.user)

            # Broadcast via Channels
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            from django.utils import timezone

            channel_layer = get_channel_layer()
            role_display = 'Telecaller' if request.user.role == 'telecaller' else ('Counselor' if request.user.role == 'counselor' else 'Admin')

            payload = {
                "note_id": note_obj.pk,
                "lead_id": lead.pk,
                "author_name": request.user.username,
                "author_role": role_display,
                "remark": note_text,
                "timestamp": timezone.localtime(note_obj.created_at).strftime("%d %b %Y %I:%M %p")
            }
            if channel_layer:
                try:
                    async_to_sync(channel_layer.group_send)(
                        f"lead_remarks_{lead.pk}",
                        {"type": "remark_message", "payload": payload}
                    )
                except Exception:
                    pass

            if is_ajax:
                return JsonResponse({"status": "success", "message": "Note added successfully."})

            messages.success(request, "Note added successfully.")
        else:
            if is_ajax:
                return JsonResponse({"status": "error", "message": "Note content cannot be empty."}, status=400)
            messages.error(request, "Note content cannot be empty.")

        return redirect(request.META.get('HTTP_REFERER') or reverse('lead_detail', kwargs={'pk': lead.pk}))

    return render(request, 'management/lead_note_form.html', {'lead': lead})


@login_required
@telecaller_required
def activities_list(request):
    if is_admin_user(request.user):
        activities = LeadActivity.objects.all()
    else:
        activities = LeadActivity.objects.filter(lead__assigned_telecaller=request.user)

    # Search
    q = request.GET.get('q', '').strip()
    if q:
        activities = activities.filter(
            Q(lead__inquiry__full_name__icontains=q) |
            Q(description__icontains=q) |
            Q(activity_type__icontains=q)
        )

    paginator = Paginator(activities, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/activity_timeline.html', {
        'page_obj': page_obj,
        'q': q,
    })


@login_required
@telecaller_required
def import_errors(request):
    import_id = request.GET.get('import_id')
    lead_import = None
    if import_id:
        lead_import = get_object_or_404(LeadImport, pk=import_id)
        if not is_admin_user(request.user) and lead_import.uploaded_by != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this record.")
        errors = ImportErrorLog.objects.filter(lead_import=lead_import)
    else:
        if is_admin_user(request.user):
            errors = ImportErrorLog.objects.all()
        else:
            errors = ImportErrorLog.objects.filter(lead_import__uploaded_by=request.user)

    # Handle CSV export
    export = request.GET.get('export', '').strip()
    if export == 'csv':
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="import_errors_{import_id or "all"}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Import ID', 'File Name', 'Row Number', 'Error Message', 'Timestamp'])
        for error in errors:
            writer.writerow([
                error.lead_import.id,
                error.lead_import.file.name,
                error.row_number,
                error.error_message,
                error.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        return response

    paginator = Paginator(errors, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/import_errors.html', {
        'page_obj': page_obj,
        'lead_import': lead_import,
        'import_id': import_id,
        'can_delete_import_errors': is_admin_user(request.user),
    })


@login_required
@admin_required
def import_error_delete(request, pk):
    if request.method != 'POST':
        return HttpResponseForbidden("Import error delete requires a POST request.")

    error = get_object_or_404(ImportErrorLog, pk=pk)
    error.delete()
    messages.success(request, "Import error record deleted successfully.")
    return redirect(request.POST.get('next') or 'import_errors')


@login_required
@admin_required
def import_error_bulk_delete(request):
    if request.method != 'POST':
        return HttpResponseForbidden("Bulk import error delete requires a POST request.")

    error_ids = request.POST.getlist('errors')
    if not error_ids:
        messages.warning(request, "Please select at least one import error to delete.")
        return redirect(request.POST.get('next') or 'import_errors')

    errors = ImportErrorLog.objects.filter(pk__in=error_ids)
    selected_count = errors.count()
    if selected_count == 0:
        messages.warning(request, "No valid import errors were selected.")
        return redirect(request.POST.get('next') or 'import_errors')

    errors.delete()
    label = "error" if selected_count == 1 else "errors"
    messages.success(request, f"{selected_count} import {label} deleted successfully.")
    return redirect(request.POST.get('next') or 'import_errors')


@login_required
@admin_required
def lead_assign(request):

    from django.contrib.auth import get_user_model
    User = get_user_model()
    telecallers = User.objects.filter(role='telecaller', is_deleted=False, is_active=True)

    # Calculate workload for dashboard
    from datetime import date
    from django.db.models import Count, Q
    from django.db import transaction
    from django.http import JsonResponse
    import json

    today = timezone.localdate()
    workload_data = []
    for tc in telecallers:
        # Active contacts: assigned leads NOT yet converted (telecalling work pending)
        stats = Lead.objects.filter(assigned_telecaller=tc).aggregate(
            active=Count('id', filter=Q(converted_at__isnull=True)),
            today=Count('id', filter=Q(assigned_at__date=today)),
            total=Count('id')
        )
        workload_data.append({
            'id': tc.id,
            'username': tc.username,
            'active': stats['active'],
            'today': stats['today'],
            'total': stats['total'],
        })

    # Telecaller Assignment display queryset:
    # Show admin-uploaded contacts in the first-assignment lifecycle.
    assignment_queue = initial_assignment_queue()
    leads = admin_uploaded_assignment_records()

    q = request.GET.get('q', '').strip()
    if q:
        leads = leads.filter(
            Q(inquiry__full_name__icontains=q) | Q(inquiry__mobile_number__icontains=q)
        )
        assignment_queue = assignment_queue.filter(
            Q(inquiry__full_name__icontains=q) | Q(inquiry__mobile_number__icontains=q)
        )

    status = request.GET.get('status', '').strip()
    if status:
        leads = leads.filter(status=status)
        assignment_queue = assignment_queue.filter(status=status)

    # Default: show all (both assigned and unassigned) so admin can see the full picture
    assigned_status = request.GET.get('assigned', 'all').strip()
    if assigned_status == 'yes':
        leads = leads.filter(Q(assigned_telecaller__isnull=False) | Q(assigned_counselor__isnull=False))
    elif assigned_status == 'no':
        leads = leads.filter(assigned_telecaller__isnull=True, assigned_counselor__isnull=True)
    elif assigned_status == 'all':
        pass  # show all

    # Handling AJAX APIs (Workload & Preview)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        action = request.POST.get('action') or request.GET.get('action')
        if action == 'workload':
            return JsonResponse({'workload': workload_data})

        if action == 'preview':
            mode = request.POST.get('mode', 'manual')
            quantity = int(request.POST.get('quantity', 0) or 0)
            assign_source = request.POST.get('assign_source', 'unassigned')

            target_leads = assignment_queue if assign_source == 'filtered' else initial_assignment_queue()

            if mode == 'quantity':
                available = target_leads.count()
                will_assign = min(quantity, available)
                return JsonResponse({
                    'requested': quantity,
                    'eligible': available,
                    'will_assign': will_assign,
                    'remaining': available - will_assign
                })
            elif mode == 'all_filtered':
                available = assignment_queue.count()
                return JsonResponse({
                    'requested': available,
                    'eligible': available,
                    'will_assign': available,
                    'remaining': 0
                })
            elif mode == 'auto_distribute':
                user_ids = request.POST.getlist('telecallers') or request.POST.getlist('telecallers[]')
                available = target_leads.count()
                return JsonResponse({
                    'requested': 'Auto',
                    'eligible': available,
                    'will_assign': available,
                    'remaining': 0,
                    'users_count': len(user_ids)
                })

    if request.method == 'POST':
        mode = request.POST.get('mode', 'manual')

        if mode == 'manual':
            telecaller_id = request.POST.get('telecaller')
            lead_ids = request.POST.getlist('leads')

            lead_id = request.POST.get('lead_id')
            if lead_id:
                lead_ids = [lead_id]

            if not telecaller_id:
                messages.error(request, "Please select a telecaller.")
            elif not lead_ids:
                messages.error(request, "Please select at least one lead.")
            else:
                telecaller = get_object_or_404(User, pk=telecaller_id, is_deleted=False, is_active=True)
                with transaction.atomic():
                    now = timezone.now()
                    lead_ids = list(initial_assignment_queue().filter(pk__in=lead_ids).values_list('id', flat=True))
                    Lead.objects.filter(pk__in=lead_ids).update(
                        assigned_telecaller=telecaller,
                        assigned_counselor=None,
                        assigned_by=request.user,
                        assigned_at=now,
                        telecaller_assigned_at=now,
                    )
                    remember_first_assignment(lead_ids, telecaller, now, 'telecaller')
                    activities = [
                        LeadActivity(
                            lead_id=int(lid),
                            activity_type='ASSIGNED',
                            description=f"Lead assigned to {telecaller.username} by admin {request.user.username}.",
                            created_by=request.user
                        ) for lid in lead_ids
                    ]
                    LeadActivity.objects.bulk_create(activities)
                messages.success(request, f"Successfully assigned {len(lead_ids)} lead(s) to {telecaller.username}.")
                return redirect(request.get_full_path())

        elif False and mode == 'counselor_telecalling':
            # Admin assigns contacts to a counselor for TELECALLING duty (NOT counselor dashboard)
            # This does NOT set converted_at — counselor must convert themselves
            counselor_id = request.POST.get('counselor_telecalling_user')
            lead_ids = request.POST.getlist('leads')

            lead_id = request.POST.get('lead_id')
            if lead_id:
                lead_ids = [lead_id]

            if not counselor_id:
                messages.error(request, "Please select a counselor for telecalling assignment.")
            elif not lead_ids:
                messages.error(request, "Please select at least one contact.")
            else:
                counselor = get_object_or_404(User, pk=counselor_id, is_deleted=False, is_active=True)
                with transaction.atomic():
                    now = timezone.now()
                    lead_ids = list(Lead.objects.filter(pk__in=lead_ids, converted_at__isnull=True).values_list('id', flat=True))
                    Lead.objects.filter(pk__in=lead_ids).update(
                        assigned_counselor=counselor,
                        assigned_telecaller=None,  # Remove any telecaller assignment
                        assigned_by=request.user,
                        assigned_at=now,
                        counselor_assigned_at=now,
                        # Do NOT set converted_at — counselor will convert themselves
                    )
                    remember_first_assignment(lead_ids, counselor, now, 'counselor')
                    activities = [
                        LeadActivity(
                            lead_id=int(lid),
                            activity_type='ASSIGNED',
                            description=f"Contact assigned to Counselor {counselor.username} for telecalling by admin {request.user.username}.",
                            created_by=request.user
                        ) for lid in lead_ids
                    ]
                    LeadActivity.objects.bulk_create(activities)
                messages.success(request, f"Successfully assigned {len(lead_ids)} contact(s) to counselor {counselor.username} for telecalling.")
                return redirect(request.get_full_path())

        elif False and mode == 'counselor_telecalling_quantity':
            counselor_id = request.POST.get('counselor_telecalling_user')
            quantity = int(request.POST.get('quantity', 0) or 0)
            assign_source = request.POST.get('assign_source', 'unassigned')

            if not counselor_id or quantity <= 0:
                messages.error(request, "Invalid counselor or quantity.")
            else:
                counselor = get_object_or_404(User, pk=counselor_id, role='counselor', is_deleted=False, is_active=True)
                target_leads = leads if assign_source == 'filtered' else Lead.objects.filter(converted_at__isnull=True, assigned_telecaller__isnull=True, assigned_counselor__isnull=True)

                with transaction.atomic():
                    lead_ids = list(target_leads.select_for_update().values_list('id', flat=True)[:quantity])
                    if lead_ids:
                        now = timezone.now()
                        Lead.objects.filter(pk__in=lead_ids).update(
                            assigned_counselor=counselor,
                            assigned_telecaller=None,
                            assigned_by=request.user,
                            assigned_at=now,
                            counselor_assigned_at=now,
                        )
                        remember_first_assignment(lead_ids, counselor, now, 'counselor')
                        activities = [
                            LeadActivity(
                                lead_id=lid,
                                activity_type='ASSIGNED',
                                description=f"Contact assigned to Counselor {counselor.username} for telecalling by admin {request.user.username}.",
                                created_by=request.user
                            ) for lid in lead_ids
                        ]
                        LeadActivity.objects.bulk_create(activities)
                        messages.success(request, f"Successfully assigned {len(lead_ids)} contact(s) to counselor {counselor.username} for telecalling.")
                    else:
                        messages.warning(request, "No eligible contacts found to assign.")
                return redirect(request.get_full_path())

        elif False and mode == 'counselor_telecalling_all_filtered':
            counselor_id = request.POST.get('counselor_telecalling_user')
            if not counselor_id:
                messages.error(request, "Please select a counselor for telecalling assignment.")
            else:
                counselor = get_object_or_404(User, pk=counselor_id, role='counselor', is_deleted=False, is_active=True)
                with transaction.atomic():
                    lead_ids = list(leads.select_for_update().values_list('id', flat=True))
                    if lead_ids:
                        now = timezone.now()
                        Lead.objects.filter(pk__in=lead_ids).update(
                            assigned_counselor=counselor,
                            assigned_telecaller=None,
                            assigned_by=request.user,
                            assigned_at=now,
                            counselor_assigned_at=now,
                        )
                        remember_first_assignment(lead_ids, counselor, now, 'counselor')
                        activities = [
                            LeadActivity(
                                lead_id=lid,
                                activity_type='ASSIGNED',
                                description=f"Contact assigned to Counselor {counselor.username} for telecalling by admin {request.user.username}.",
                                created_by=request.user
                            ) for lid in lead_ids
                        ]
                        LeadActivity.objects.bulk_create(activities)
                        messages.success(request, f"Successfully assigned {len(lead_ids)} filtered contact(s) to counselor {counselor.username} for telecalling.")
                    else:
                        messages.warning(request, "No contacts matched the filters.")
                return redirect(request.get_full_path())

        elif False and mode == 'counselor_telecalling_auto_distribute':
            user_ids = request.POST.getlist('counselors')
            assign_source = request.POST.get('assign_source', 'unassigned')

            if not user_ids:
                messages.error(request, "Please select at least one counselor for telecalling distribution.")
            else:
                selected_users = list(User.objects.filter(id__in=user_ids, role='counselor', is_deleted=False, is_active=True))
                if not selected_users:
                    messages.error(request, "Invalid counselors selected.")
                else:
                    target_leads = leads if assign_source == 'filtered' else Lead.objects.filter(converted_at__isnull=True, assigned_telecaller__isnull=True, assigned_counselor__isnull=True)

                    with transaction.atomic():
                        lead_ids = list(target_leads.select_for_update().values_list('id', flat=True))
                        if not lead_ids:
                            messages.warning(request, "No eligible contacts found for counselor telecalling distribution.")
                        else:
                            pool_size = len(lead_ids)
                            num_users = len(selected_users)
                            workload_map = {w['id']: w['active'] for w in counselor_workload_data}
                            selected_users.sort(key=lambda u: workload_map.get(u.id, 0))

                            base_count = pool_size // num_users
                            remainder = pool_size % num_users

                            current_idx = 0
                            activities = []
                            for i, counselor in enumerate(selected_users):
                                count_for_counselor = base_count + (1 if i < remainder else 0)
                                if count_for_counselor > 0:
                                    chunk_ids = lead_ids[current_idx:current_idx+count_for_counselor]
                                    now = timezone.now()
                                    Lead.objects.filter(pk__in=chunk_ids).update(
                                        assigned_counselor=counselor,
                                        assigned_telecaller=None,
                                        assigned_by=request.user,
                                        assigned_at=now,
                                        counselor_assigned_at=now,
                                    )
                                    remember_first_assignment(chunk_ids, counselor, now, 'counselor')
                                    for lid in chunk_ids:
                                        activities.append(LeadActivity(
                                            lead_id=lid,
                                            activity_type='ASSIGNED',
                                            description=f"Contact assigned to Counselor {counselor.username} for telecalling via auto-distribution by admin.",
                                            created_by=request.user
                                        ))
                                    current_idx += count_for_counselor

                            LeadActivity.objects.bulk_create(activities)
                            messages.success(request, f"Successfully distributed {pool_size} contact(s) among {num_users} counselor(s) for telecalling.")
                    return redirect(request.get_full_path())

        elif mode == 'quantity':
            telecaller_id = request.POST.get('telecaller')
            quantity = int(request.POST.get('quantity', 0) or 0)
            assign_source = request.POST.get('assign_source', 'unassigned')

            if not telecaller_id or quantity <= 0:
                messages.error(request, "Invalid telecaller or quantity.")
            else:
                telecaller = get_object_or_404(User, pk=telecaller_id, is_deleted=False, is_active=True)
                target_leads = assignment_queue if assign_source == 'filtered' else initial_assignment_queue()

                with transaction.atomic():
                    lead_ids = list(target_leads.select_for_update().values_list('id', flat=True)[:quantity])
                    if lead_ids:
                        now = timezone.now()
                        Lead.objects.filter(pk__in=lead_ids).update(
                            assigned_telecaller=telecaller,
                            assigned_counselor=None,
                            assigned_by=request.user,
                            assigned_at=now,
                            telecaller_assigned_at=now,
                        )
                        remember_first_assignment(lead_ids, telecaller, now, 'telecaller')
                        activities = [
                            LeadActivity(
                                lead_id=lid,
                                activity_type='ASSIGNED',
                                description=f"Lead assigned to {telecaller.username} by admin {request.user.username}.",
                                created_by=request.user
                            ) for lid in lead_ids
                        ]
                        LeadActivity.objects.bulk_create(activities)
                        messages.success(request, f"Successfully assigned {len(lead_ids)} lead(s) to {telecaller.username}.")
                    else:
                        messages.warning(request, "No eligible leads found to assign.")
                return redirect(request.get_full_path())

        elif mode == 'all_filtered':
            telecaller_id = request.POST.get('telecaller')
            if not telecaller_id:
                messages.error(request, "Please select a telecaller.")
            else:
                telecaller = get_object_or_404(User, pk=telecaller_id, is_deleted=False, is_active=True)
                with transaction.atomic():
                    lead_ids = list(assignment_queue.select_for_update().values_list('id', flat=True))
                    if lead_ids:
                        now = timezone.now()
                        Lead.objects.filter(pk__in=lead_ids).update(
                            assigned_telecaller=telecaller,
                            assigned_counselor=None,
                            assigned_by=request.user,
                            assigned_at=now,
                            telecaller_assigned_at=now,
                        )
                        remember_first_assignment(lead_ids, telecaller, now, 'telecaller')
                        activities = [
                            LeadActivity(
                                lead_id=lid,
                                activity_type='ASSIGNED',
                                description=f"Lead assigned to {telecaller.username} by admin {request.user.username}.",
                                created_by=request.user
                            ) for lid in lead_ids
                        ]
                        LeadActivity.objects.bulk_create(activities)
                        messages.success(request, f"Successfully assigned {len(lead_ids)} lead(s) to {telecaller.username}.")
                    else:
                        messages.warning(request, "No leads matched the filters.")
                return redirect(request.get_full_path())

        elif mode == 'auto_distribute':
            user_ids = request.POST.getlist('telecallers')
            assign_source = request.POST.get('assign_source', 'unassigned')

            if not user_ids:
                messages.error(request, "Please select at least one telecaller for auto-distribution.")
            else:
                selected_users = list(User.objects.filter(id__in=user_ids, role='telecaller', is_active=True))
                if not selected_users:
                    messages.error(request, "Invalid telecallers selected.")
                else:
                    target_leads = assignment_queue if assign_source == 'filtered' else initial_assignment_queue()

                    with transaction.atomic():
                        lead_ids = list(target_leads.select_for_update().values_list('id', flat=True))
                        if not lead_ids:
                            messages.warning(request, "No eligible leads found for auto distribution.")
                        else:
                            pool_size = len(lead_ids)
                            num_users = len(selected_users)

                            # Sort users by current workload (active leads) ascending
                            # We can use the pre-calculated workload_data for sorting
                            workload_map = {w['id']: w['active'] for w in workload_data}
                            selected_users.sort(key=lambda u: workload_map.get(u.id, 0))

                            base_count = pool_size // num_users
                            remainder = pool_size % num_users

                            current_idx = 0
                            activities = []
                            for i, tc in enumerate(selected_users):
                                count_for_tc = base_count + (1 if i < remainder else 0)
                                if count_for_tc > 0:
                                    chunk_ids = lead_ids[current_idx:current_idx+count_for_tc]
                                    now = timezone.now()
                                    Lead.objects.filter(pk__in=chunk_ids).update(
                                        assigned_telecaller=tc,
                                        assigned_counselor=None,
                                        assigned_by=request.user,
                                        assigned_at=now,
                                        telecaller_assigned_at=now,
                                    )
                                    remember_first_assignment(chunk_ids, tc, now, 'telecaller')
                                    for lid in chunk_ids:
                                        activities.append(LeadActivity(
                                            lead_id=lid,
                                            activity_type='ASSIGNED',
                                            description=f"Lead assigned to {tc.username} via auto-distribution by admin.",
                                            created_by=request.user
                                        ))
                                    current_idx += count_for_tc

                            LeadActivity.objects.bulk_create(activities)
                            messages.success(request, f"Successfully distributed {pool_size} lead(s) among {num_users} telecaller(s).")
                    return redirect(request.get_full_path())

    paginator = Paginator(leads, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/telecaller_assignment.html', {
        'page_obj': page_obj,
        'telecallers': telecallers,
        'workload_data': workload_data,
        'q': q,
        'status': status,
        'assigned': assigned_status,
        'status_choices': Lead.STATUS_CHOICES,
    })


@login_required
@telecaller_required
def lead_bulk_action(request):
    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        lead_ids = request.POST.getlist('leads')

        action_map = {
            'Mark Contacted': 'Contacted',
            'Mark Interested': 'Interested',
            'Mark Follow Up': 'Follow Up',
            'Mark Qualified': 'Qualified',
            'Mark Rejected': 'Rejected',
            'Mark Invalid Number': 'Invalid Number'
        }

        if not action or action not in action_map:
            messages.error(request, "Invalid action selected.")
            return redirect('lead_list')

        if not lead_ids:
            messages.error(request, "No leads selected.")
            return redirect('lead_list')

        status_value = action_map[action]
        success_count = 0

        for lid in lead_ids:
            lead = get_object_or_404(Lead, pk=lid)
            if not is_admin_user(request.user) and lead.assigned_telecaller != request.user:
                continue

            old_status = lead.status
            lead.status = status_value
            lead.save()

            log_lead_activity(lead, 'STATUS_CHANGED', f"Status updated via bulk action from {old_status} to {status_value}.", request.user)
            success_count += 1

        messages.success(request, f"Successfully executed '{action}' on {success_count} lead(s).")
        return redirect('lead_list')

    if is_admin_user(request.user):
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_telecaller=request.user)

    paginator = Paginator(leads, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/lead_bulk_actions.html', {
        'page_obj': page_obj,
    })


@login_required
@telecaller_required
def reports_dashboard(request):
    sources = [choice[0] for choice in Inquiry.SOURCE_CHOICES]
    if is_admin_user(request.user):
        leads_qs = Lead.objects.select_related('inquiry')
    else:
        leads_qs = Lead.objects.filter(assigned_telecaller=request.user).select_related('inquiry')

    source_stats = []
    total_leads_all = leads_qs.count()
    total_qualified_all = leads_qs.filter(status='Qualified').count()
    total_rejected_all = leads_qs.filter(status='Rejected').count()

    for src in sources:
        source_leads = leads_qs.filter(inquiry__source=src)
        total_leads = source_leads.count()
        qualified_leads = source_leads.filter(status='Qualified').count()
        rejected_leads = source_leads.filter(status='Rejected').count()

        conversion_rate = 0.0
        if total_leads > 0:
            conversion_rate = round((qualified_leads / total_leads) * 100, 2)

        source_stats.append({
            'source': src,
            'total_leads': total_leads,
            'qualified_leads': qualified_leads,
            'rejected_leads': rejected_leads,
            'conversion_rate': conversion_rate,
            'share_rate': round((total_leads / total_leads_all) * 100, 2) if total_leads_all > 0 else 0.0,
        })

    chart_labels = sources
    source_distribution_data = [stat['total_leads'] for stat in source_stats]
    conversion_by_source_data = [stat['conversion_rate'] for stat in source_stats]
    best_source = max(source_stats, key=lambda item: item['conversion_rate']) if total_leads_all > 0 else None

    return render(request, 'management/lead_source_analytics.html', {
        'source_stats': source_stats,
        'chart_labels': chart_labels,
        'source_distribution_data': source_distribution_data,
        'conversion_by_source_data': conversion_by_source_data,
        'total_leads_all': total_leads_all,
        'total_qualified_all': total_qualified_all,
        'total_rejected_all': total_rejected_all,
        'overall_conversion_rate': round((total_qualified_all / total_leads_all) * 100, 2) if total_leads_all > 0 else 0.0,
        'best_source': best_source,
    })


@login_required
@telecaller_required
def telecaller_report(request):
    from django.contrib.auth import get_user_model
    User = get_user_model()

    if is_admin_user(request.user):
        telecallers = User.objects.filter(role='telecaller', is_deleted=False, is_active=True)
    else:
        telecallers = User.objects.filter(pk=request.user.pk, is_deleted=False, is_active=True)

    date_filter = request.GET.get('date', '').strip()

    report_data = []
    for tc in telecallers:
        inquiries_qs = Inquiry.objects.filter(created_by=tc)
        leads_qs = Lead.objects.filter(assigned_telecaller=tc)
        calls_qs = CallLog.objects.filter(created_by=tc)
        followups_qs = FollowUp.objects.filter(created_by=tc)

        if date_filter:
            inquiries_qs = inquiries_qs.filter(created_at__date=date_filter)
            leads_qs = leads_qs.filter(created_at__date=date_filter)
            calls_qs = calls_qs.filter(call_date__date=date_filter)
            followups_qs = followups_qs.filter(followup_date=date_filter)

        total_inquiries = inquiries_qs.count()
        total_leads = leads_qs.count()
        calls_made = calls_qs.count()
        followups_completed = followups_qs.filter(status='Completed').count()
        qualified_leads = leads_qs.filter(status='Qualified').count()
        rejected_leads = leads_qs.filter(status='Rejected').count()
        pending_followups = followups_qs.filter(status='Pending').count()

        # Call Status Metrics
        cs_accepted = inquiries_qs.filter(call_status='ACCEPTED').count()
        cs_busy = inquiries_qs.filter(call_status='BUSY').count()
        cs_call_back = inquiries_qs.filter(call_status='CALL_BACK').count()
        cs_interested = inquiries_qs.filter(call_status='INTERESTED').count()
        cs_not_interested = inquiries_qs.filter(call_status='NOT_INTERESTED').count()

        conversion_pct = 0.0
        if total_leads > 0:
            conversion_pct = round((qualified_leads / total_leads) * 100, 2)

        report_data.append({
            'telecaller': tc.username,
            'total_inquiries': total_inquiries,
            'total_leads': total_leads,
            'calls_made': calls_made,
            'followups_completed': followups_completed,
            'qualified_leads': qualified_leads,
            'rejected_leads': rejected_leads,
            'pending_followups': pending_followups,
            'conversion_pct': conversion_pct,
            'cs_accepted': cs_accepted,
            'cs_busy': cs_busy,
            'cs_call_back': cs_call_back,
            'cs_interested': cs_interested,
            'cs_not_interested': cs_not_interested,
        })

    totals = {
        'telecallers': len(report_data),
        'total_inquiries': sum(row['total_inquiries'] for row in report_data),
        'total_leads': sum(row['total_leads'] for row in report_data),
        'calls_made': sum(row['calls_made'] for row in report_data),
        'followups_completed': sum(row['followups_completed'] for row in report_data),
        'qualified_leads': sum(row['qualified_leads'] for row in report_data),
        'rejected_leads': sum(row['rejected_leads'] for row in report_data),
        'pending_followups': sum(row['pending_followups'] for row in report_data),
    }
    totals['conversion_pct'] = round((totals['qualified_leads'] / totals['total_leads']) * 100, 2) if totals['total_leads'] > 0 else 0.0

    export_format = request.GET.get('export', '').strip()
    if export_format == 'csv':
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="telecaller_performance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Telecaller', 'Total Inquiries', 'Total Leads', 'Calls Made', 'Follow-Ups Completed', 'Qualified Leads', 'Rejected Leads', 'Pending Follow-Ups', 'Conversion %', 'Accepted', 'Busy', 'Call Back', 'Interested', 'Not Interested'])
        for row in report_data:
            writer.writerow([
                row['telecaller'], row['total_inquiries'], row['total_leads'], row['calls_made'],
                row['followups_completed'], row['qualified_leads'], row['rejected_leads'],
                row['pending_followups'], row['conversion_pct'],
                row['cs_accepted'], row['cs_busy'], row['cs_call_back'],
                row['cs_interested'], row['cs_not_interested']
            ])
        return response

    elif export_format == 'excel':
        import openpyxl
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="telecaller_performance_report.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Report"
        ws.append(['Telecaller', 'Total Inquiries', 'Total Leads', 'Calls Made', 'Follow-Ups Completed', 'Qualified Leads', 'Rejected Leads', 'Pending Follow-Ups', 'Conversion %', 'Accepted', 'Busy', 'Call Back', 'Interested', 'Not Interested'])
        for row in report_data:
            ws.append([
                row['telecaller'], row['total_inquiries'], row['total_leads'], row['calls_made'],
                row['followups_completed'], row['qualified_leads'], row['rejected_leads'],
                row['pending_followups'], row['conversion_pct'],
                row['cs_accepted'], row['cs_busy'], row['cs_call_back'],
                row['cs_interested'], row['cs_not_interested']
            ])
        wb.save(response)
        return response

    return render(request, 'management/telecaller_report.html', {
        'report_data': report_data,
        'totals': totals,
        'date_filter': date_filter,
    })


# ==================================================
# COUNSELOR OPERATIONS (Phase 11.2)
# ==================================================

@login_required
@counselor_required
def counselor_dashboard(request):

    today = timezone.localdate()

    # Base Lead Queryset Scoping (Strict Isolation)
    if is_admin_user(request.user):
        leads_qs = Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=False)
        calls_qs = CallLog.objects.all()
        followups_qs = FollowUp.objects.filter(lead__assigned_counselor__isnull=False, lead__converted_at__isnull=False)
        activities_qs = LeadActivity.objects.filter(lead__assigned_counselor__isnull=False, lead__converted_at__isnull=False)
        visits_qs = VisitSheet.objects.all()
        admissions_qs = AdmissionSheet.objects.all()
    else:
        leads_qs = Lead.objects.filter(assigned_counselor=request.user, converted_at__isnull=False)
        calls_qs = CallLog.objects.filter(lead__assigned_counselor=request.user, lead__converted_at__isnull=False)
        followups_qs = FollowUp.objects.filter(lead__assigned_counselor=request.user, lead__converted_at__isnull=False)
        activities_qs = LeadActivity.objects.filter(lead__assigned_counselor=request.user, lead__converted_at__isnull=False)
        visits_qs = VisitSheet.objects.filter(counselor=request.user)
        admissions_qs = AdmissionSheet.objects.filter(counselor=request.user)

    # Counselor Status Counts
    total_assigned = {'today': leads_qs.filter(assigned_at__date=today).count(), 'total': leads_qs.count()}
    new_leads = {'today': leads_qs.filter(counselor_status='NEW', counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status='NEW').count()}
    contacted_leads = {'today': leads_qs.filter(counselor_status='CONTACTED', counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status='CONTACTED').count()}
    counseling_done = {'today': leads_qs.filter(counselor_status='COUNSELING_DONE', counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status='COUNSELING_DONE').count()}
    followup_req = {'today': leads_qs.filter(counselor_status='FOLLOW_UP_REQUIRED', counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status='FOLLOW_UP_REQUIRED').count()}
    interested_leads = {'today': leads_qs.filter(counselor_status='INTERESTED', counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status='INTERESTED').count()}
    converted_leads = {'today': leads_qs.filter(counselor_status__in=['CONVERTED', 'ADMISSION'], counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status__in=['CONVERTED', 'ADMISSION']).count()}
    not_interested = {'today': leads_qs.filter(counselor_status='NOT_INTERESTED', counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status='NOT_INTERESTED').count()}
    lost_leads = {'today': leads_qs.filter(counselor_status='LOST', counselor_status_updated_at__date=today).count(), 'total': leads_qs.filter(counselor_status='LOST').count()}

    # Followup statistics
    pending_followups = followups_qs.filter(status='Pending').count()
    overdue_followups = followups_qs.filter(status='Pending', followup_date__lt=today).count()

    today_visits = visits_qs.filter(visit_date=today).count()
    upcoming_visits = visits_qs.filter(visit_date__gt=today, status='Scheduled').count()
    completed_visits = visits_qs.filter(status__in=['Visited', 'Admission Done']).count()
    no_shows = visits_qs.filter(status='No Show').count()

    # Admission statistics
    admission_metrics = {
        'total': admissions_qs.count(),
        'confirmed': leads_qs.filter(counselor_status='ADMISSION').count(),
        'pending': admissions_qs.filter(admission_status='PENDING').count(),
        'cancelled': admissions_qs.filter(admission_status='CANCELLED').count(),
    }

    # Table Contexts
    recent_leads = leads_qs.select_related('inquiry', 'assigned_telecaller', 'assigned_counselor').order_by('-created_at')[:5]
    today_followups_list = followups_qs.filter(status='Pending', followup_date=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]
    overdue_followups_list = followups_qs.filter(status='Pending', followup_date__lt=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]
    recent_activities = activities_qs.select_related('lead__inquiry', 'created_by').order_by('-created_at')[:5]

    context = {
        'total_assigned': total_assigned,
        'new_leads': new_leads,
        'contacted_leads': contacted_leads,
        'counseling_done': counseling_done,
        'followup_req': followup_req,
        'interested_leads': interested_leads,
        'converted_leads': converted_leads,
        'not_interested': not_interested,
        'lost_leads': lost_leads,
        'pending_followups': pending_followups,
        'overdue_followups': overdue_followups,
        'recent_leads': recent_leads,
        'today_followups_list': today_followups_list,
        'overdue_followups_list': overdue_followups_list,
        'recent_activities': recent_activities,
        'today_visits': today_visits,
        'upcoming_visits': upcoming_visits,
        'completed_visits': completed_visits,
        'no_shows': no_shows,
        'admission_metrics': admission_metrics,
    }
    return render(request, 'management/counselor_dashboard.html', context)


@login_required
@counselor_required
def counselor_lead_list(request):
    if is_admin_user(request.user):
        leads = Lead.objects.filter(assigned_counselor__isnull=False, converted_at__isnull=False)
    else:
        leads = Lead.objects.filter(assigned_counselor=request.user, converted_at__isnull=False)

    # Search candidates
    q = request.GET.get('q', '').strip()
    if q:
        leads = leads.filter(
            Q(inquiry__full_name__icontains=q) | Q(inquiry__mobile_number__icontains=q)
        )

    # Filters
    status = request.GET.get('status', '').strip()
    if status:
        leads = leads.filter(counselor_status=status)

    priority = request.GET.get('priority', '').strip()
    if priority:
        leads = leads.filter(priority=priority)

    date_filter = request.GET.get('date', '').strip()
    selected_date = parse_filter_date(date_filter)
    if selected_date:
        leads = filter_datetime_field_by_local_date(leads, 'created_at', selected_date)
        date_filter = selected_date.isoformat()

    status_date_filter = request.GET.get('status_date', '').strip()
    selected_status_date = parse_filter_date(status_date_filter)
    if selected_status_date:
        leads = filter_datetime_field_by_local_date(leads, 'counselor_status_updated_at', selected_status_date)

    assigned_date_filter = request.GET.get('assigned_date', '').strip()
    selected_assigned_date = parse_filter_date(assigned_date_filter)
    if selected_assigned_date:
        leads = filter_datetime_field_by_local_date(leads, 'assigned_at', selected_assigned_date)

    paginator = Paginator(leads, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/counselor_lead_list.html', {
        'page_obj': page_obj,
        'q': q,
        'status': status,
        'priority': priority,
        'date_filter': date_filter,
        'status_choices': Lead.COUNSELOR_STATUS_CHOICES,
        'priority_choices': Lead.PRIORITY_CHOICES,
    })


@login_required
@counselor_required
def counselor_lead_detail(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    if not is_admin_user(request.user) and lead.assigned_counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this lead record.")

    sessions = lead.counseling_sessions.all().order_by('-session_date')
    followups = lead.followups.all().order_by('-followup_date')
    notes = lead.notes_timeline.all().select_related('created_by').order_by('-created_at')
    activities = lead.activities.all().order_by('-created_at')
    visits = lead.visit_sheets.all().order_by('-visit_date', '-visit_time')

    # Aggregate communication history
    communication_history = []
    for note in notes:
        communication_history.append({
            'type': 'Note',
            'timestamp': note.created_at,
            'content': note.note,
        })
    for call in lead.call_logs.all().order_by('-call_date'):
        communication_history.append({
            'type': 'Call',
            'timestamp': normalize_timestamp(call.call_date),
            'content': f"{call.call_status}: {call.remarks}",
        })
    for fu in followups:
        communication_history.append({
            'type': 'FollowUp',
            'timestamp': normalize_timestamp(fu.followup_date),
            'content': fu.response or fu.outcome or "",
        })
    communication_history.sort(key=lambda x: x['timestamp'], reverse=True)

    # Check admission sheet
    try:
        admission = lead.admission_sheet
    except AdmissionSheet.DoesNotExist:
        admission = None

    return render(request, 'management/counselor_lead_detail.html', {
        'lead': lead,
        'sessions': sessions,
        'followups': followups,
        'notes': notes,
        'activities': activities,
        'visits': visits,
        'admission': admission,
        'communication_history': communication_history,
    })


@login_required
@counselor_required
def counselor_lead_status_update(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    if not is_admin_user(request.user) and lead.assigned_counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this lead record.")

    if request.method == 'POST':
        form = CounselorLeadStatusForm(request.POST, instance=lead)
        if form.is_valid():
            old_status = lead.counselor_status
            updated_lead = form.save(commit=False)
            updated_lead.counselor_status_updated_at = timezone.now()
            updated_lead.save()

            if old_status != updated_lead.counselor_status:
                remark = updated_lead.notes.strip() if updated_lead.notes else ''
                activity_msg = f"Counselor status updated from {old_status} to {updated_lead.counselor_status}."
                if remark:
                    activity_msg += f" Remark: {remark}"
                log_lead_activity(
                    updated_lead,
                    'STATUS_CHANGED',
                    activity_msg,
                    request.user
                )

            messages.success(request, f"Counselor status updated successfully to {updated_lead.counselor_status}.")
            return redirect('counselor_lead_detail', pk=lead.pk)
    else:
        form = CounselorLeadStatusForm(instance=lead)
    return render(request, 'management/counselor_lead_status_update.html', {
        'form': form,
        'lead': lead,
    })


@login_required
@counselor_required
def counselor_mark_admission(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    if not is_admin_user(request.user) and lead.assigned_counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this lead record.")

    if request.method == 'POST':
        old_status = lead.counselor_status
        lead.counselor_status = 'ADMISSION'
        lead.counselor_status_updated_at = timezone.now()
        lead.save()

        if old_status != 'ADMISSION':
            log_lead_activity(
                lead,
                'STATUS_CHANGED',
                f"Counselor requested Admission. Status updated from {old_status} to ADMISSION.",
                request.user
            )

        messages.success(request, "Lead successfully marked for Admission. It has been sent to the Admin queue.")

    return redirect('counselor_lead_detail', pk=lead.pk)


@login_required
@counselor_required
def counselor_session_list(request):
    if is_admin_user(request.user):
        sessions = CounselingSession.objects.all()
    else:
        sessions = CounselingSession.objects.filter(counselor=request.user)

    date_filter = request.GET.get('date', '').strip()
    if date_filter:
        sessions = sessions.filter(session_date__date=date_filter)

    paginator = Paginator(sessions, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/counselor_session_list.html', {
        'page_obj': page_obj,
        'date_filter': date_filter,
    })


@login_required
@counselor_required
def counselor_session_add(request):
    lead_id = request.GET.get('lead_id')
    lead = None
    if lead_id:
        lead = get_object_or_404(Lead, pk=lead_id)
        if not is_admin_user(request.user) and lead.assigned_counselor != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

    if request.method == 'POST':
        form = CounselingSessionForm(request.POST)
        selected_lead_id = request.POST.get('lead')
        selected_lead = get_object_or_404(Lead, pk=selected_lead_id)

        if not is_admin_user(request.user) and selected_lead.assigned_counselor != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

        if form.is_valid():
            session = form.save(commit=False)
            session.lead = selected_lead
            session.counselor = request.user
            session.save()

            # Automatically transition lead status to COUNSELING_DONE
            if selected_lead.counselor_status != 'COUNSELING_DONE':
                old_status = selected_lead.counselor_status
                selected_lead.counselor_status = 'COUNSELING_DONE'
                selected_lead.counselor_status_updated_at = timezone.now()
                selected_lead.save()
                log_lead_activity(
                    selected_lead,
                    'STATUS_CHANGED',
                    f"Counselor status updated from {old_status} to COUNSELING_DONE upon recording session.",
                    request.user
                )

            log_lead_activity(
                selected_lead,
                'STATUS_CHANGED',
                f"Counseling session recorded by {request.user.username}.",
                request.user
            )
            messages.success(request, f"Counseling session recorded successfully for {selected_lead.inquiry.full_name}.")
            return redirect('counselor_lead_detail', pk=selected_lead.pk)
    else:
        form = CounselingSessionForm(initial={'lead': lead})

    if is_admin_user(request.user):
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_counselor=request.user)

    return render(request, 'management/counselor_session_form.html', {
        'form': form,
        'leads': leads,
        'selected_lead': lead,
    })


@login_required
@counselor_required
def counselor_session_detail(request, pk):
    session = get_object_or_404(CounselingSession, pk=pk)
    if not is_admin_user(request.user) and session.counselor != request.user:
        return HttpResponseForbidden("Access Denied: You did not record this counseling session.")

    return render(request, 'management/counselor_session_detail.html', {
        'session': session,
    })


@login_required
@counselor_required
def counselor_followup_list(request):
    if is_admin_user(request.user):
        followups = FollowUp.objects.all()
    else:
        followups = FollowUp.objects.filter(lead__assigned_counselor=request.user)

    status = request.GET.get('status', '').strip()
    if status:
        followups = followups.filter(status=status)

    date_filter = request.GET.get('date', '').strip()
    if date_filter:
        followups = followups.filter(followup_date=date_filter)

    paginator = Paginator(followups, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/counselor_followup_list.html', {
        'page_obj': page_obj,
        'status': status,
        'date_filter': date_filter,
    })


@login_required
@counselor_required
def counselor_followup_add(request):
    lead_id = request.GET.get('lead_id')
    lead = None
    if lead_id:
        lead = get_object_or_404(Lead, pk=lead_id)
        if not is_admin_user(request.user) and lead.assigned_counselor != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

    if request.method == 'POST':
        form = CounselorFollowUpForm(request.POST)
        selected_lead_id = request.POST.get('lead')
        selected_lead = get_object_or_404(Lead, pk=selected_lead_id)

        if not is_admin_user(request.user) and selected_lead.assigned_counselor != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead.")

        if form.is_valid():
            followup = form.save(commit=False)
            followup.lead = selected_lead
            followup.created_by = request.user
            followup.save()

            # Sync dates and log activity
            selected_lead.next_followup_date = followup.followup_date
            # Transition lead status to FOLLOW_UP_REQUIRED
            if selected_lead.counselor_status != 'FOLLOW_UP_REQUIRED':
                old_status = selected_lead.counselor_status
                selected_lead.counselor_status = 'FOLLOW_UP_REQUIRED'
                selected_lead.counselor_status_updated_at = timezone.now()
                log_lead_activity(
                    selected_lead,
                    'STATUS_CHANGED',
                    f"Counselor status updated from {old_status} to FOLLOW_UP_REQUIRED upon scheduling follow-up.",
                    request.user
                )
            selected_lead.save()

            log_lead_activity(selected_lead, 'FOLLOWUP_CREATED', f"Counselor follow-up scheduled for {followup.followup_date}.", request.user)
            messages.success(request, f"Follow-up scheduled successfully for {selected_lead.inquiry.full_name}.")
            return redirect('counselor_lead_detail', pk=selected_lead.pk)
    else:
        form = CounselorFollowUpForm(initial={'lead': lead})

    if is_admin_user(request.user):
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_counselor=request.user)

    return render(request, 'management/counselor_followup_form.html', {
        'form': form,
        'leads': leads,
        'selected_lead': lead,
    })


@login_required
@counselor_required
def counselor_followup_edit(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if not is_admin_user(request.user) and followup.lead.assigned_counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this followup record.")

    if request.method == 'POST':
        form = CounselorFollowUpForm(request.POST, instance=followup)
        if form.is_valid():
            fp = form.save()

            lead = fp.lead
            if fp.next_followup_date:
                lead.next_followup_date = fp.next_followup_date
            lead.save()

            messages.success(request, "Follow-up rescheduled successfully.")
            return redirect('counselor_lead_detail', pk=fp.lead.pk)
    else:
        form = CounselorFollowUpForm(instance=followup)

    return render(request, 'management/counselor_followup_form.html', {
        'form': form,
        'followup': followup,
        'selected_lead': followup.lead,
    })


@login_required
@counselor_required
def counselor_followup_complete(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if not is_admin_user(request.user) and followup.lead.assigned_counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    followup.status = 'Completed'
    followup.save()
    log_lead_activity(followup.lead, 'FOLLOWUP_COMPLETED', f"Follow-up scheduled on {followup.followup_date} completed by counselor.", request.user)
    messages.success(request, "Follow-up marked as Completed.")
    return redirect('counselor_lead_detail', pk=followup.lead.pk)


@login_required
@counselor_required
def counselor_followup_miss(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if not is_admin_user(request.user) and followup.lead.assigned_counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this record.")

    followup.status = 'Missed'
    followup.save()
    log_lead_activity(followup.lead, 'FOLLOWUP_COMPLETED', f"Follow-up scheduled on {followup.followup_date} marked as Missed by counselor.", request.user)
    messages.warning(request, "Follow-up marked as Missed.")
    return redirect('counselor_lead_detail', pk=followup.lead.pk)


@login_required
@counselor_required
def counselor_note_add(request, lead_pk):
    lead = get_object_or_404(Lead, pk=lead_pk)
    if not is_admin_user(request.user) and lead.assigned_counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this lead record.")

    if request.method == 'POST':
        note_text = request.POST.get('note', '').strip()
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', '')

        if note_text:
            note_obj = LeadNote.objects.create(
                lead=lead,
                note=note_text,
                created_by=request.user
            )
            log_lead_activity(lead, 'NOTE_ADDED', f"Counselor note added: '{note_text[:50]}...'", request.user)

            # Broadcast via Channels
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            from django.utils import timezone

            channel_layer = get_channel_layer()
            role_display = 'Telecaller' if request.user.role == 'telecaller' else ('Counselor' if request.user.role == 'counselor' else 'Admin')

            payload = {
                "note_id": note_obj.pk,
                "lead_id": lead.pk,
                "author_name": request.user.username,
                "author_role": role_display,
                "remark": note_text,
                "timestamp": timezone.localtime(note_obj.created_at).strftime("%d %b %Y %I:%M %p")
            }
            if channel_layer:
                try:
                    async_to_sync(channel_layer.group_send)(
                        f"lead_remarks_{lead.pk}",
                        {"type": "remark_message", "payload": payload}
                    )
                except Exception:
                    pass

            if is_ajax:
                return JsonResponse({"status": "success", "message": "Note added successfully."})

            messages.success(request, "Counselor note timeline updated successfully.")
        else:
            if is_ajax:
                return JsonResponse({"status": "error", "message": "Note content cannot be empty."}, status=400)
            messages.error(request, "Note content cannot be empty.")

    return redirect('counselor_lead_detail', pk=lead.pk)


@login_required
@counselor_required
def counselor_reports_dashboard(request):
    from django.contrib.auth import get_user_model
    User = get_user_model()

    if is_admin_user(request.user):
        counselors = User.objects.filter(role='counselor', is_deleted=False, is_active=True)
    else:
        counselors = User.objects.filter(pk=request.user.pk, is_deleted=False, is_active=True)

    date_filter = request.GET.get('date', '').strip()

    report_type = request.GET.get('report_type', 'performance').strip()
    report_data = []

    # Default values for visit metrics
    total_scheduled = 0
    total_completed = 0
    total_no_shows = 0
    total_admissions_visit = 0

    if report_type == 'visit':
        if is_admin_user(request.user):
            visits_qs = VisitSheet.objects.all()
        else:
            visits_qs = VisitSheet.objects.filter(counselor=request.user)

        if date_filter:
            visits_qs = visits_qs.filter(visit_date=date_filter)

        total_scheduled = visits_qs.filter(status='Scheduled').count()
        total_completed = visits_qs.filter(status='Visited').count()
        total_no_shows = visits_qs.filter(status='No Show').count()
        total_admissions_visit = visits_qs.filter(status='Admission Done').count()

        for v in visits_qs:
            report_data.append({
                'candidate': v.lead.inquiry.full_name,
                'mobile_number': v.lead.inquiry.mobile_number,
                'course': v.lead.inquiry.course_interest,
                'visit_date': v.visit_date.strftime('%Y-%m-%d'),
                'visit_time': v.visit_time.strftime('%H:%M'),
                'status': v.status,
                'counselor': v.counselor.username,
                'remarks': v.remarks or '-',
            })

    elif report_type == 'performance':
        for cs in counselors:
            leads_qs = Lead.objects.filter(assigned_counselor=cs)
            sessions_qs = CounselingSession.objects.filter(counselor=cs)
            followups_qs = FollowUp.objects.filter(created_by=cs)

            if date_filter:
                leads_qs = leads_qs.filter(created_at__date=date_filter)
                sessions_qs = sessions_qs.filter(created_at__date=date_filter)
                followups_qs = followups_qs.filter(created_at__date=date_filter)

            total_leads = leads_qs.count()
            sessions_conducted = sessions_qs.count()
            followups_completed = followups_qs.filter(status='Completed').count()
            converted = leads_qs.filter(counselor_status='CONVERTED').count()
            lost = leads_qs.filter(counselor_status='LOST').count()

            conversion_pct = 0.0
            if total_leads > 0:
                conversion_pct = round((converted / total_leads) * 100, 2)

            report_data.append({
                'counselor': cs.username,
                'total_leads': total_leads,
                'sessions_conducted': sessions_conducted,
                'followups_completed': followups_completed,
                'converted': converted,
                'lost': lost,
                'conversion_pct': conversion_pct,
            })

    elif report_type == 'conversion':
        if is_admin_user(request.user):
            leads = Lead.objects.all()
        else:
            leads = Lead.objects.filter(assigned_counselor=request.user)

        if date_filter:
            leads = leads.filter(created_at__date=date_filter)

        for l in leads:
            report_data.append({
                'candidate': l.inquiry.full_name,
                'counselor': l.assigned_counselor.username if l.assigned_counselor else '-',
                'status': l.counselor_status,
                'priority': l.priority,
                'date': l.created_at.strftime('%Y-%m-%d')
            })

    elif report_type == 'followup':
        if is_admin_user(request.user):
            followups = FollowUp.objects.all()
        else:
            followups = FollowUp.objects.filter(lead__assigned_counselor=request.user)

        if date_filter:
            followups = followups.filter(followup_date=date_filter)

        for fp in followups:
            report_data.append({
                'candidate': fp.lead.inquiry.full_name,
                'followup_date': fp.followup_date.strftime('%Y-%m-%d'),
                'status': fp.status,
                'outcome': fp.outcome or '-',
                'notes': fp.response or '-'
            })

    elif report_type == 'lost':
        if is_admin_user(request.user):
            leads = Lead.objects.filter(counselor_status='LOST')
        else:
            leads = Lead.objects.filter(assigned_counselor=request.user, counselor_status='LOST')

        if date_filter:
            leads = leads.filter(counselor_status_updated_at__date=date_filter)

        for l in leads:
            report_data.append({
                'candidate': l.inquiry.full_name,
                'counselor': l.assigned_counselor.username if l.assigned_counselor else '-',
                'reason': l.notes or 'No details provided.',
                'date': l.counselor_status_updated_at.strftime('%Y-%m-%d') if l.counselor_status_updated_at else '-'
            })

    export_format = request.GET.get('export', '').strip()
    if export_format in ('csv', 'excel') and report_type == 'visit':
        return HttpResponseForbidden("Exporting is not supported for visit reports.")

    if export_format == 'csv':
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="counselor_{report_type}_report.csv"'
        writer = csv.writer(response)

        if report_type == 'performance':
            writer.writerow(['Counselor', 'Total Leads', 'Sessions Conducted', 'Follow-Ups Completed', 'Converted Leads', 'Lost Leads', 'Conversion %'])
            for row in report_data:
                writer.writerow([row['counselor'], row['total_leads'], row['sessions_conducted'], row['followups_completed'], row['converted'], row['lost'], row['conversion_pct']])
        elif report_type == 'conversion':
            writer.writerow(['Candidate Name', 'Counselor', 'Status', 'Priority', 'Assigned Date'])
            for row in report_data:
                writer.writerow([row['candidate'], row['counselor'], row['status'], row['priority'], row['date']])
        elif report_type == 'followup':
            writer.writerow(['Candidate Name', 'Follow-up Date', 'Status', 'Outcome', 'Notes'])
            for row in report_data:
                writer.writerow([row['candidate'], row['followup_date'], row['status'], row['outcome'], row['notes']])
        elif report_type == 'lost':
            writer.writerow(['Candidate Name', 'Counselor', 'Notes / Reason', 'Lost Date'])
            for row in report_data:
                writer.writerow([row['candidate'], row['counselor'], row['reason'], row['date']])

        return response

    elif export_format == 'excel':
        import openpyxl
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="counselor_{report_type}_report.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.title()

        if report_type == 'performance':
            ws.append(['Counselor', 'Total Leads', 'Sessions Conducted', 'Follow-Ups Completed', 'Converted Leads', 'Lost Leads', 'Conversion %'])
            for row in report_data:
                ws.append([row['counselor'], row['total_leads'], row['sessions_conducted'], row['followups_completed'], row['converted'], row['lost'], row['conversion_pct']])
        elif report_type == 'conversion':
            ws.append(['Candidate Name', 'Counselor', 'Status', 'Priority', 'Assigned Date'])
            for row in report_data:
                ws.append([row['candidate'], row['counselor'], row['status'], row['priority'], row['date']])
        elif report_type == 'followup':
            ws.append(['Candidate Name', 'Follow-up Date', 'Status', 'Outcome', 'Notes'])
            for row in report_data:
                ws.append([row['candidate'], row['followup_date'], row['status'], row['outcome'], row['notes']])
        elif report_type == 'lost':
            ws.append(['Candidate Name', 'Counselor', 'Notes / Reason', 'Lost Date'])
            for row in report_data:
                ws.append([row['candidate'], row['counselor'], row['reason'], row['date']])

        wb.save(response)
        return response

    return render(request, 'management/counselor_reports.html', {
        'report_data': report_data,
        'report_type': report_type,
        'date_filter': date_filter,
        'total_scheduled': total_scheduled,
        'total_completed': total_completed,
        'total_no_shows': total_no_shows,
        'total_admissions_visit': total_admissions_visit,
    })


@login_required
@admin_required
def lead_assign_counselor(request):

    from django.contrib.auth import get_user_model
    User = get_user_model()
    counselors = User.objects.filter(role='counselor', is_deleted=False, is_active=True)
    assignment_type = (request.POST.get('assignment_type') or request.GET.get('assignment_type') or 'telecalling').strip()
    if assignment_type not in ('telecalling', 'counseling'):
        assignment_type = 'telecalling'
    is_telecalling_assignment = assignment_type == 'telecalling'

    # Calculate workload for dashboard
    from datetime import date
    from django.db.models import Count, Q
    from django.db import transaction
    from django.http import JsonResponse

    today = timezone.localdate()
    workload_data = []
    for c in counselors:
        workload_qs = Lead.objects.filter(assigned_counselor=c)
        if is_telecalling_assignment:
            workload_qs = workload_qs.filter(converted_at__isnull=True)
            active_filter = Q(converted_at__isnull=True)
        else:
            workload_qs = workload_qs.filter(converted_at__isnull=False)
            active_filter = Q(counselor_status__in=['NEW', 'CONTACTED', 'FOLLOW_UP_REQUIRED', 'INTERESTED'])

        stats = workload_qs.aggregate(
            active=Count('id', filter=active_filter),
            today=Count('id', filter=Q(assigned_at__date=today)),
            total=Count('id')
        )
        workload_data.append({
            'id': c.id,
            'username': c.username,
            'active': stats['active'],
            'today': stats['today'],
            'total': stats['total'],
        })

    if is_telecalling_assignment:
        # Counselor Telecalling Assignment:
        # Show admin-uploaded raw contacts in the first-assignment lifecycle.
        assignment_queue = initial_assignment_queue()
        leads = admin_uploaded_assignment_records()
    else:
        # Counselor Lead Assignment (Admin Queue):
        # Shows converted leads that need a counselor assignment.
        assignment_queue = None
        leads = Lead.objects.select_related('inquiry', 'assigned_telecaller', 'assigned_counselor').filter(
            converted_at__isnull=False
        )

    override_ready = request.GET.get('all_leads', 'no') == 'yes'

    q = request.GET.get('q', '').strip()
    if q:
        leads = leads.filter(
            Q(inquiry__full_name__icontains=q) | Q(inquiry__mobile_number__icontains=q)
        )
        if is_telecalling_assignment:
            assignment_queue = assignment_queue.filter(
                Q(inquiry__full_name__icontains=q) | Q(inquiry__mobile_number__icontains=q)
            )

    status = request.GET.get('status', '').strip()
    if status:
        if is_telecalling_assignment:
            leads = leads.filter(status=status)
            assignment_queue = assignment_queue.filter(status=status)
        else:
            leads = leads.filter(counselor_status=status)

    # Default: show records without the owner relevant to the selected assignment flow.
    assigned_status = request.GET.get('assigned', 'no').strip()
    if is_telecalling_assignment:
        if assigned_status == 'yes':
            leads = leads.filter(Q(assigned_telecaller__isnull=False) | Q(assigned_counselor__isnull=False))
        elif assigned_status == 'no':
            leads = leads.filter(assigned_telecaller__isnull=True, assigned_counselor__isnull=True)
        elif assigned_status == 'all':
            pass
    elif assigned_status == 'yes':
        leads = leads.filter(assigned_counselor__isnull=False)
    elif assigned_status == 'no':
        leads = leads.filter(assigned_counselor__isnull=True)
    elif assigned_status == 'all':
        pass

    # Handling AJAX APIs (Workload & Preview)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        action = request.POST.get('action') or request.GET.get('action')
        if action == 'workload':
            return JsonResponse({'workload': workload_data})

        if action == 'preview':
            mode = request.POST.get('mode', 'manual')
            quantity = int(request.POST.get('quantity', 0) or 0)
            assign_source = request.POST.get('assign_source', 'unassigned')

            if is_telecalling_assignment:
                target_leads = assignment_queue if assign_source == 'filtered' else initial_assignment_queue()
            else:
                target_leads = leads
                if assign_source == 'unassigned':
                    target_leads = Lead.objects.filter(
                        converted_at__isnull=False,
                        assigned_counselor__isnull=True,
                    )

            if mode == 'quantity':
                available = target_leads.count()
                will_assign = min(quantity, available)
                return JsonResponse({
                    'requested': quantity,
                    'eligible': available,
                    'will_assign': will_assign,
                    'remaining': available - will_assign
                })
            elif mode == 'all_filtered':
                available = assignment_queue.count() if is_telecalling_assignment else leads.count()
                return JsonResponse({
                    'requested': available,
                    'eligible': available,
                    'will_assign': available,
                    'remaining': 0
                })
            elif mode == 'auto_distribute':
                user_ids = request.POST.getlist('counselors') or request.POST.getlist('counselors[]')
                available = target_leads.count()
                return JsonResponse({
                    'requested': 'Auto',
                    'eligible': available,
                    'will_assign': available,
                    'remaining': 0,
                    'users_count': len(user_ids)
                })

    if request.method == 'POST':
        mode = request.POST.get('mode', 'manual')

        if mode == 'manual':
            counselor_id = request.POST.get('counselor')
            lead_ids = request.POST.getlist('leads')

            lead_id = request.POST.get('lead_id')
            if lead_id:
                lead_ids = [lead_id]

            if not counselor_id:
                messages.error(request, "Please select a counselor.")
            elif not lead_ids:
                messages.error(request, "Please select at least one lead.")
            else:
                counselor = get_object_or_404(User, pk=counselor_id, is_deleted=False, is_active=True)
                with transaction.atomic():
                    assigned_at = timezone.now()
                    if is_telecalling_assignment:
                        lead_ids = list(initial_assignment_queue().filter(pk__in=lead_ids).values_list('id', flat=True))
                    else:
                        lead_filter = {'pk__in': lead_ids, 'converted_at__isnull': False}
                        lead_ids = list(Lead.objects.filter(**lead_filter).values_list('id', flat=True))
                    update_fields = {
                        'assigned_counselor': counselor,
                        'assigned_by': request.user,
                        'assigned_at': assigned_at,
                        'counselor_assigned_at': assigned_at,
                    }
                    if is_telecalling_assignment:
                        update_fields['assigned_telecaller'] = None
                    Lead.objects.filter(pk__in=lead_ids).update(**update_fields)
                    remember_first_assignment(lead_ids, counselor, assigned_at, 'counselor')
                    activity_description = (
                        f"Contact assigned to Counselor {counselor.username} for telecalling by admin {request.user.username}."
                        if is_telecalling_assignment
                        else f"Lead assigned to Counselor {counselor.username} by admin {request.user.username}."
                    )
                    activities = [
                        LeadActivity(
                            lead_id=int(lid),
                            activity_type='ASSIGNED',
                            description=activity_description,
                            created_by=request.user
                        ) for lid in lead_ids
                    ]
                    LeadActivity.objects.bulk_create(activities)
                record_label = "contact(s) for telecalling" if is_telecalling_assignment else "lead(s)"
                messages.success(request, f"Successfully assigned {len(lead_ids)} {record_label} to Counselor {counselor.username}.")
                return redirect(request.get_full_path())

        elif mode == 'quantity':
            counselor_id = request.POST.get('counselor')
            quantity = int(request.POST.get('quantity', 0) or 0)
            assign_source = request.POST.get('assign_source', 'unassigned')

            if not counselor_id or quantity <= 0:
                messages.error(request, "Invalid counselor or quantity.")
            else:
                counselor = get_object_or_404(User, pk=counselor_id, is_deleted=False, is_active=True)
                if assign_source == 'filtered':
                    target_leads = assignment_queue if is_telecalling_assignment else leads
                elif is_telecalling_assignment:
                    target_leads = initial_assignment_queue()
                else:
                    target_leads = Lead.objects.filter(converted_at__isnull=False, assigned_counselor__isnull=True)

                with transaction.atomic():
                    lead_ids = list(target_leads.select_for_update().values_list('id', flat=True)[:quantity])
                    if lead_ids:
                        assigned_at = timezone.now()
                        update_fields = {
                            'assigned_counselor': counselor,
                            'assigned_by': request.user,
                            'assigned_at': assigned_at,
                            'counselor_assigned_at': assigned_at,
                        }
                        if is_telecalling_assignment:
                            update_fields['assigned_telecaller'] = None
                        Lead.objects.filter(pk__in=lead_ids).update(**update_fields)
                        remember_first_assignment(lead_ids, counselor, assigned_at, 'counselor')
                        activity_description = (
                            f"Contact assigned to Counselor {counselor.username} for telecalling by admin {request.user.username}."
                            if is_telecalling_assignment
                            else f"Lead assigned to Counselor {counselor.username} by admin {request.user.username}."
                        )
                        activities = [
                            LeadActivity(
                                lead_id=lid,
                                activity_type='ASSIGNED',
                                description=activity_description,
                                created_by=request.user
                            ) for lid in lead_ids
                        ]
                        LeadActivity.objects.bulk_create(activities)
                        record_label = "contact(s) for telecalling" if is_telecalling_assignment else "lead(s)"
                        messages.success(request, f"Successfully assigned {len(lead_ids)} {record_label} to Counselor {counselor.username}.")
                    else:
                        messages.warning(request, "No eligible records found to assign.")
                return redirect(request.get_full_path())

        elif mode == 'all_filtered':
            counselor_id = request.POST.get('counselor')
            if not counselor_id:
                messages.error(request, "Please select a counselor.")
            else:
                counselor = get_object_or_404(User, pk=counselor_id, is_deleted=False, is_active=True)
                with transaction.atomic():
                    source_leads = assignment_queue if is_telecalling_assignment else leads
                    lead_ids = list(source_leads.select_for_update().values_list('id', flat=True))
                    if lead_ids:
                        assigned_at = timezone.now()
                        update_fields = {
                            'assigned_counselor': counselor,
                            'assigned_by': request.user,
                            'assigned_at': assigned_at,
                            'counselor_assigned_at': assigned_at,
                        }
                        if is_telecalling_assignment:
                            update_fields['assigned_telecaller'] = None
                        Lead.objects.filter(pk__in=lead_ids).update(**update_fields)
                        remember_first_assignment(lead_ids, counselor, assigned_at, 'counselor')
                        activity_description = (
                            f"Contact assigned to Counselor {counselor.username} for telecalling by admin {request.user.username}."
                            if is_telecalling_assignment
                            else f"Lead assigned to Counselor {counselor.username} by admin {request.user.username}."
                        )
                        activities = [
                            LeadActivity(
                                lead_id=lid,
                                activity_type='ASSIGNED',
                                description=activity_description,
                                created_by=request.user
                            ) for lid in lead_ids
                        ]
                        LeadActivity.objects.bulk_create(activities)
                        record_label = "filtered contact(s) for telecalling" if is_telecalling_assignment else "lead(s)"
                        messages.success(request, f"Successfully assigned {len(lead_ids)} {record_label} to Counselor {counselor.username}.")
                    else:
                        messages.warning(request, "No records matched the filters.")
                return redirect(request.get_full_path())

        elif mode == 'auto_distribute':
            user_ids = request.POST.getlist('counselors')
            assign_source = request.POST.get('assign_source', 'unassigned')

            if not user_ids:
                messages.error(request, "Please select at least one counselor for auto-distribution.")
            else:
                selected_users = list(User.objects.filter(id__in=user_ids, role='counselor', is_active=True))
                if not selected_users:
                    messages.error(request, "Invalid counselors selected.")
                else:
                    if assign_source == 'filtered':
                        target_leads = assignment_queue if is_telecalling_assignment else leads
                    elif is_telecalling_assignment:
                        target_leads = initial_assignment_queue()
                    else:
                        target_leads = Lead.objects.filter(converted_at__isnull=False, assigned_counselor__isnull=True)

                    with transaction.atomic():
                        lead_ids = list(target_leads.select_for_update().values_list('id', flat=True))
                        if not lead_ids:
                            messages.warning(request, "No eligible records found for auto distribution.")
                        else:
                            pool_size = len(lead_ids)
                            num_users = len(selected_users)

                            workload_map = {w['id']: w['active'] for w in workload_data}
                            selected_users.sort(key=lambda u: workload_map.get(u.id, 0))

                            base_count = pool_size // num_users
                            remainder = pool_size % num_users

                            current_idx = 0
                            activities = []
                            for i, c in enumerate(selected_users):
                                count_for_c = base_count + (1 if i < remainder else 0)
                                if count_for_c > 0:
                                    chunk_ids = lead_ids[current_idx:current_idx+count_for_c]
                                    assigned_at = timezone.now()
                                    update_fields = {
                                        'assigned_counselor': c,
                                        'assigned_by': request.user,
                                        'assigned_at': assigned_at,
                                        'counselor_assigned_at': assigned_at,
                                    }
                                    if is_telecalling_assignment:
                                        update_fields['assigned_telecaller'] = None
                                    Lead.objects.filter(pk__in=chunk_ids).update(**update_fields)
                                    remember_first_assignment(chunk_ids, c, assigned_at, 'counselor')
                                    activity_description = (
                                        f"Contact assigned to Counselor {c.username} for telecalling via auto-distribution by admin."
                                        if is_telecalling_assignment
                                        else f"Lead assigned to Counselor {c.username} via auto-distribution by admin."
                                    )
                                    for lid in chunk_ids:
                                        activities.append(LeadActivity(
                                            lead_id=lid,
                                            activity_type='ASSIGNED',
                                            description=activity_description,
                                            created_by=request.user
                                        ))
                                    current_idx += count_for_c

                            LeadActivity.objects.bulk_create(activities)
                            record_label = "contact(s) for telecalling" if is_telecalling_assignment else "lead(s)"
                            messages.success(request, f"Successfully distributed {pool_size} {record_label} among {num_users} counselor(s).")
                    return redirect(request.get_full_path())

    paginator = Paginator(leads, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/counselor_assignment.html', {
        'page_obj': page_obj,
        'counselors': counselors,
        'workload_data': workload_data,
        'q': q,
        'status': status,
        'assigned': assigned_status,
        'assignment_type': assignment_type,
        'all_leads': 'yes' if override_ready else 'no',
        'status_choices': Lead.STATUS_CHOICES if is_telecalling_assignment else Lead.COUNSELOR_STATUS_CHOICES,
    })


@login_required
@counselor_required
def counselor_visit_list(request):
    if is_admin_user(request.user):
        visits = VisitSheet.objects.all()
    elif request.user.role == 'counselor':
        visits = VisitSheet.objects.filter(counselor=request.user)
    else:
        return HttpResponseForbidden("Access Denied.")

    # Search candidates
    q = request.GET.get('q', '').strip()
    if q:
        visits = visits.filter(
            Q(lead__inquiry__full_name__icontains=q) |
            Q(lead__inquiry__mobile_number__icontains=q) |
            Q(lead__inquiry__course_interest__icontains=q)
        )

    # Filters
    status = request.GET.get('status', '').strip()
    if status:
        visits = visits.filter(status=status)

    date_filter = request.GET.get('date', '').strip()
    if date_filter:
        try:
            visits = visits.filter(visit_date=datetime.strptime(date_filter, "%Y-%m-%d").date())
        except ValueError:
            pass

    paginator = Paginator(visits, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/counselor_visit_list.html', {
        'page_obj': page_obj,
        'q': q,
        'status': status,
        'date_filter': date_filter,
        'status_choices': VisitSheet.STATUS_CHOICES,
    })


@login_required
@counselor_required
def counselor_visit_add(request):
    lead_id = request.GET.get('lead_id')
    lead = None
    if lead_id:
        lead = get_object_or_404(Lead, pk=lead_id)
        if not is_admin_user(request.user) and lead.assigned_counselor != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead record.")

    if request.method == 'POST':
        form = VisitSheetForm(request.POST, user=request.user)
        selected_lead_id = request.POST.get('lead')
        selected_lead = get_object_or_404(Lead, pk=selected_lead_id)

        if not is_admin_user(request.user) and selected_lead.assigned_counselor != request.user:
            return HttpResponseForbidden("Access Denied: You do not own this lead record.")

        if form.is_valid():
            visit = form.save(commit=False)
            visit.lead = selected_lead
            visit.counselor = selected_lead.assigned_counselor or request.user
            visit.created_by = request.user
            visit.save()

            # Log activity
            log_lead_activity(
                selected_lead,
                'NOTE_ADDED',
                f"Counselor scheduled candidate visit for {visit.visit_date} at {visit.visit_time}.",
                request.user
            )

            messages.success(request, f"Visit scheduled successfully for {selected_lead.inquiry.full_name}.")
            return redirect('counselor_lead_detail', pk=selected_lead.pk)
    else:
        form = VisitSheetForm(initial={'lead': lead}, user=request.user)

    if is_admin_user(request.user):
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_counselor=request.user)

    return render(request, 'management/counselor_visit_form.html', {
        'form': form,
        'leads': leads,
        'selected_lead': lead,
        'title': 'Schedule Visit',
    })


@login_required
@counselor_required
def counselor_visit_edit(request, pk):
    visit = get_object_or_404(VisitSheet, pk=pk)
    if not is_admin_user(request.user) and visit.counselor != request.user:
        return HttpResponseForbidden("Access Denied: You do not own this visit record.")

    if request.method == 'POST':
        form = VisitSheetForm(request.POST, instance=visit, user=request.user)
        if form.is_valid():
            old_status = visit.status
            updated_visit = form.save()

            if old_status != updated_visit.status:
                log_lead_activity(
                    updated_visit.lead,
                    'STATUS_CHANGED',
                    f"Visit status updated from {old_status} to {updated_visit.status}.",
                    request.user
                )

            messages.success(request, "Visit sheet updated successfully.")
            return redirect('counselor_lead_detail', pk=updated_visit.lead.pk)
    else:
        form = VisitSheetForm(instance=visit, user=request.user)

    return render(request, 'management/counselor_visit_form.html', {
        'form': form,
        'visit': visit,
        'selected_lead': visit.lead,
        'title': 'Edit Visit Sheet',
    })


# ==================================================
# PHASE 11.5 — ADMISSION SHEET MANAGEMENT
# ==================================================

@login_required
@admin_required
def admission_list(request):
    """List all admissions with search, filters, and pagination."""
    admissions = AdmissionSheet.objects.all()

    # Search
    q = request.GET.get('q', '').strip()
    if q:
        admissions = admissions.filter(
            Q(student_name__icontains=q) | Q(mobile_number__icontains=q) | Q(admission_number__icontains=q)
        )

    # Filters
    status = request.GET.get('status', '').strip()
    if status:
        admissions = admissions.filter(admission_status=status)

    counselor_id = request.GET.get('counselor', '').strip()
    if counselor_id:
        admissions = admissions.filter(counselor_id=counselor_id)

    date_filter = request.GET.get('date', '').strip()
    if date_filter:
        admissions = admissions.filter(admission_date=date_filter)

    # Counselor choices for filter (admin only)
    from django.contrib.auth import get_user_model
    User = get_user_model()
    counselor_choices = User.objects.filter(role='counselor', is_deleted=False, is_active=True).order_by('username')

    paginator = Paginator(admissions, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'management/admission_list.html', {
        'page_obj': page_obj,
        'q': q,
        'status': status,
        'counselor_id': counselor_id,
        'date_filter': date_filter,
        'status_choices': AdmissionSheet.ADMISSION_STATUS_CHOICES,
        'counselor_choices': counselor_choices,
    })


@login_required
@admin_required
def admission_create(request, lead_pk):
    """Create admission from a lead. Auto-populates student info."""
    lead = get_object_or_404(Lead, pk=lead_pk)

    # Check duplicate
    if AdmissionSheet.objects.filter(lead=lead).exists():
        messages.error(request, "Admission Sheet already exists for this lead.")
        return redirect('counselor_lead_detail', pk=lead.pk)

    if request.method == 'POST':
        form = AdmissionSheetForm(request.POST, user=request.user)
        if form.is_valid():
            admission = form.save(commit=False)
            admission.lead = lead
            admission.created_by = request.user
            if not admission.counselor:
                admission.counselor = lead.assigned_counselor or request.user
            admission.save()

            # Update counselor status to Converted
            lead.counselor_status = 'CONVERTED'
            lead.counselor_status_updated_at = timezone.now()
            lead.save()

            log_lead_activity(
                lead,
                'STATUS_CHANGED',
                f"Admission Sheet created: {admission.admission_number}.",
                request.user
            )

            messages.success(request, f"Admission {admission.admission_number} created successfully.")
            return redirect('admission_detail', pk=admission.pk)
    else:
        # Auto-populate from lead
        inquiry = lead.inquiry
        initial = {
            'student_name': inquiry.full_name,
            'mobile_number': inquiry.mobile_number,
            'email_id': inquiry.email or '',
            'college_name': '',
            'department': '',
            'course_name': inquiry.course_interest or '',
            'admission_date': date.today(),
        }
        form = AdmissionSheetForm(initial=initial, user=request.user)

    return render(request, 'management/admission_form.html', {
        'form': form,
        'lead': lead,
        'title': 'Create Admission',
    })


@login_required
@admin_required
def admission_edit(request, pk):
    """Edit an existing admission."""
    admission = get_object_or_404(AdmissionSheet, pk=pk)

    if request.method == 'POST':
        form = AdmissionSheetForm(request.POST, instance=admission, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Admission {admission.admission_number} updated successfully.")
            return redirect('admission_detail', pk=admission.pk)
    else:
        form = AdmissionSheetForm(instance=admission, user=request.user)

    return render(request, 'management/admission_form.html', {
        'form': form,
        'lead': admission.lead,
        'admission': admission,
        'title': 'Edit Admission',
    })


@login_required
@admin_required
def admission_detail(request, pk):
    """View admission details."""
    admission = get_object_or_404(AdmissionSheet, pk=pk)

    return render(request, 'management/admission_detail.html', {
        'admission': admission,
    })


@login_required
@admin_required
def admission_report(request):
    """Simple admission report with tabular data."""
    admissions = AdmissionSheet.objects.all()

    # Aggregate metrics
    from django.db.models import Count
    total_admissions = admissions.count()

    # By counselor
    by_counselor = admissions.values(
        'counselor__username'
    ).annotate(
        count=Count('id')
    ).order_by('-count')

    # By course
    by_course = admissions.values(
        'course_name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')

    return render(request, 'management/admission_report.html', {
        'total_admissions': total_admissions,
        'by_counselor': by_counselor,
        'by_course': by_course,
    })


# ============================================================
# PHASE 2 ENTERPRISE ENHANCEMENT VIEWS
# ============================================================

def _get_date_range(request):
    """Helper: parse date_filter from request.GET and return (start_date, end_date)."""
    today = timezone.localdate()
    date_filter = request.GET.get('date_filter', 'this_month')
    start_date, end_date = None, today

    if date_filter == 'today':
        start_date = today
    elif date_filter == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = today - timedelta(days=1)
    elif date_filter == 'this_week':
        start_date = today - timedelta(days=today.weekday())
    elif date_filter == 'this_month':
        start_date = today.replace(day=1)
    elif date_filter == 'this_quarter':
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=quarter_start_month, day=1)
    elif date_filter == 'this_year':
        start_date = today.replace(month=1, day=1)
    elif date_filter == 'custom':
        try:
            s = request.GET.get('start_date', '')
            e = request.GET.get('end_date', '')
            if s: start_date = datetime.strptime(s, "%Y-%m-%d").date()
            if e: end_date = datetime.strptime(e, "%Y-%m-%d").date()
        except ValueError:
            pass
    else:
        start_date = today.replace(day=1)  # default to this_month

    PERIOD_LABELS = {
        'today': 'Today',
        'yesterday': 'Yesterday',
        'this_week': 'This Week',
        'this_month': 'This Month',
        'this_quarter': 'This Quarter',
        'this_year': 'This Year',
        'custom': 'Custom Range',
    }
    period_label = PERIOD_LABELS.get(date_filter, 'This Month')
    return start_date, end_date, date_filter, period_label


def _compute_telecaller_board(start_date, end_date, date_filter):
    """Compute leaderboard data for all telecallers."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    telecallers = User.objects.filter(role='telecaller', is_deleted=False, is_active=True)
    board = []
    for tc in telecallers:
        inq_qs = Inquiry.objects.filter(created_by=tc)
        leads_qs = Lead.objects.filter(assigned_telecaller=tc)
        calls_qs = CallLog.objects.filter(created_by=tc)
        followups_qs = FollowUp.objects.filter(created_by=tc)
        if start_date:
            inq_qs = inq_qs.filter(created_at__date__gte=start_date)
            leads_qs = leads_qs.filter(created_at__date__gte=start_date)
            calls_qs = calls_qs.filter(call_date__date__gte=start_date)
            followups_qs = followups_qs.filter(followup_date__gte=start_date)
        if end_date and date_filter not in ('all', ''):
            inq_qs = inq_qs.filter(created_at__date__lte=end_date)
            leads_qs = leads_qs.filter(created_at__date__lte=end_date)
            calls_qs = calls_qs.filter(call_date__date__lte=end_date)
            followups_qs = followups_qs.filter(followup_date__lte=end_date)

        total_inquiries = inq_qs.count()
        total_leads = leads_qs.count()
        calls_made = calls_qs.count()
        followups_completed = followups_qs.filter(status='Completed').count()
        qualified_leads = leads_qs.filter(status='Qualified').count()
        rejected_leads = leads_qs.filter(status='Rejected').count()
        pending_followups = followups_qs.filter(status='Pending').count()
        conversion_pct = round((qualified_leads / total_leads * 100), 2) if total_leads > 0 else 0.0
        # Weighted performance score
        score = (total_inquiries * 1.0 + calls_made * 0.5 + followups_completed * 2.0 + qualified_leads * 5.0)
        board.append({
            'telecaller': tc.username,
            'total_inquiries': total_inquiries,
            'total_leads': total_leads,
            'calls_made': calls_made,
            'followups_completed': followups_completed,
            'qualified_leads': qualified_leads,
            'rejected_leads': rejected_leads,
            'pending_followups': pending_followups,
            'conversion_pct': conversion_pct,
            'score': score,
        })
    return sorted(board, key=lambda x: x['score'], reverse=True)


def _compute_counselor_board(start_date, end_date, date_filter):
    """Compute leaderboard data for all counselors."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    counselors = User.objects.filter(role='counselor', is_deleted=False, is_active=True)
    board = []
    for c in counselors:
        leads_qs = Lead.objects.filter(assigned_counselor=c)
        sessions_qs = CounselingSession.objects.filter(counselor=c)
        followups_qs = FollowUp.objects.filter(lead__assigned_counselor=c)
        admissions_qs = AdmissionSheet.objects.filter(counselor=c)
        visits_qs = VisitSheet.objects.filter(counselor=c)
        if start_date:
            sessions_qs = sessions_qs.filter(session_date__date__gte=start_date)
            admissions_qs = admissions_qs.filter(admission_date__gte=start_date)
            visits_qs = visits_qs.filter(visit_date__gte=start_date)
        if end_date and date_filter not in ('all', ''):
            sessions_qs = sessions_qs.filter(session_date__date__lte=end_date)
            admissions_qs = admissions_qs.filter(admission_date__lte=end_date)
            visits_qs = visits_qs.filter(visit_date__lte=end_date)

        total_leads = leads_qs.count()
        sessions = sessions_qs.count()
        followups_completed = followups_qs.filter(status='Completed').count()
        admissions = admissions_qs.filter(admission_status='CONFIRMED').count()
        visits = visits_qs.count()
        conversion_pct = round((admissions / total_leads * 100), 2) if total_leads > 0 else 0.0
        score = (sessions * 3.0 + admissions * 10.0 + followups_completed * 2.0)
        board.append({
            'counselor': c.username,
            'total_leads': total_leads,
            'sessions': sessions,
            'followups_completed': followups_completed,
            'admissions': admissions,
            'visits': visits,
            'conversion_pct': conversion_pct,
            'score': score,
        })
    return sorted(board, key=lambda x: x['score'], reverse=True)


@login_required
def leaderboard(request):
    """Enterprise leaderboard for Telecallers and Counselors."""
    if request.user.role != 'admin':
        return HttpResponseForbidden("Access Denied: Admin only.")

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    telecaller_board = _compute_telecaller_board(start_date, end_date, date_filter)
    counselor_board = _compute_counselor_board(start_date, end_date, date_filter)

    # Search filter
    q = request.GET.get('q', '').strip().lower()
    if q:
        telecaller_board = [x for x in telecaller_board if q in x['telecaller'].lower()]
        counselor_board = [x for x in counselor_board if q in x['counselor'].lower()]

    # Pagination
    from django.core.paginator import Paginator
    tc_page = Paginator(telecaller_board, 10).get_page(request.GET.get('t_page'))
    c_page = Paginator(counselor_board, 10).get_page(request.GET.get('c_page'))

    return render(request, 'management/leaderboard.html', {
        'telecaller_board': tc_page,
        'counselor_board': c_page,
        'date_filter': date_filter,
        'period_label': period_label,
        'q': q,
    })


@login_required
def user_performance(request):
    """User Performance Center with per-user metrics and department comparison."""
    if request.user.role != 'admin':
        return HttpResponseForbidden("Access Denied: Admin only.")

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    telecaller_report_data = _compute_telecaller_board(start_date, end_date, date_filter)
    counselor_report_data = _compute_counselor_board(start_date, end_date, date_filter)

    avg_telecaller_score = sum(r['score'] for r in telecaller_report_data) / len(telecaller_report_data) if telecaller_report_data else 0
    avg_counselor_score = sum(r['score'] for r in counselor_report_data) / len(counselor_report_data) if counselor_report_data else 0

    # Search filter
    q = request.GET.get('q', '').strip().lower()
    if q:
        telecaller_report_data = [x for x in telecaller_report_data if q in x['telecaller'].lower()]
        counselor_report_data = [x for x in counselor_report_data if q in x['counselor'].lower()]

    # Pagination
    from django.core.paginator import Paginator
    tc_page = Paginator(telecaller_report_data, 10).get_page(request.GET.get('t_page'))
    c_page = Paginator(counselor_report_data, 10).get_page(request.GET.get('c_page'))

    return render(request, 'management/user_performance.html', {
        'telecaller_report': tc_page,
        'counselor_report': c_page,
        'total_telecallers': len(telecaller_report_data),
        'total_counselors': len(counselor_report_data),
        'avg_telecaller_score': avg_telecaller_score,
        'avg_counselor_score': avg_counselor_score,
        'date_filter': date_filter,
        'period_label': period_label,
    })


@login_required
def executive_reports(request):
    """Executive Reports Dashboard — high-level business KPIs."""
    if request.user.role != 'admin':
        return HttpResponseForbidden("Access Denied: Admin only.")

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    today = timezone.localdate()

    # Base querysets
    inq_qs = Inquiry.objects.all()
    leads_qs = Lead.objects.all()
    calls_qs = CallLog.objects.all()
    admissions_qs = AdmissionSheet.objects.filter(admission_status='CONFIRMED')

    if start_date:
        inq_qs = inq_qs.filter(created_at__date__gte=start_date)
        leads_qs = leads_qs.filter(created_at__date__gte=start_date)
        calls_qs = calls_qs.filter(call_date__date__gte=start_date)
        admissions_qs = admissions_qs.filter(admission_date__gte=start_date)
    if end_date and date_filter not in ('all', ''):
        inq_qs = inq_qs.filter(created_at__date__lte=end_date)
        leads_qs = leads_qs.filter(created_at__date__lte=end_date)
        calls_qs = calls_qs.filter(call_date__date__lte=end_date)
        admissions_qs = admissions_qs.filter(admission_date__lte=end_date)

    total_inquiries = inq_qs.count()
    total_leads = leads_qs.count()
    total_calls = calls_qs.count()
    total_admissions = admissions_qs.count()
    total_overdue = FollowUp.objects.filter(status='Pending', followup_date__lt=today).count()
    conversion_rate = round(total_admissions / total_leads * 100, 2) if total_leads > 0 else 0

    # Pipeline funnel
    pipeline = [
        {'label': 'Total Leads', 'count': total_leads, 'color': '#5b5fef'},
        {'label': 'Contacted', 'count': leads_qs.filter(status='Contacted').count(), 'color': '#38bdf8'},
        {'label': 'Interested', 'count': leads_qs.filter(status='Interested').count(), 'color': '#818cf8'},
        {'label': 'Qualified', 'count': leads_qs.filter(status='Qualified').count(), 'color': '#34d399'},
        {'label': 'Admissions', 'count': total_admissions, 'color': '#10b981'},
    ]

    # Source distribution
    source_distribution = inq_qs.values('source').annotate(count=Count('id')).order_by('-count')[:8]

    # Call status breakdown
    call_breakdown = [
        {'label': 'Accepted', 'count': inq_qs.filter(call_status='ACCEPTED').count()},
        {'label': 'Busy', 'count': inq_qs.filter(call_status='BUSY').count()},
        {'label': 'Call Back', 'count': inq_qs.filter(call_status='CALL_BACK').count()},
        {'label': 'Interested', 'count': inq_qs.filter(call_status='INTERESTED').count()},
        {'label': 'Not Interested', 'count': inq_qs.filter(call_status='NOT_INTERESTED').count()},
        {'label': 'No Answer', 'count': inq_qs.filter(call_status='NO_ANSWER').count()},
    ]

    top_telecallers = _compute_telecaller_board(start_date, end_date, date_filter)[:5]
    top_counselors = _compute_counselor_board(start_date, end_date, date_filter)[:5]

    return render(request, 'management/executive_reports.html', {
        'total_inquiries': total_inquiries,
        'total_leads': total_leads,
        'total_calls': total_calls,
        'total_admissions': total_admissions,
        'total_overdue': total_overdue,
        'conversion_rate': conversion_rate,
        'pipeline': pipeline,
        'source_distribution': source_distribution,
        'call_breakdown': call_breakdown,
        'top_telecallers': top_telecallers,
        'top_counselors': top_counselors,
        'date_filter': date_filter,
        'period_label': period_label,
    })


@login_required
def global_search(request):
    """Global search returning JSON results for the omnibar."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 2:
        return JsonResponse({'inquiries': [], 'leads': [], 'admissions': []})

    # Scope: admins see all, others see their own
    if request.user.role == 'admin':
        inquiries = Inquiry.objects.filter(Q(full_name__icontains=q) | Q(mobile_number__icontains=q) | Q(email__icontains=q))[:6]
        leads = Lead.objects.filter(Q(inquiry__full_name__icontains=q) | Q(inquiry__mobile_number__icontains=q)).select_related('inquiry')[:6]
        admissions = AdmissionSheet.objects.filter(Q(student_name__icontains=q) | Q(mobile_number__icontains=q) | Q(admission_number__icontains=q))[:4]
    elif request.user.role == 'telecaller':
        inquiries = Inquiry.objects.filter(created_by=request.user).filter(Q(full_name__icontains=q) | Q(mobile_number__icontains=q))[:6]
        leads = Lead.objects.filter(assigned_telecaller=request.user).filter(Q(inquiry__full_name__icontains=q)).select_related('inquiry')[:6]
        admissions = []
    else:
        inquiries = []
        leads = Lead.objects.filter(assigned_counselor=request.user).filter(Q(inquiry__full_name__icontains=q)).select_related('inquiry')[:6]
        admissions = AdmissionSheet.objects.filter(counselor=request.user).filter(Q(student_name__icontains=q))[:4]

    return JsonResponse({
        'inquiries': [{'id': i.pk, 'name': i.full_name, 'phone': i.mobile_number, 'status': i.status} for i in inquiries],
        'leads': [{'id': l.pk, 'name': l.inquiry.full_name, 'status': l.status, 'priority': l.priority} for l in leads],
        'admissions': [{'id': a.pk, 'name': a.student_name, 'number': a.admission_number, 'status': a.admission_status} for a in admissions],
    })


# ─── CSV Export Views ───────────────────────────────────────

@login_required
def export_leads_csv(request):
    """Export leads to CSV, respecting date filter."""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    qs = Lead.objects.select_related('inquiry', 'assigned_telecaller').all()
    if start_date:
        qs = qs.filter(created_at__date__gte=start_date)
    if end_date and date_filter not in ('all', ''):
        qs = qs.filter(created_at__date__lte=end_date)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="leads_{period_label.replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Lead Name', 'Mobile', 'Email', 'Status', 'Priority', 'Assigned Telecaller', 'Created At'])
    for lead in qs:
        writer.writerow([
            lead.inquiry.full_name, lead.inquiry.mobile_number,
            lead.inquiry.email or '', lead.status, lead.priority,
            lead.assigned_telecaller.username if lead.assigned_telecaller else '',
            lead.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    # Track in session
    history = request.session.get('export_history', [])
    history.insert(0, {'name': f'Leads — {period_label}', 'type': 'csv', 'time': datetime.now().strftime('%I:%M %p')})
    request.session['export_history'] = history[:5]
    return response


@login_required
def export_telecaller_csv(request):
    """Export telecaller performance report to CSV."""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    board = _compute_telecaller_board(start_date, end_date, date_filter)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="telecaller_performance_{period_label.replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Rank', 'Telecaller', 'Inquiries', 'Calls Made', 'Follow-ups Done', 'Qualified', 'Rejected', 'Pending F/U', 'Conversion %', 'Score'])
    for i, row in enumerate(board, 1):
        writer.writerow([i, row['telecaller'], row['total_inquiries'], row['calls_made'], row['followups_completed'],
                         row['qualified_leads'], row['rejected_leads'], row['pending_followups'],
                         f"{row['conversion_pct']}%", f"{row['score']:.1f}"])
    history = request.session.get('export_history', [])
    history.insert(0, {'name': f'Telecaller Report — {period_label}', 'type': 'csv', 'time': datetime.now().strftime('%I:%M %p')})
    request.session['export_history'] = history[:5]
    return response


@login_required
def export_counselor_csv(request):
    """Export counselor performance report to CSV."""
    if request.user.role not in ('admin', 'counselor'):
        return HttpResponseForbidden()

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    board = _compute_counselor_board(start_date, end_date, date_filter)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="counselor_performance_{period_label.replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Rank', 'Counselor', 'Leads Handled', 'Sessions', 'Follow-ups Done', 'Admissions', 'Conversion %', 'Visits', 'Score'])
    for i, row in enumerate(board, 1):
        writer.writerow([i, row['counselor'], row['total_leads'], row['sessions'], row['followups_completed'],
                         row['admissions'], f"{row['conversion_pct']:.1f}%", row['visits'], f"{row['score']:.1f}"])
    history = request.session.get('export_history', [])
    history.insert(0, {'name': f'Counselor Report — {period_label}', 'type': 'csv', 'time': datetime.now().strftime('%I:%M %p')})
    request.session['export_history'] = history[:5]
    return response


@login_required
def export_leaderboard_csv(request):
    """Export combined leaderboard to CSV."""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    tc_board = _compute_telecaller_board(start_date, end_date, date_filter)
    c_board = _compute_counselor_board(start_date, end_date, date_filter)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="leaderboard_{period_label.replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['=== TELECALLER LEADERBOARD ==='])
    writer.writerow(['Rank', 'Telecaller', 'Inquiries', 'Calls', 'Followups', 'Qualified', 'Conversion%', 'Score'])
    for i, row in enumerate(tc_board, 1):
        writer.writerow([i, row['telecaller'], row['total_inquiries'], row['calls_made'], row['followups_completed'], row['qualified_leads'], f"{row['conversion_pct']}%", f"{row['score']:.1f}"])
    writer.writerow([])
    writer.writerow(['=== COUNSELOR LEADERBOARD ==='])
    writer.writerow(['Rank', 'Counselor', 'Sessions', 'Admissions', 'Conversion%', 'Visits', 'Score'])
    for i, row in enumerate(c_board, 1):
        writer.writerow([i, row['counselor'], row['sessions'], row['admissions'], f"{row['conversion_pct']:.1f}%", row['visits'], f"{row['score']:.1f}"])
    return response


@login_required
def export_user_performance_csv(request):
    """Export user performance report to CSV."""
    if request.user.role != 'admin':
        return HttpResponseForbidden()
    return export_leaderboard_csv(request)


@login_required
def export_executive_csv(request):
    """Export executive summary to CSV."""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    start_date, end_date, date_filter, period_label = _get_date_range(request)
    today = timezone.localdate()

    inq_qs = Inquiry.objects.all()
    leads_qs = Lead.objects.all()
    calls_qs = CallLog.objects.all()
    admissions_qs = AdmissionSheet.objects.filter(admission_status='CONFIRMED')

    if start_date:
        inq_qs = inq_qs.filter(created_at__date__gte=start_date)
        leads_qs = leads_qs.filter(created_at__date__gte=start_date)
        calls_qs = calls_qs.filter(call_date__date__gte=start_date)
        admissions_qs = admissions_qs.filter(admission_date__gte=start_date)
    if end_date and date_filter not in ('all', ''):
        inq_qs = inq_qs.filter(created_at__date__lte=end_date)
        leads_qs = leads_qs.filter(created_at__date__lte=end_date)
        calls_qs = calls_qs.filter(call_date__date__lte=end_date)
        admissions_qs = admissions_qs.filter(admission_date__lte=end_date)

    total_leads = leads_qs.count()
    total_admissions = admissions_qs.count()
    conversion_rate = round(total_admissions / total_leads * 100, 2) if total_leads > 0 else 0

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="executive_summary_{period_label.replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow([f'Executive Summary — {period_label}'])
    writer.writerow([])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total Inquiries', inq_qs.count()])
    writer.writerow(['Total Leads', total_leads])
    writer.writerow(['Total Calls', calls_qs.count()])
    writer.writerow(['Confirmed Admissions', total_admissions])
    writer.writerow(['Overall Conversion Rate', f'{conversion_rate}%'])
    writer.writerow(['Overdue Follow-ups', FollowUp.objects.filter(status='Pending', followup_date__lt=today).count()])
    return response



@login_required
@admin_required
def admin_counselor_updates(request):
    # Admin tracking page showing all leads assigned to counselors
    leads = Lead.objects.filter(
        assigned_counselor__isnull=False,
        converted_at__isnull=False,
    ).select_related('inquiry', 'assigned_counselor').order_by('-counselor_status_updated_at')

    q = request.GET.get('q', '').strip()
    if q:
        leads = leads.filter(
            Q(inquiry__full_name__icontains=q) |
            Q(inquiry__mobile_number__icontains=q) |
            Q(assigned_counselor__first_name__icontains=q) |
            Q(assigned_counselor__username__icontains=q)
        )

    status = request.GET.get('status', '').strip()
    if status:
        leads = leads.filter(counselor_status=status)

    priority = request.GET.get('priority', '').strip()
    if priority:
        leads = leads.filter(priority=priority)

    date_filter = request.GET.get('date', '').strip()
    selected_date = parse_filter_date(date_filter)
    if selected_date:
        leads = filter_datetime_field_by_local_date(leads, 'counselor_status_updated_at', selected_date)
        date_filter = selected_date.isoformat()

    paginator = Paginator(leads, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    counselor_status_choices = [
        ('NEW', 'New'),
        ('CONTACTED', 'Contacted'),
        ('COUNSELING_DONE', 'Counseling Done'),
        ('FOLLOW_UP_REQUIRED', 'Follow Up Required'),
        ('INTERESTED', 'Interested'),
        ('CONVERTED', 'Converted'),
        ('ADMISSION', 'Admission'),
        ('NOT_INTERESTED', 'Not Interested'),
        ('LOST', 'Lost'),
    ]

    return render(request, 'management/admin_counselor_updates.html', {
        'page_obj': page_obj,
        'q': q,
        'status': status,
        'priority': priority,
        'date_filter': date_filter,
        'status_choices': counselor_status_choices,
        'priority_choices': Lead.PRIORITY_CHOICES,
    })


@login_required
@telecaller_counselor_admin_required
def inquiry_bulk_assign(request):
    if request.method == 'POST':
        if not is_admin_user(request.user):
            return HttpResponseForbidden("Access Denied: Only administrators can bulk assign inquiries.")

        inquiry_ids = request.POST.getlist('inquiries')
        counselor_id = request.POST.get('counselor')

        if not counselor_id:
            messages.error(request, "Please select a counselor to assign.")
            return redirect('inquiry_list')

        if not inquiry_ids:
            messages.error(request, "Please select at least one inquiry.")
            return redirect('inquiry_list')

        from django.contrib.auth import get_user_model
        User = get_user_model()
        counselor = get_object_or_404(User, pk=counselor_id, is_deleted=False, is_active=True)
        inquiries = Inquiry.objects.filter(pk__in=inquiry_ids).select_related('lead')

        assigned_count = 0
        skipped_count = 0
        with transaction.atomic():
            for inq in inquiries:
                lead = getattr(inq, 'lead', None)
                if not lead or not lead.converted_at:
                    skipped_count += 1
                    continue

                assigned_at = timezone.now()
                lead.assigned_counselor = counselor
                lead.assigned_by = request.user
                lead.assigned_at = assigned_at
                lead.counselor_assigned_at = assigned_at
                lead.save()

                LeadActivity.objects.create(
                    lead=lead,
                    activity_type='ASSIGNED',
                    description=f"Lead bulk-assigned to Counselor {counselor.username} by admin {request.user.username}.",
                    created_by=request.user
                )
                assigned_count += 1

        if assigned_count:
            messages.success(request, f"Successfully assigned {assigned_count} converted lead(s) to Counselor {counselor.username}.")
        if skipped_count:
            messages.warning(request, f"Skipped {skipped_count} inquiry/inquiries because they are not converted leads yet.")

    return redirect(request.META.get('HTTP_REFERER') or 'inquiry_list')


@login_required
@counselor_required
def counselor_telecalling_dashboard(request):
    today = timezone.localdate()
    timeout_threshold = timezone.now() - timedelta(hours=48)
    User = get_user_model()

    if is_admin_user(request.user):
        counselor_users = User.objects.filter(role='counselor', is_deleted=False, is_active=True)
    else:
        counselor_users = User.objects.filter(pk=request.user.pk)

    from django.db.models import Q

    # COUNSELOR TELECALLING DASHBOARD:
    # Shows ONLY contacts assigned by admin for counselors to do telecalling.
    # These are Inquiries whose Lead has assigned_counselor set but converted_at IS NULL.
    # Once converted, they move to Admin Queue (counselor_dashboard after admin re-assigns).

    # Active telecalling contacts: Lead exists, assigned to this counselor, NOT yet converted
    active_contacts_qs = Inquiry.objects.filter(
        lead__assigned_counselor__in=counselor_users,
        lead__converted_at__isnull=True
    ).distinct()

    # Leads converted from telecalling by this counselor (historical, for stats)
    converted_from_telecalling_qs = Lead.objects.filter(
        first_assigned_counselor__in=counselor_users,
        converted_at__isnull=False,
    )

    calls_qs = CallLog.objects.filter(created_by__in=counselor_users)
    followups_qs = FollowUp.objects.filter(created_by__in=counselor_users)

    # Statistics: based on active telecalling contacts
    assigned_inquiries = {
        'today': active_contacts_qs.filter(lead__assigned_at__date=today).count(),
        'total': active_contacts_qs.count()
    }
    total_leads = {
        'today': converted_from_telecalling_qs.filter(converted_at__date=today).count(),
        'total': converted_from_telecalling_qs.count()
    }
    assigned_leads = total_leads
    contacted_leads = {
        'today': active_contacts_qs.filter(call_status='INTERESTED', updated_at__date=today).count(),
        'total': active_contacts_qs.filter(call_status='INTERESTED').count()
    }
    interested_leads = {
        'today': active_contacts_qs.filter(call_status='INTERESTED', updated_at__date=today).count(),
        'total': active_contacts_qs.filter(call_status='INTERESTED').count()
    }
    qualified_leads = total_leads
    rejected_leads = {
        'today': active_contacts_qs.filter(call_status='NOT_INTERESTED', updated_at__date=today).count(),
        'total': active_contacts_qs.filter(call_status='NOT_INTERESTED').count()
    }
    calls_stats = {'today': calls_qs.filter(call_date__date=today).count(), 'total': calls_qs.count()}
    followups_stats = {'pending': followups_qs.filter(status='Pending').count(), 'overdue': followups_qs.filter(status='Pending', followup_date__lt=today).count()}

    # Call Outcomes (from active contacts)
    call_outcomes = {
        'accepted': active_contacts_qs.filter(call_status='ACCEPTED').count(),
        'busy': active_contacts_qs.filter(call_status='BUSY').count(),
        'ringing': active_contacts_qs.filter(call_status='NO_ANSWER').count(),
        'call_back': active_contacts_qs.filter(call_status='CALL_BACK').count(),
        'wrong_number': active_contacts_qs.filter(call_status='WRONG_NUMBER').count(),
        'interested': active_contacts_qs.filter(call_status='INTERESTED').count(),
        'not_interested': active_contacts_qs.filter(call_status='NOT_INTERESTED').count(),
        'pending_follow_up': active_contacts_qs.filter(call_status='PENDING_FOLLOW_UP', updated_at__gte=timeout_threshold).count(),
        'overdue_follow_up': active_contacts_qs.filter(call_status='PENDING_FOLLOW_UP', updated_at__lt=timeout_threshold).count(),
    }

    # Tables
    recent_inquiries = active_contacts_qs.select_related('lead', 'lead__assigned_counselor').order_by('-lead__assigned_at')[:5]
    pending_inquiry_followups = active_contacts_qs.filter(call_status='PENDING_FOLLOW_UP', updated_at__gte=timeout_threshold).order_by('-updated_at')[:5]
    overdue_inquiry_followups = active_contacts_qs.filter(call_status='PENDING_FOLLOW_UP', updated_at__lt=timeout_threshold).order_by('updated_at')[:5]
    recent_leads = converted_from_telecalling_qs.select_related('inquiry', 'assigned_counselor').order_by('-converted_at')[:5]
    recent_activities = LeadActivity.objects.filter(
        lead__in=converted_from_telecalling_qs
    ).select_related('lead__inquiry', 'created_by').order_by('-created_at')[:5]
    today_followups_list = followups_qs.filter(status='Pending', followup_date=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]
    overdue_followups_list = followups_qs.filter(status='Pending', followup_date__lt=today).select_related('lead__inquiry', 'created_by').order_by('followup_date')[:5]

    context = {
        'assigned_inquiries': assigned_inquiries,
        'converted_to_lead': total_leads,
        'total_leads': total_leads,
        'assigned_leads': assigned_leads,
        'contacted_leads': contacted_leads,
        'interested_leads': interested_leads,
        'qualified_leads': qualified_leads,
        'rejected_leads': rejected_leads,
        'today_calls': calls_stats,
        'pending_followups': followups_stats['pending'],
        'overdue_followups': followups_stats['overdue'],
        'call_outcomes': call_outcomes,
        'recent_inquiries': recent_inquiries,
        'pending_inquiry_followups': pending_inquiry_followups,
        'overdue_inquiry_followups': overdue_inquiry_followups,
        'recent_leads': recent_leads,
        'recent_activities': recent_activities,
        'today_followups_list': today_followups_list,
        'overdue_followups_list': overdue_followups_list,
        'active_contacts_count': active_contacts_qs.count(),
    }
    return render(request, 'management/counselor_telecalling_dashboard.html', context)

